from __future__ import annotations
import pandas as pd
from mongoengine import Q
from omnisight.models import OrderItems,Order,Marketplace,Product,CityDetails,user,notes_data,chooseMatrix,Fee,Refund,Brand,inventry_log,productPriceChange
from mongoengine.queryset.visitor import Q
from concurrent.futures import ThreadPoolExecutor, as_completed
from dateutil.relativedelta import relativedelta
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime,timedelta
from bson.son import SON
from bson import ObjectId
import numpy as np
import json
import time
import asyncio
from collections import defaultdict
from bson import ObjectId
from django.http import JsonResponse
from datetime import datetime
from django.core.cache import cache 
from django.http import JsonResponse
from django.http import HttpResponse
import openpyxl
import csv
from datetime import datetime
import math
import pytz 
import threading
import re
from bson import ObjectId
from django.http import JsonResponse
from rest_framework.parsers import JSONParser
from collections import OrderedDict, defaultdict
from io import StringIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import io
from pytz import timezone
from bson import ObjectId
from calendar import monthrange
from ecommerce_tool.settings import MARKETPLACE_ID,SELLER_ID
from django.db.models import Sum, Q
from omnisight.operations.helium_utils import calculate_metricss, get_date_range, grossRevenue, get_previous_periods, refundOrder,AnnualizedRevenueAPIView,getOrdersListBasedonProductId, getproductIdListBasedonbrand, getdaywiseproductssold, pageViewsandSessionCount, getproductIdListBasedonManufacture,totalRevenueCalculation,get_graph_data, totalRevenueCalculationForProduct, get_top_movers, convertLocalTimeToUTC, convertdateTotimezone
from ecommerce_tool.crud import DatabaseModel
from omnisight.operations.common_utils import calculate_listing_score
import threading
from concurrent.futures import ThreadPoolExecutor
import math
from omnisight.models import *
from django.utils import timezone
from rest_framework.parsers import JSONParser
from datetime import datetime
from datetime import datetime, timedelta
from rest_framework.parsers import JSONParser
from django.http import JsonResponse
import logging
logger = logging.getLogger(__name__)
def sanitize_data(data):
    """Recursively sanitize data to ensure all float values are JSON compliant."""
    if isinstance(data, dict):
        return {key: sanitize_data(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [sanitize_data(item) for item in data]
    elif isinstance(data, float):
        if math.isnan(data) or data == float('inf') or data == float('-inf'):
            return 0  
    return data



def calculate_margin(records):
    gross=sum(r['gross_revenue'] for r in records)
    net=sum(r['net_profit'] for r in records)
    return round((net/gross)*100,2) if gross else 0

@csrf_exempt
def get_metrics_by_date_range(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    target_date_str = json_request.get('target_date')
    brand_id = json_request.get('brand_id', None)
    product_id = json_request.get('product_id', None)
    manufacturer_name = json_request.get('manufacturer_name', [])
    fulfillment_channel = json_request.get('fulfillment_channel', None)
    timezone_str="US/Pacific"
    preset=json_request.get('preset','Today')
    start_date_str=json_request.get("start_date",None)
    end_date_str=json_request.get('end_date',None)
    if start_date_str and end_date_str:
        start_date_dt=datetime.strptime(start_date_str,"%d/%m/%Y")
        end_date_dt=datetime.strptime(end_date_str,"%d/%m/%Y").replace(hour=23,minute=59,second=59)
    else:
        start_date_dt,end_date_dt=get_date_range(preset,time_zone_str=timezone_str)
    target_date = datetime.strptime(target_date_str, "%d/%m/%Y").date()
    local_tz = pytz.timezone(timezone_str)
    current_time = datetime.now(local_tz).replace(year=target_date.year, month=target_date.month, day=target_date.day)
    target_date = current_time.replace(hour=0, minute=0, second=0, microsecond=0)
    previous_date = target_date - timedelta(days=1)
    eight_days_ago = target_date - timedelta(days=8)
    date_filters = {
        "targeted": {
            "start": start_date_dt,
            "end": end_date_dt
        },
        "previous": {
            "start": start_date_dt-timedelta(days=1),
            "end": end_date_dt-timedelta(days=1)
        }
    }
    if start_date_str and end_date_str:
        graph_days_filter = {}
        current_day = start_date_dt
        while current_day <= end_date_dt:
            day_key = current_day.strftime("%B %d, %Y").lower()
            graph_days_filter[day_key] = {
            "start": datetime(current_day.year, current_day.month, current_day.day),
            "end": datetime(current_day.year, current_day.month, current_day.day, 23, 59, 59)
            }
            current_day += timedelta(days=1)
    else:
        eight_days_ago = target_date - timedelta(days=8)
        graph_days_filter = {}
        for i in range(1, 9):
            day = eight_days_ago + timedelta(days=i)
            day_key = day.strftime("%B %d, %Y").lower()
            graph_days_filter[day_key] = {
            "start": datetime(day.year, day.month, day.day),
            "end": datetime(day.year, day.month, day.day, 23, 59, 59)
        }
    metrics = {}
    graph_data = {}
    def process_date_range(key, date_range, results):
        gross_revenue = 0
        result = grossRevenue(date_range["start"], date_range["end"], marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel, timezone_str)
        if result != []:
            for ins in result:
                gross_revenue += ins['order_total']
        results[key] = {
            "gross_revenue": round(gross_revenue, 2),
        }
    from concurrent.futures import ThreadPoolExecutor
    results = {}
    with ThreadPoolExecutor(max_workers=min(len(graph_days_filter), 10)) as executor:
        futures = {executor.submit(process_date_range, key, date_range, results): key 
                  for key, date_range in graph_days_filter.items()}
        for future in futures:
            future.result()  
    graph_data = {key: results[key] for key in graph_days_filter.keys()}
    metrics["graph_data"] = graph_data
    all_order_item_ids = set()
    all_raw_results = {}
    for key, date_range in date_filters.items():
        raw_result = grossRevenue(date_range["start"], date_range["end"], marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel, timezone_str)
        result=[
            r for r in raw_result
            if r.get('order_status') not in ['Cancelled','Canceled'] and r.get('order_total')>0
        ]
        all_raw_results[key] = result
        for ins in result:
            all_order_item_ids.update(ins['order_items'])
    bulk_pipeline = [
        {
            "$match": {
                "_id": {"$in": list(all_order_item_ids)}
            }
        },
        {
            "$lookup": {
                "from": "product",
                "localField": "ProductDetails.product_id",
                "foreignField": "_id",
                "as": "product_ins"
            }
        },
        {
        "$unwind": {
            "path": "$product_ins",
            "preserveNullAndEmptyArrays": True
        }
        },
        {
            "$project": {
                "_id": 1,
                "price": {"$ifNull": ["$Pricing.ItemPrice.Amount", 0]},
                "cogs": {"$ifNull": ["$product_ins.cogs", 0.0]},
                "tax_price": {"$ifNull": ["$Pricing.ItemTax.Amount", 0]},
                "total_cogs": {"$ifNull": ["$product_ins.total_cogs", 0]},
                "w_total_cogs": {"$ifNull": ["$product_ins.w_total_cogs", 0]},
                "vendor_funding": {"$ifNull": ["$product_ins.vendor_funding", 0]},
            }
        }
    ]
    order_items_lookup = {}
    bulk_results = list(OrderItems.objects.aggregate(*bulk_pipeline))
    for item in bulk_results:
        order_items_lookup[item['_id']] = item
    for key, date_range in date_filters.items():
        gross_revenue = 0
        total_cogs = 0
        refund = 0
        margin = 0
        net_profit = 0
        total_units = 0
        total_orders = 0
        tax_price = 0
        temp_other_price = 0
        vendor_funding = 0
        result = all_raw_results[key]
        refund_ins = refundOrder(date_range["start"], date_range["end"], marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel)
        if refund_ins != []:
            for ins in refund_ins:
                refund += len(ins['order_items'])
        total_orders = len(result)
        if result != []:
            for ins in result:
                gross_revenue += ins['order_total']
                total_units += ins['items_order_quantity']
                for j in ins['order_items']:
                    item_result = order_items_lookup.get(j)
                    if item_result:
                        tax_price += item_result['tax_price']
                        temp_other_price += item_result['price']
                        if ins['marketplace_name'] == "Amazon":
                            total_cogs += item_result['total_cogs']
                        else:
                            total_cogs += item_result['w_total_cogs']
                        vendor_funding += item_result['vendor_funding']
            net_profit = (temp_other_price - total_cogs) + vendor_funding
            margin = (net_profit / gross_revenue) * 100 if gross_revenue != 0 else 0
        metrics[key] = {
            "gross_revenue": round(gross_revenue, 2),
            "total_tax":round(tax_price,2),
            "total_cogs": round(total_cogs, 2),
            "refund": round(refund, 2),
            "margin": round(margin, 2),
            "net_profit": round(net_profit, 2),
            "total_orders": round(total_orders, 2),
            "total_units": round(total_units, 2)
        }
    difference = {
        "gross_revenue": round(metrics["targeted"]["gross_revenue"] - metrics["previous"]["gross_revenue"], 2),
        "total_cogs": round(metrics["targeted"]["total_cogs"] - metrics["previous"]["total_cogs"], 2),
        "refund": round(metrics["targeted"]["refund"] - metrics["previous"]["refund"], 2),
        "margin": round(metrics["targeted"]["margin"] - metrics["previous"]["margin"], 2),
        "total_tax":round(metrics['targeted']['total_tax']-metrics['previous']['total_tax'],2),
        "net_profit": round(metrics["targeted"]["net_profit"] - metrics["previous"]["net_profit"], 2),
        "total_orders": round(metrics["targeted"]["total_orders"] - metrics["previous"]["total_orders"], 2),
        "total_units": round(metrics["targeted"]["total_units"] - metrics["previous"]["total_units"], 2),
    }
    name = "Today Snapshot"
    item_pipeline = [
        {"$match": {"name": name}}
    ]
    item_result = list(chooseMatrix.objects.aggregate(*item_pipeline))
    if item_result:
        item_result = item_result[0]
        if item_result['select_all']:
            pass
        if item_result['gross_revenue'] == False:
            del metrics['targeted']["gross_revenue"]
            del metrics['previous']["gross_revenue"]
        if item_result['units_sold'] == False:
            del metrics['targeted']["total_units"]
            del metrics['previous']["total_units"]
        if item_result['total_cogs'] == False:
            del metrics['targeted']["total_cogs"]
            del metrics['previous']["total_cogs"]
        if item_result['orders'] == False:
            del metrics['targeted']["total_orders"]
            del metrics['previous']["total_orders"]
        if item_result['refund_quantity'] == False:
            del metrics['targeted']["refund"]
            del metrics['previous']["refund"]
        if item_result['profit_margin'] == False:
            del metrics['targeted']["margin"]
            del metrics['previous']["margin"]
    metrics["difference"] = difference
    metrics = sanitize_data(metrics)
    return metrics

@csrf_exempt
def LatestOrdersTodayAPIView(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    product_id = json_request.get('product_id', [])
    brand_id = json_request.get('brand_id', [])
    manufacturer_name = json_request.get('manufacturer_name', [])
    fulfillment_channel = json_request.get('fulfillment_channel', None)
    pacific = pytz.timezone("US/Pacific")
    utc = pytz.UTC
    now_pacific = datetime.now(pacific)
    start_of_day_pacific = now_pacific.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_day_pacific = now_pacific.replace(hour=23, minute=59, second=59, microsecond=999999)
    start_of_day_utc = start_of_day_pacific.astimezone(utc)
    end_of_day_utc = end_of_day_pacific.astimezone(utc)
    match = dict()
    match['order_date'] = {"$gte": start_of_day_utc, "$lte": end_of_day_utc}
    match['order_status'] = {"$in": ['Shipped', 'Delivered','Acknowledged','Pending','Unshipped','PartiallyShipped']}
    if fulfillment_channel:
        match['fulfillment_channel'] = fulfillment_channel
    if marketplace_id != None and marketplace_id != "" and marketplace_id != "all" and marketplace_id != "custom":
        match['marketplace_id'] = ObjectId(marketplace_id)
    if manufacturer_name != None and manufacturer_name != "" and manufacturer_name != []:
        ids = getproductIdListBasedonManufacture(manufacturer_name, start_of_day_utc, end_of_day_utc)
        match["_id"] = {"$in": ids}
    elif product_id != None and product_id != "" and product_id != []:
        product_id = [ObjectId(pid) for pid in product_id]
        ids = getOrdersListBasedonProductId(product_id, start_of_day_utc, end_of_day_utc)
        match["_id"] = {"$in": ids}
    elif brand_id != None and brand_id != "" and brand_id != []:
        brand_id = [ObjectId(bid) for bid in brand_id]
        ids = getproductIdListBasedonbrand(brand_id, start_of_day_utc, end_of_day_utc)
        match["_id"] = {"$in": ids}
    pipeline = [
        {
            "$match": match
        },
        {
            "$sort": {"order_date": -1}  
        },
        {
            "$project": {
                "_id" : 1,
                "order_date": 1,
                "order_items": 1
            }
        }
    ]
    qs = list(Order.objects.aggregate(*pipeline))
    chart = OrderedDict()
    bucket = start_of_day_pacific.replace(minute=0, second=0, microsecond=0)
    for _ in range(24):  
        key = bucket.strftime("%Y-%m-%d %H:00:00")
        chart[key] = {"ordersCount": 0, "unitsCount": 0}
        bucket += timedelta(hours=1)
    orders_out = []
    for order in qs:
        order_utc_time = order.get('order_date')
        if isinstance(order_utc_time, datetime):
            if order_utc_time.tzinfo is None:
                order_utc_time = utc.localize(order_utc_time)
            order_pacific_time = order_utc_time.astimezone(pacific)
            bk = order_pacific_time.replace(minute=0, second=0, microsecond=0).strftime("%Y-%m-%d %H:00:00")
        else:
            raise TypeError("'order_date' must be a datetime object")
        if bk in chart:
            chart[bk]["ordersCount"] += 1
            try:
                item_pipeline = [
                    {"$match": {"_id": {"$in": order['order_items']}}},
                    {
                        "$lookup": {
                            "from": "product",
                            "localField": "ProductDetails.product_id",
                            "foreignField": "_id",
                            "as": "product_ins"
                        }
                    },
                    {"$unwind": {"path": "$product_ins", "preserveNullAndEmptyArrays": True}},
                    {
                        "$group": {
                            "_id": {
                                "id" : "$product_ins._id",
                                "sku": {"$ifNull": ["$product_ins.sku", ""]},
                                "asin": {"$ifNull": ["$product_ins.product_id", ""]},
                                "title": "$product_ins.product_title",
                                "imageUrl": "$product_ins.image_url",
                                "purchaseDate": order_pacific_time  
                            },
                            "quantityOrdered": {"$sum": "$ProductDetails.QuantityOrdered"},
                            "unitPrice": {"$first": "$Pricing.ItemPrice.Amount"},
                            "Platform" : {"$first": "$Platform"},
                        }
                    },
                    {
                        "$project": {
                            "_id": 0,
                            "id" : "$_id.id",
                            "sku": "$_id.sku",
                            "asin": "$_id.asin",
                            "title": "$_id.title",
                            "imageUrl": "$_id.imageUrl",
                            "quantityOrdered": "$quantityOrdered",
                            "unitPrice": "$unitPrice",
                            "Platform" : "$Platform",
                            "purchaseDate": "$_id.purchaseDate"
                        }
                    }
                ]
                item_results = list(OrderItems.objects.aggregate(*item_pipeline))
                for item in item_results:
                    try:
                        if item["Platform"] == "Walmart":
                            price = item["unitPrice"] * item['quantityOrdered']
                        elif item["Platform"] == "Amazon":
                            price = item["unitPrice"]
                        else:
                            price = item["unitPrice"]
                    except:
                        price = 0
                    orders_out.append({
                        "id" : str(item["id"]),
                        "sellerSku": item["sku"],
                        "asin": item["asin"],
                        "title": item["title"],
                        "quantityOrdered": item["quantityOrdered"],
                        "imageUrl": item["imageUrl"],
                        "price": price,
                        "purchaseDate": order_pacific_time.strftime("%Y-%m-%d %H:%M:%S"),  
                        "purchaseDatetime": order_pacific_time  
                    })
                    chart[bk]["unitsCount"] += item["quantityOrdered"]
            except Exception as e:
                continue
    orders_out.sort(key=lambda o: o["purchaseDatetime"], reverse=True)
    for order in orders_out:
        order.pop("purchaseDatetime", None)
    chart_list = [{"hour": hour, **data} for hour, data in chart.items()]
    data = {
        "orders": orders_out,
        "hourly_order_count": chart_list
    }
    return data
@csrf_exempt
def RevenueWidgetAPIView(request):
    from django.utils import timezone
    json_request = JSONParser().parse(request)
    preset = json_request.get("preset", "Today")
    compare_startdate = json_request.get("compare_startdate")
    compare_enddate = json_request.get("compare_enddate")
    marketplace_id = json_request.get("marketplace_id", None)
    product_id = json_request.get("product_id", None)
    brand_id = json_request.get("brand_id", None)
    manufacturer_name = json_request.get("manufacturer_name", None)
    fulfillment_channel = json_request.get("fulfillment_channel", None)
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    timezone_str = json_request.get('timezone', 'US/Pacific')
    if start_date != None and start_date != "":
        start_date, end_date = convertdateTotimezone(start_date,end_date,timezone_str)
    else:
        start_date, end_date = get_date_range(preset,timezone_str)
    comapre_past = get_previous_periods(start_date, end_date)
    def fetch_total():
        return totalRevenueCalculation(start_date, end_date, marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,timezone_str)
    def fetch_graph_data():
        return get_graph_data(start_date, end_date, preset, marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel)
    def fetch_compare_total():
        return totalRevenueCalculation(compare_startdate, compare_enddate, marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,timezone_str)
    def fetch_compare_graph_data():
        return get_graph_data(compare_startdate, compare_enddate, initial, marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel)
    executor = ThreadPoolExecutor(max_workers=4)
    try:
        future_total = executor.submit(fetch_total)
        future_graph_data = executor.submit(fetch_graph_data)
        compare_total = None
        compare_graph = None
        if compare_startdate != None and compare_startdate != "":
            compare_startdate = datetime.strptime(compare_startdate, "%Y-%m-%d").replace(hour=0, minute=0, second=0, microsecond=0)
            compare_enddate = datetime.strptime(compare_enddate, "%Y-%m-%d").replace(hour=23, minute=59, second=59, microsecond=0)
            initial = "Today" if compare_startdate.date() == compare_enddate.date() else None
            future_compare_total = executor.submit(fetch_compare_total)
            future_compare_graph_data = executor.submit(fetch_compare_graph_data)
    finally:
        executor.shutdown(wait=True)
    total = future_total.result()
    graph_data = future_graph_data.result()
    if compare_startdate != None and compare_startdate != "":
        compare_total = future_compare_total.result()
        compare_graph = future_compare_graph_data.result()
    data = {
        "total": total,
        "graph": graph_data,
        "comapre_past": comapre_past
    }
    if compare_total:
        difference = {
            "gross_revenue": round(((total["gross_revenue"] - compare_total["gross_revenue"]) / compare_total["gross_revenue"] * 100) if compare_total["gross_revenue"] else 0, 2),
            "net_profit": round(((total["net_profit"] - compare_total["net_profit"]) / compare_total["net_profit"] * 100) if compare_total["net_profit"] else 0, 2),
            "profit_margin": round(((total["profit_margin"] - compare_total["profit_margin"]) / compare_total["profit_margin"] * 100) if compare_total["profit_margin"] else 0, 2),
            "orders": round(((total["orders"] - compare_total["orders"]) / compare_total["orders"] * 100) if compare_total["orders"] else 0, 2),
            "units_sold": round(((total["units_sold"] - compare_total["units_sold"]) / compare_total["units_sold"] * 100) if compare_total["units_sold"] else 0, 2),
            "refund_amount": round(((total["refund_amount"] - compare_total["refund_amount"]) / compare_total["refund_amount"] * 100) if compare_total["refund_amount"] else 0, 2),
            "refund_quantity": round(((total["refund_quantity"] - compare_total["refund_quantity"]) / compare_total["refund_quantity"] * 100) if compare_total["refund_quantity"] else 0, 2),
        }
        data['compare_total'] = difference
        data['previous_total'] = compare_total
        data['compare_graph'] = compare_graph
    name = "Revenue"
    item_pipeline = [
        {"$match": {"name": name}}
    ]
    item_result = list(chooseMatrix.objects.aggregate(*item_pipeline))
    if item_result:
        item_result = item_result[0]
        if item_result['select_all']:
            pass
        if item_result['gross_revenue'] == False:
            del data['total']["gross_revenue"]
        if item_result['units_sold'] == False:
            del data['total']["units_sold"]
        if item_result['refund_quantity'] == False:
            del data['total']["refund_quantity"]
        if item_result['refund_amount'] == False:
            del data['total']["refund_amount"]
        if item_result['net_profit'] == False:
            del data['total']["net_profit"]
        if item_result['profit_margin'] == False:
            del data['total']["profit_margin"]
        if item_result['orders'] == False:
            del data['total']["orders"]
    return data
@csrf_exempt
def updatedRevenueWidgetAPIView(request):
    from django.utils import timezone
    from concurrent.futures import ThreadPoolExecutor
    from datetime import datetime

    json_request = JSONParser().parse(request)
    preset = json_request.get("preset", "Today")
    compare_startdate = json_request.get("compare_startdate")
    compare_enddate = json_request.get("compare_enddate")
    marketplace_id = json_request.get("marketplace_id", None)
    product_id = json_request.get("product_id", None)
    brand_id = json_request.get("brand_id", None)
    manufacturer_name = json_request.get("manufacturer_name", None)
    fulfillment_channel = json_request.get("fulfillment_channel", None)

    timezone_str = "US/Pacific"
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)

    if start_date not in [None, ""]:
        start_date, end_date = convertdateTotimezone(start_date, end_date, timezone_str)
    else:
        start_date, end_date = get_date_range(preset, timezone_str)

    compare_enabled = compare_startdate not in [None, ""]
    if compare_enabled:
        compare_startdate = datetime.strptime(compare_startdate, "%Y-%m-%d").replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        compare_enddate = datetime.strptime(compare_enddate, "%Y-%m-%d").replace(
            hour=23, minute=59, second=59, microsecond=0
        )

    comapre_past = get_previous_periods(start_date, end_date)

    def fetch_total():
        return totalRevenueCalculation(
            start_date, end_date, marketplace_id, brand_id,
            product_id, manufacturer_name, fulfillment_channel, timezone_str
        )

    def fetch_graph_data():
        return get_graph_data(
            start_date, end_date, preset, marketplace_id, brand_id,
            product_id, manufacturer_name, fulfillment_channel, timezone_str
        )

    def fetch_compare_total():
        return totalRevenueCalculation(
            compare_startdate, compare_enddate, marketplace_id,
            brand_id, product_id, manufacturer_name, fulfillment_channel, timezone_str
        )

    def fetch_compare_graph_data():
        initial = "Today" if compare_startdate.date() == compare_enddate.date() else None
        return get_graph_data(
            compare_startdate, compare_enddate, initial, marketplace_id,
            brand_id, product_id, manufacturer_name, fulfillment_channel, timezone_str
        )

    with ThreadPoolExecutor(max_workers=4) as executor:
        future_total = executor.submit(fetch_total)
        future_graph_data = executor.submit(fetch_graph_data)
        future_compare_total = executor.submit(fetch_compare_total) if compare_enabled else None
        future_compare_graph_data = executor.submit(fetch_compare_graph_data) if compare_enabled else None

        total = future_total.result()
        graph_data = future_graph_data.result()
        compare_total = future_compare_total.result() if compare_enabled else None
        compare_graph = future_compare_graph_data.result() if compare_enabled else None

    updated_graph = {}

    if compare_enabled:
        for index, (key, metrics) in enumerate(graph_data.items()):
            compare_metrics = list(compare_graph.values())[index] if index < len(compare_graph) else {}
            updated_graph[key] = {
                "current_date": key,
                "gross_revenue": metrics.get("gross_revenue", 0),
                "net_profit": metrics.get("net_profit", 0),
                "profit_margin": metrics.get("profit_margin", 0),
                "orders": metrics.get("orders", 0),
                "units_sold": metrics.get("units_sold", 0),
                "refund_amount": metrics.get("refund_amount", 0),
                "refund_quantity": metrics.get("refund_quantity", 0),
                "compare_gross_revenue": compare_metrics.get("gross_revenue", 0),
                "compare_net_profit": compare_metrics.get("net_profit", 0),
                "compare_profit_margin": compare_metrics.get("profit_margin", 0),
                "compare_orders": compare_metrics.get("orders", 0),
                "compare_units_sold": compare_metrics.get("units_sold", 0),
                "compare_refund_amount": compare_metrics.get("refund_amount", 0),
                "compare_refund_quantity": compare_metrics.get("refund_quantity", 0),
                "compare_date": list(compare_graph.keys())[index] if index < len(compare_graph) else None,
            }
    else:
        for key, metrics in graph_data.items():
            updated_graph[key] = {
                "current_date": key,
                "gross_revenue": metrics.get("gross_revenue", 0),
                "net_profit": metrics.get("net_profit", 0),
                "profit_margin": metrics.get("profit_margin", 0),
                "orders": metrics.get("orders", 0),
                "units_sold": metrics.get("units_sold", 0),
                "refund_amount": metrics.get("refund_amount", 0),
                "refund_quantity": metrics.get("refund_quantity", 0),
            }

    data = {
        "total": total,
        "graph": updated_graph,
        "comapre_past": comapre_past,
    }

    if compare_enabled:
        difference = {
            "gross_revenue": round(((total["gross_revenue"] - compare_total["gross_revenue"]) / compare_total["gross_revenue"] * 100) if compare_total["gross_revenue"] else 0, 2),
            "net_profit": round(((total["net_profit"] - compare_total["net_profit"]) / compare_total["net_profit"] * 100) if compare_total["net_profit"] else 0, 2),
            "profit_margin": round(((total["profit_margin"] - compare_total["profit_margin"]) / compare_total["profit_margin"] * 100) if compare_total["profit_margin"] else 0, 2),
            "orders": round(((total["orders"] - compare_total["orders"]) / compare_total["orders"] * 100) if compare_total["orders"] else 0, 2),
            "units_sold": round(((total["units_sold"] - compare_total["units_sold"]) / compare_total["units_sold"] * 100) if compare_total["units_sold"] else 0, 2),
            "refund_amount": round(((total["refund_amount"] - compare_total["refund_amount"]) / compare_total["refund_amount"] * 100) if compare_total["refund_amount"] else 0, 2),
            "refund_quantity": round(((total["refund_quantity"] - compare_total["refund_quantity"]) / compare_total["refund_quantity"] * 100) if compare_total["refund_quantity"] else 0, 2),
        }
        data['compare_total'] = difference

    name = "Revenue"
    item_pipeline = [{"$match": {"name": name}}]
    item_result = list(chooseMatrix.objects.aggregate(*item_pipeline))
    if item_result:
        item_result = item_result[0]
        if not item_result['select_all']:
            for field in ['gross_revenue', 'units_sold', 'refund_quantity',
                          'refund_amount', 'net_profit', 'profit_margin', 'orders']:
                if not item_result.get(field, True):
                    data['total'].pop(field, None)

    return data

import pytz
@csrf_exempt
def get_top_products(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', None)
    product_id = json_request.get('product_id', None)
    metric = json_request.get("sortBy", "units_sold")  
    preset = json_request.get("preset", "Today")  
    start_date_str = json_request.get("start_date", None)
    end_date_str = json_request.get("end_date", None)
    timezone_str = "US/Pacific"
    if start_date_str and end_date_str:
        local_tz = pytz.timezone(timezone_str)
        naive_from_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        naive_to_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        localized_from_date = local_tz.localize(naive_from_date)
        localized_to_date = local_tz.localize(naive_to_date).replace(hour=23, minute=59, second=59)
        start_date = localized_from_date.astimezone(pytz.UTC)
        end_date = localized_to_date.astimezone(pytz.UTC)
    else:
        start_date, end_date = get_date_range(preset, timezone_str)  
    duration_hours = (end_date - start_date).total_seconds() / 3600
    if duration_hours <= 24:
        chart_date_format = "%Y-%m-%d %H:00:00+00:00"
    else:
        chart_date_format = "%Y-%m-%d 00:00:00+00:00"
    sort_field = {
        "units_sold": "total_units",
        "price": "total_price",
        "refund": "refund_qty"
    }.get(metric, "total_units")
    chart_value_field = {
        "units_sold": "$order_items_ins.ProductDetails.QuantityOrdered",
        "price": {
            "$multiply": [
                "$order_items_ins.Pricing.ItemPrice.Amount",
                "$order_items_ins.ProductDetails.QuantityOrdered"
            ]
        },
        "refund": "$order_items_ins.ProductDetails.QuantityShipped"
    }.get(metric, "$order_items_ins.ProductDetails.QuantityOrdered")
    match = dict()
    match['order_date'] = {"$gte": start_date, "$lt": end_date}
    match['order_status'] = {"$in": ['Shipped', 'Delivered', 'Acknowledged', 'Pending', 'Unshipped', 'PartiallyShipped']}
    if marketplace_id and marketplace_id not in ["all", "custom"]:
        match['marketplace_id'] = ObjectId(marketplace_id)
    if metric == "refund":
        match['order_status'] = "Refunded"
    brand_ids_for_match = []
    if brand_id:
        if isinstance(brand_id, str):
            brand_ids_for_match = [ObjectId(brand_id)]
        elif isinstance(brand_id, list):
            brand_ids_for_match = [ObjectId(bid) for bid in brand_id]
    product_ids_for_match = []
    if product_id:
        if isinstance(product_id, str):
            product_ids_for_match = [ObjectId(product_id)]
        elif isinstance(product_id, list):
            product_ids_for_match = [ObjectId(pid) for pid in product_id]
    if product_ids_for_match:
        ids_from_products = getOrdersListBasedonProductId(product_ids_for_match, start_date, end_date)
        if ids_from_products:
            match["_id"] = {"$in": ids_from_products}
        else:
            return {"results": {"items": []}}
    pipeline = [
        {"$match": match},
        {"$lookup": {
            "from": "order_items",
            "localField": "order_items",
            "foreignField": "_id",
            "as": "order_items_ins"
        }},
        {"$unwind": {
            "path": "$order_items_ins",
            "preserveNullAndEmptyArrays": True
        }},
        {"$lookup": {
            "from": "product",
            "localField": "order_items_ins.ProductDetails.product_id",
            "foreignField": "_id",
            "as": "product_ins"
        }},
        {"$unwind": {
            "path": "$product_ins",
            "preserveNullAndEmptyArrays": True
        }},
    ]
    if brand_ids_for_match:
        pipeline.append({
            "$match": {
                "product_ins.brand_id": {"$in": brand_ids_for_match}
            }
        })
    pipeline.extend([
    {"$addFields": {
        "chart_key_raw": "$order_date",
        "chart_value": chart_value_field
    }},
    {"$group": {
        "_id": {
            "productId": "$product_ins._id",
            "timeBucket": {
                "$dateToString": {
                    "format": chart_date_format,
                    "date": "$chart_key_raw",
                }
            }
        },
        "productTitle": {"$first": "$product_ins.product_title"},
        "asin": {"$first": "$product_ins.product_id"},
        "sellerSku": {"$first": "$product_ins.sku"},
        "imageUrl": {"$first": "$product_ins.image_url"},
        "total_units_sum": {"$sum": "$order_items_ins.ProductDetails.QuantityOrdered"},
        "total_price_sum": {
            "$sum": {
                "$multiply": [
                    "$order_items_ins.Pricing.ItemPrice.Amount",
                    "$order_items_ins.ProductDetails.QuantityOrdered"
                ]
            }
        },
        "refund_qty_sum": {"$sum": "$order_items_ins.ProductDetails.QuantityShipped"},
        "hourly_or_daily_sale": {"$sum": "$chart_value"}
    }},
    {"$group": {
        "_id": "$_id.productId",
        "product": {
            "$first": {
                "title": "$productTitle",
                "asin": "$asin",
                "sellerSku": "$sellerSku",
                "imageUrl": "$imageUrl"
            }
        },
        "chart": {
            "$push": {
                "k": "$_id.timeBucket",
                "v": "$hourly_or_daily_sale"
            }
        },
        "total_units": {"$sum": "$total_units_sum"},
        "total_price": {"$sum": "$total_price_sum"},
        "refund_qty": {"$sum": "$refund_qty_sum"}
    }},
    {"$project": {
        "_id": 1,
        "product": 1,
        "chart": {
            "$arrayToObject": {
                "$filter": {
                    "input": "$chart",
                    "as": "item",
                    "cond": {
                        "$and": [
                            {"$ne": ["$$item.k", None]},
                            {"$ne": ["$$item.v", None]},
                            {"$eq": [{"$type": "$$item.k"}, "string"]}
                        ]
                    }
                }
            }
        },
        "total_units": 1,
        "total_price": 1,
        "refund_qty": 1
    }},
    {"$sort": SON([(sort_field, -1)])},
    {"$limit": 11}
    ])
    result = list(Order.objects.aggregate(pipeline))
    formatted_results = []
    for item in result:
        product_info = item.get("product") or {}
        chart = item.get("chart", {})
        chart = {str(k): float(v) for k, v in chart.items() if k and v is not None}
        product_dict = {}
        _id = item.get("_id")
        if _id:
            product_dict["id"] = str(_id)
        if product_info.get("title"):
            product_dict["product"] = product_info["title"]
        if product_info.get("asin"):
            product_dict["asin"] = product_info["asin"]
        if product_info.get("sellerSku"):
            product_dict["sku"] = product_info["sellerSku"]
        if product_info.get("imageUrl"):
            product_dict["product_image"] = product_info["imageUrl"]
        if item.get("total_units") is not None:
            product_dict["total_units"] = item["total_units"]
        if item.get("total_price"):
            product_dict["total_price"] = item["total_price"]
        if item.get("refund_qty"):
            product_dict["refund_qty"] = item["refund_qty"]
        if chart:
            product_dict["chart"] = chart
        title=product_dict.get('product',"").strip()
        if title:
            formatted_results.append(product_dict)
    data = {"results": {"items": formatted_results}}
    # After you create the formatted_results list but before returning the data
    for item in formatted_results:
        if "chart" in item:
            if duration_hours <= 24:
                end_date_str = end_date.strftime("%Y-%m-%d %H:00:00+00:00")
            else :
                end_date_str = end_date.strftime("%Y-%m-%d 00:00:00+00:00")
            item["chart"] = {k: v for k, v in item["chart"].items() if k < end_date_str}
    return data
def getPreviousDateRange(start_date, end_date):
    duration = end_date - start_date
    previous_start_date = start_date - duration - timedelta(days=1)
    previous_end_date = start_date - timedelta(days=1)
    return previous_start_date.strftime("%Y-%m-%d"), previous_end_date.strftime("%Y-%m-%d")

@csrf_exempt
def get_products_with_pagination(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', None)
    product_id = json_request.get('product_id', None)
    manufacturer_name = json_request.get('manufacturer_name', [])
    page = int(json_request.get("page", 1))
    page_size = int(json_request.get("page_size", 10))
    preset = json_request.get("preset", "Today")
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    parent = json_request.get('parent', True)
    sort_by = json_request.get('sort_by')
    sort_by_value = json_request.get('sort_by_value', 1)
    parent_search = json_request.get('parent_search')
    sku_search = json_request.get('sku_search')
    search_query = json_request.get('search_query')
    timezone_str = 'US/Pacific'
    if start_date and start_date != "":
        start_date, end_date = convertdateTotimezone(start_date, end_date, timezone_str)
    else:
        start_date, end_date = get_date_range(preset, timezone_str)
    today_start_date, today_end_date = get_date_range("Today", timezone_str)
    if timezone_str != 'UTC':
        today_start_date, today_end_date = convertLocalTimeToUTC(today_start_date, today_end_date, timezone_str)
        start_date, end_date = convertLocalTimeToUTC(start_date, end_date, timezone_str)
    match = {}
    if marketplace_id and marketplace_id not in ["", "all", "custom"]:
        match['marketplace_id'] = ObjectId(marketplace_id)
    if product_id and product_id != []:
        match["_id"] = {"$in": [ObjectId(pid) for pid in product_id]}
    elif brand_id and brand_id != []:
        match["brand_id"] = {"$in": [ObjectId(bid) for bid in brand_id]}
    elif manufacturer_name and manufacturer_name != []:
        match["manufacturer_name"] = {"$in": manufacturer_name}
    if parent_search:
        match["parent_sku"] = {"$regex": parent_search, "$options": "i"}
    if not parent and sku_search:
        match["sku"] = {"$regex": sku_search, "$options": "i"}
    if search_query:
        search_query = re.escape(search_query.strip())
        match["$or"] = [
            {"product_title": {"$regex": search_query, "$options": "i"}},
            {"sku": {"$regex": search_query, "$options": "i"}},
            {'asin':{"$regex":search_query,"$options":'i'}},
            {'product_id':{"$regex":search_query,"$options":'i'}}
        ]
    if parent:
        return get_parent_products(match, page, page_size, start_date, end_date, 
                                              today_start_date, today_end_date, sort_by, sort_by_value)
    else:
        return get_individual_products(match, page, page_size, start_date, end_date, 
                                                  today_start_date, today_end_date, sort_by, sort_by_value)
def get_parent_products(match, page, page_size, start_date, end_date, 
                                   today_start_date, today_end_date, sort_by, sort_by_value):
    pipeline = []
    if match:
        pipeline.append({"$match": match})
    pipeline.extend([
        {
            "$lookup": {
                "from": "marketplace",
                "localField": "marketplace_id",
                "foreignField": "_id",
                "as": "marketplace_info"
            }
        },
        {"$unwind": "$marketplace_info"},
        {
            "$addFields": {
                "calculated_cogs": {
                    "$cond": {
                        "if": {"$eq": ["$marketplace_info.name", "Amazon"]},
                        "then": {"$ifNull": ["$total_cogs", 0]},
                        "else": {"$ifNull": ["$w_total_cogs", 0]}
                    }
                }
            }
        },
        {
            "$group": {
                "_id": "$parent_sku",
                "first_product": {"$first": "$$ROOT"},
                "total_stock": {"$sum": {"$ifNull": ["$quantity", 0]}},
                "min_price": {"$min": {"$ifNull": ["$price", 0]}},
                "max_price": {"$max": {"$ifNull": ["$price", 0]}},
                "sku_count": {"$sum": 1},
                "total_cogs": {"$sum": "$calculated_cogs"},
                "product_ids": {"$push": {"$toString": "$_id"}},
                "vendor_funding_sum": {"$sum": {"$ifNull": ["$vendor_funding", 0]}}
            }
        },
        {
            "$project": {
                "_id": 0,
                "parent_sku": "$_id",
                "id": {"$toString": "$first_product._id"},
                "title": {"$ifNull": ["$first_product.product_title", ""]},
                "imageUrl": {"$ifNull": ["$first_product.image_url", ""]},
                "marketplace": {"$ifNull": ["$first_product.marketplace_info.name", ""]},
                "category": {"$ifNull": ["$first_product.category", ""]},
                "product_id": {"$ifNull": ["$first_product.product_id", ""]},
                "sku_count": 1,
                "stock": "$total_stock",
                "price_start": "$min_price",
                "price_end": "$max_price",
                "cogs": {"$round": ["$total_cogs", 2]},
                "totalchannelFees": {"$round": ["$total_cogs", 2]},
                "product_ids": 1,
                "vendor_funding": "$vendor_funding_sum"
            }
        }
    ])
    count_pipeline = pipeline + [{"$count": "total"}]
    if sort_by and sort_by in ["price_start", "price_end", "stock", "sku_count"]:
        pipeline.append({"$sort": {sort_by: int(sort_by_value)}})
    pipeline.extend([
        {"$skip": (page - 1) * page_size},
        {"$limit": page_size}
    ])
    total_result = list(Product.objects.aggregate(*count_pipeline))
    total_products = total_result[0]["total"] if total_result else 0
    products_result = list(Product.objects.aggregate(*pipeline))
    all_product_ids = []
    for group in products_result:
        all_product_ids.extend(group["product_ids"])
    sales_data = batch_get_sales_data_optimized(all_product_ids, start_date, end_date, today_start_date, today_end_date)
    processed_products = []
    for group in products_result:
        total_sales_today = 0
        total_units_today = 0
        total_revenue = 0
        total_net_profit = 0
        total_units_period = 0
        total_revenue_period = 0
        total_net_profit_period = 0
        for product_id in group["product_ids"]:
            product_sales = sales_data.get(product_id, {
                "today": {"revenue": 0, "units": 0},
                "period": {"revenue": 0, "units": 0},
                "compare": {"revenue": 0, "units": 0}
            })
            cogs = group["cogs"] / len(group["product_ids"])  
            vendor_funding = group["vendor_funding"] / len(group["product_ids"])  
            total_sales_today += product_sales["today"]["revenue"]
            total_units_today += product_sales["period"]["units"]
            total_revenue += product_sales["period"]["revenue"]
            period_profit = (product_sales["period"]["revenue"] - (cogs * product_sales["period"]["units"]) + 
                           (vendor_funding * product_sales["period"]["units"]))
            total_net_profit += period_profit
            total_units_period += product_sales["compare"]["units"] - product_sales["period"]["units"]
            total_revenue_period += product_sales["compare"]["revenue"] - product_sales["period"]["revenue"]
            compare_profit = (product_sales["compare"]["revenue"] - (cogs * product_sales["compare"]["units"]) + 
                            (vendor_funding * product_sales["compare"]["units"]))
            total_net_profit_period += compare_profit - period_profit
        margin = (total_net_profit / total_revenue) * 100 if total_revenue > 0 else 0
        margin_period = ((total_net_profit_period / (total_revenue + total_revenue_period)) * 100 if (total_revenue + total_revenue_period) > 0 else 0) - margin
        group.update({
            "salesForToday": round(total_sales_today, 2),
            "unitsSoldForToday": total_units_today,
            "unitsSoldForPeriod": total_units_period,
            "refunds": 0,
            "refundsforPeriod": 0,
            "refundsAmount": 0,
            "refundsAmountforPeriod": 0,
            "grossRevenue": round(total_revenue, 2),
            "grossRevenueforPeriod": round(total_revenue_period, 2),
            "netProfit": round(total_net_profit, 2),
            "netProfitforPeriod": round(total_net_profit_period, 2),
            "margin": round(margin, 2),
            "marginforPeriod": round(margin_period, 2)
        })
        group.pop("product_ids", None)
        group.pop("vendor_funding", None)
        processed_products.append(group)
    calculated_fields = {'salesForToday', 'unitsSoldForToday', 'grossRevenue', 'netProfit', 'margin', 
                        'unitsSoldForPeriod', 'grossRevenueforPeriod', 'netProfitforPeriod', 'marginforPeriod'}
    if sort_by and sort_by in calculated_fields:
        reverse_sort = sort_by_value == -1
        processed_products.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse_sort)
    response_data = {
        "total_products": total_products,
        "page": page,
        "page_size": page_size,
        "products": clean_json_floats(processed_products),
        "tab_type": "parent"
    }
    return JsonResponse(response_data, safe=False)
def get_individual_products(match, page, page_size, start_date, end_date, 
                            today_start_date, today_end_date, sort_by, sort_by_value):
    db_sortable_fields = {
        'price': 'price',
        'stock': 'quantity',
        'title': 'product_title',
        'category': 'category',
        'product_id': 'product_id',
        'parent_sku': 'sku'
    }
    calculated_fields = {
        'salesForToday', 'unitsSoldForToday', 'grossRevenue', 'netProfit', 
        'margin', 'unitsSoldForPeriod', 'grossRevenueforPeriod', 
        'netProfitforPeriod', 'marginforPeriod'
    }
    pipeline = []
    if match:
        pipeline.append({"$match": match})
    pipeline.append({
        "$project": {
            "product_id": 1, "sku": 1, "price": 1, "quantity": 1, "marketplace_id": 1,
            "total_cogs": 1, "w_total_cogs": 1, "referral_fee": 1, "a_shipping_cost": 1,
            "walmart_fee": 1, "w_shiping_cost": 1, "fullfillment_by_channel": 1,
            "image_url": 1, "product_title": 1, "listing_quality_score": 1,
            "category": 1, "vendor_funding": 1
        }
    })
    pipeline.extend([
        {
            "$lookup": {
                "from": "marketplace",
                "localField": "marketplace_id",
                "foreignField": "_id",
                "as": "marketplace_ins"
            }
        },
        {"$unwind": "$marketplace_ins"},
        {
            "$addFields": {
                "calculated_cogs": {
                    "$cond": {
                        "if": {"$eq": ["$marketplace_ins.name", "Amazon"]},
                        "then": {"$ifNull": ["$total_cogs", 0]},
                        "else": {"$ifNull": ["$w_total_cogs", 0]}
                    }
                },
                "calculated_fees": {
                    "$cond": {
                        "if": {"$eq": ["$marketplace_ins.name", "Amazon"]},
                        "then": {"$add": [{"$ifNull": ["$referral_fee", 0]}, {"$ifNull": ["$a_shipping_cost", 0]}]},
                        "else": {"$add": [{"$ifNull": ["$walmart_fee", 0]}, {"$ifNull": ["$w_shiping_cost", 0]}]}
                    }
                }
            }
        }
    ])
    if sort_by and sort_by in db_sortable_fields:
        pipeline.append({"$sort": {db_sortable_fields[sort_by]: int(sort_by_value)}})
    facet_stage = {
        "$facet": {
            "metadata": [{"$count": "total"}],
            "data": [
                {"$skip": (page - 1) * page_size},
                {"$limit": page_size},
                {
                    "$project": {
                        "_id": 0,
                        "id": {"$toString": "$_id"},
                        "product_id": {"$ifNull": ["$product_id", "N/A"]},
                        "parent_sku": {"$ifNull": ["$sku", "N/A"]},
                        "imageUrl": {"$ifNull": ["$image_url", "N/A"]},
                        "title": {"$ifNull": ["$product_title", "N/A"]},
                        "marketplace": {"$ifNull": ["$marketplace_ins.name", "N/A"]},
                        "fulfillmentChannel": {
                            "$cond": {
                                "if": {"$eq": ["$fullfillment_by_channel", True]},
                                "then": "FBA",
                                "else": "FBM"
                            }
                        },
                        "price": {"$round": [{"$ifNull": ["$price", 0]}, 2]},
                        "stock": {"$ifNull": ["$quantity", 0]},
                        "listingScore": {"$ifNull": ["$listing_quality_score", 0]},
                        "cogs": {"$round": ["$calculated_cogs", 2]},
                        "category": {"$ifNull": ["$category", "N/A"]},
                        "vendor_funding": {"$ifNull": ["$vendor_funding", 0]},
                        "totalchannelFees": {"$round": ["$calculated_fees", 2]}
                    }
                }
            ]
        }
    }
    pipeline.append(facet_stage)
    result = list(Product.objects.aggregate(*pipeline))
    total_products = result[0]["metadata"][0]["total"] if result[0]["metadata"] else 0
    products = result[0]["data"]
    product_ids = [p["id"] for p in products]
    sales_data = batch_get_sales_data_optimized(product_ids, start_date, end_date, today_start_date, today_end_date)
    cogs = np.array([p["cogs"] for p in products])
    vendor_funding = np.array([p["vendor_funding"] for p in products])
    for i, product in enumerate(products):
        product_id = product["id"]
        product_sales = sales_data.get(product_id, {
            "today": {"revenue": 0, "units": 0},
            "period": {"revenue": 0, "units": 0},
            "compare": {"revenue": 0, "units": 0}
        })
        today_revenue = product_sales["today"]["revenue"]
        period_revenue = product_sales["period"]["revenue"]
        period_units = product_sales["period"]["units"]
        compare_revenue = product_sales["compare"]["revenue"]
        compare_units = product_sales["compare"]["units"]
        net_profit = (period_revenue - (cogs[i] * period_units)) + (vendor_funding[i] * period_units)
        compare_profit = (compare_revenue - (cogs[i] * compare_units)) + (vendor_funding[i] * compare_units)
        product.update({
            "salesForToday": round(today_revenue, 2),
            "unitsSoldForToday": round(period_units, 2),
            "grossRevenue": round(period_revenue, 2),
            "netProfit": round(net_profit, 2),
            "margin": round((net_profit / period_revenue) * 100 if period_revenue > 0 else 0, 2),
            "unitsSoldForPeriod": round(compare_units - period_units, 2),
            "grossRevenueforPeriod": round(compare_revenue - period_revenue, 2),
            "netProfitforPeriod": round(compare_profit - net_profit, 2),
            "marginforPeriod": round(((compare_profit / compare_revenue) * 100 if compare_revenue > 0 else 0) - 
                                   ((net_profit / period_revenue) * 100 if period_revenue > 0 else 0), 2),
            "refunds": 0,
            "refundsforPeriod": 0,
            "refundsAmount": 0,
            "refundsAmountforPeriod": 0
        })
    if sort_by and sort_by in calculated_fields:
        reverse_sort = sort_by_value == -1
        products.sort(key=lambda x: x.get(sort_by, 0), reverse=reverse_sort)
    response_data = {
        "total_products": total_products,
        "page": page,
        "page_size": page_size,
        "products": clean_json_floats(products),
        "tab_type": "sku"
    }
    return JsonResponse(response_data, safe=False)
def batch_get_sales_data_optimized(product_ids, start_date, end_date, today_start_date, today_end_date):
    """Highly optimized batch fetch sales data with caching and connection pooling"""
    if not product_ids:
        return {}
    compare_start, compare_end = getPreviousDateRange(start_date, end_date)
    import asyncio
    from concurrent.futures import ThreadPoolExecutor
    sales_data = {}
    chunk_size = 50
    chunks = [product_ids[i:i + chunk_size] for i in range(0, len(product_ids), chunk_size)]
    def process_chunk(chunk):
        chunk_sales = {}
        try:
            with ThreadPoolExecutor(max_workers=min(len(chunk), 20)) as executor:
                futures = []
                for product_id in chunk:
                    future = executor.submit(get_single_product_sales, product_id, 
                                           today_start_date, today_end_date, 
                                           start_date, end_date, 
                                           compare_start, compare_end)
                    futures.append((product_id, future))
                for product_id, future in futures:
                    try:
                        chunk_sales[product_id] = future.result(timeout=5)  
                    except Exception as e:
                        chunk_sales[product_id] = {
                            "today": {"revenue": 0, "units": 0},
                            "period": {"revenue": 0, "units": 0},
                            "compare": {"revenue": 0, "units": 0}
                        }
        except Exception as e:
            for product_id in chunk:
                chunk_sales[product_id] = {
                    "today": {"revenue": 0, "units": 0},
                    "period": {"revenue": 0, "units": 0},
                    "compare": {"revenue": 0, "units": 0}
                }
        return chunk_sales
    with ThreadPoolExecutor(max_workers=min(len(chunks), 5)) as executor:
        chunk_futures = [executor.submit(process_chunk, chunk) for chunk in chunks]
        for future in chunk_futures:
            try:
                chunk_result = future.result(timeout=30)  
                sales_data.update(chunk_result)
            except Exception as e:
                logger.error("Error processing chunk:%s",e)
        return sales_data
def get_single_product_sales(product_id, today_start_date, today_end_date, 
                           start_date, end_date, compare_start, compare_end):
    """Optimized single product sales data fetch"""
    try:
        today_sales = getdaywiseproductssold(today_start_date, today_end_date, product_id, False)
        period_sales = getdaywiseproductssold(start_date, end_date, product_id, False)
        compare_sales = getdaywiseproductssold(compare_start, compare_end, product_id, False)
        return {
            "today": {
                "revenue": sum(sale["total_price"] for sale in today_sales),
                "units": sum(sale["total_quantity"] for sale in today_sales)
            },
            "period": {
                "revenue": sum(sale["total_price"] for sale in period_sales),
                "units": sum(sale["total_quantity"] for sale in period_sales)
            },
            "compare": {
                "revenue": sum(sale["total_price"] for sale in compare_sales),
                "units": sum(sale["total_quantity"] for sale in compare_sales)
            }
        }
    except Exception as e:
        return {
            "today": {"revenue": 0, "units": 0},
            "period": {"revenue": 0, "units": 0},
            "compare": {"revenue": 0, "units": 0}
        }
def clean_json_floats(obj):
    """Optimized NaN and Inf cleaning"""
    import math
    if isinstance(obj, float):
        return None if (math.isnan(obj) or math.isinf(obj)) else obj
    elif isinstance(obj, dict):
        return {k: clean_json_floats(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_json_floats(i) for i in obj]
    return obj
@csrf_exempt
def getPeriodWiseData(request):
    def to_utc_format(dt):
        return dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id', [])
    manufacturer_name = json_request.get('manufacturer_name', [])
    fulfillment_channel = json_request.get('fulfillment_channel', None)
    timezone_str = 'US/Pacific'
    def calculate_metrics_sync(start_date, end_date):
        return calculate_metricss(
            start_date, end_date, 
            marketplace_id, brand_id, 
            product_id, manufacturer_name, 
            fulfillment_channel,
            timezone_str, False,
            use_threads=False
        )
    def format_period_metrics(label, current_start, current_end, prev_start, prev_end):
        current_metrics = calculate_metrics_sync(current_start, current_end)
        period_format = {
            "current": {"from": to_utc_format(current_start)},
            "previous": {"from": to_utc_format(prev_start)}
        }
        if label not in ['Today', 'Yesterday']:
            period_format["current"]["to"] = to_utc_format(current_end)
            period_format["previous"]["to"] = to_utc_format(prev_end)
        output = {
            "label": label,
            "period": period_format
        }
        for key in current_metrics:
            output[key] = {
                "current": current_metrics[key],
            }
        return output
    date_ranges = {
        "yesterday": get_date_range("Yesterday", timezone_str),
        "last7Days": get_date_range("Last 7 days", timezone_str),
        "last30Days": get_date_range("Last 30 days", timezone_str),
        "yearToDate": get_date_range("This Year", timezone_str),
        "lastYear": get_date_range("Last Year", timezone_str)
    }
    period_jobs = [
        ("yesterday", "Yesterday", date_ranges["yesterday"][0], date_ranges["yesterday"][1],
         date_ranges["yesterday"][0] - timedelta(days=1), date_ranges["yesterday"][0] - timedelta(seconds=1)),
        ("last7Days", "Last 7 Days", date_ranges["last7Days"][0], date_ranges["last7Days"][1],
         date_ranges["last7Days"][0] - timedelta(days=7), date_ranges["last7Days"][0] - timedelta(seconds=1)),
        ("last30Days", "Last 30 Days", date_ranges["last30Days"][0], date_ranges["last30Days"][1],
         date_ranges["last30Days"][0] - timedelta(days=30), date_ranges["last30Days"][0] - timedelta(seconds=1)),
        ("yearToDate", "Year to Date", date_ranges["yearToDate"][0], date_ranges["yearToDate"][1],
         date_ranges["lastYear"][0], date_ranges["lastYear"][1])
    ]
    response_data = {}
    with ThreadPoolExecutor(max_workers=8) as executor:
        future_to_label = {
            executor.submit(format_period_metrics, label, cur_start, cur_end, prev_start, prev_end): key
            for key, label, cur_start, cur_end, prev_start, prev_end in period_jobs
        }
        for future in as_completed(future_to_label):
            key = future_to_label[future]
            try:
                response_data[key] = future.result()
            except Exception as exc:
                response_data[key] = {"error": str(exc)}
    return JsonResponse(response_data, safe=False)

@csrf_exempt
def getPeriodWiseDataXl(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id', [])
    manufacturer_name = json_request.get('manufacturer_name', [])
    fulfillment_channel = json_request.get('fulfillment_channel')
    timezone_str = json_request.get('timezone', 'US/Pacific')
    periods = {
        "Yesterday": get_date_range("Yesterday", timezone_str),
        "Last 7 Days": get_date_range("Last 7 days", timezone_str),
        "Last 30 Days": get_date_range("Last 30 days", timezone_str),
        "Month to Date": get_date_range("This Month", timezone_str),
        "Year to Date": get_date_range("This Year", timezone_str),
    }
    def create_row(label, start, end):
        data = calculate_metricss(
            start, end,
            marketplace_id,
            brand_id,
            product_id,
            manufacturer_name,
            fulfillment_channel,
            timezone_str,
            include_extra_fields=True
        )
        return [
            label,
            data.get("seller", ""),
            data.get("marketplace", ""),
            start.strftime("%Y-%m-%d"),
            end.strftime("%Y-%m-%d"),
            data.get("grossRevenue", 0),
            data.get("expenses", 0),
            data.get("netProfit", 0),
            data.get("roi", 0),
            data.get("unitsSold", 0),
            data.get("refunds", 0),
            data.get("skuCount", 0),
            data.get("sessions", 0),
            data.get("pageViews", 0),
            data.get("unitSessionPercentage", 0),
            data.get("margin", 0)
        ]
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {
            label: executor.submit(create_row, label, start, end)
            for label, (start, end) in periods.items()
        }
        period_rows = [futures[label].result() for label in periods]
    headers = [
        "Period", "Seller", "Marketplace", "Start Date", "End Date",
        "Gross Revenue", "Expenses", "Net Profit", "ROI %",
        "Units Sold", "Refunds", "SKU Count", "Sessions",
        "Page Views", "Unit Session %", "Margin %"
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Period Metrics"
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    for row in period_rows:
        ws.append(row)
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=PeriodWiseMetrics.xlsx'
    wb.save(response)
    return response
@csrf_exempt
def exportPeriodWiseCSV(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id', [])
    manufacturer_name = json_request.get('manufacturer_name', [])
    fulfillment_channel = json_request.get('fulfillment_channel', None)
    timezone_str = json_request.get('timezone', 'US/Pacific')
    periods = {
        "Yesterday": get_date_range("Yesterday", timezone_str),
        "Last 7 Days": get_date_range("Last 7 days", timezone_str),
        "Last 30 Days": get_date_range("Last 30 days", timezone_str),
        "Month to Date": get_date_range("This Month", timezone_str),
        "Year to Date": get_date_range("This Year", timezone_str),
    }
    def create_row(label, start, end):
        data = calculate_metricss(
            start, end,
            marketplace_id,
            brand_id,
            product_id,
            manufacturer_name,
            fulfillment_channel,
            timezone_str=timezone_str,
            include_extra_fields=True
        )
        return [
            label,
            data.get("seller", ""),
            data.get("marketplace", ""),
            start.strftime("%Y-%m-%d"),
            end.strftime("%Y-%m-%d"),
            str(data.get("grossRevenue", 0)),
            str(data.get("expenses", 0)),
            str(data.get("netProfit", 0)),
            str(data.get("roi", 0)),
            str(data.get("unitsSold", 0)),
            str(data.get("refunds", 0)),
            str(data.get("skuCount", 0)),
            str(data.get("sessions", 0)),
            str(data.get("pageViews", 0)),
            str(data.get("unitSessionPercentage", 0)),
            str(data.get("margin", 0))
        ]
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {
            label: executor.submit(create_row, label, start, end)
            for label, (start, end) in periods.items()
        }
        period_rows = [futures[label].result() for label in periods]
    headers = [
        "Period", "Seller", "Marketplace", "Start Date", "End Date",
        "Gross Revenue", "Expenses", "Net Profit", "ROI %",
        "Units Sold", "Refunds", "SKU Count", "Sessions",
        "Page Views", "Unit Session %", "Margin %"
    ]
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="PeriodWiseMetrics.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(period_rows)
    return response
@csrf_exempt
def getPeriodWiseDataCustom(request):
    import pytz
    from datetime import datetime, timedelta
    from rest_framework.parsers import JSONParser
    from django.http import JsonResponse
    from concurrent.futures import ThreadPoolExecutor
    
    def to_utc_format(dt):
        return dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id', [])
    manufacturer_name = json_request.get('manufacturer_name', [])
    fulfillment_channel = json_request.get('fulfillment_channel', None)
    timezone_str = "US/Pacific"
    preset = json_request.get("preset")
    start_date = json_request.get("start_date")
    end_date = json_request.get("end_date")
    if start_date:
        local_tz = pytz.timezone(timezone_str)
        naive_from_date = datetime.strptime(start_date, '%Y-%m-%d')
        naive_to_date = datetime.strptime(end_date, '%Y-%m-%d')
        localized_from_date = local_tz.localize(naive_from_date)
        localized_to_date = local_tz.localize(naive_to_date.replace(hour=23, minute=59, second=59))  
        from_date = localized_from_date.astimezone(pytz.UTC)
        to_date = localized_to_date.astimezone(pytz.UTC)
    else:
        from_date, to_date = get_date_range(preset, timezone_str)
    duration = to_date - from_date
    prev_from, prev_to = from_date - duration, to_date - duration
    today_start, today_end = get_date_range("Today", timezone_str)
    yesterday_start, yesterday_end = get_date_range("Yesterday", timezone_str)
    last7_start, last7_end = get_date_range("Last 7 days", timezone_str)
    last7_prev_start = today_start - timedelta(days=14)
    last7_prev_end = last7_start - timedelta(seconds=1)
    day_before_yesterday_start = yesterday_start - timedelta(days=1)
    day_before_yesterday_end = yesterday_end - timedelta(days=1)
    def format_metrics_response(current, previous):
        def format_metric(metric):
            current_value = sanitize_data(current.get(metric, 0))
            previous_value = sanitize_data(previous.get(metric, 0))
            delta = round(current_value - previous_value, 2)
            return {
                "current": current_value,
                "previous": previous_value,
                "delta": delta
            }
        summary_metrics = [
            "grossRevenue", "netProfit", "expenses", "unitsSold", "refunds", "skuCount",
            "sessions", "pageViews", "unitSessionPercentage", "margin", "roi", "orders"
        ]
        summary = {metric: format_metric(metric) for metric in summary_metrics}
        def net_profit_calc(metrics):
            return {
                "gross": sanitize_data(metrics.get("grossRevenue", 0)),
                "totalCosts": sanitize_data(metrics.get("expenses", 0)),
                "productRefunds": sanitize_data(metrics.get("refunds", 0)),
                "totalTax": sanitize_data(metrics.get("tax_price", 0)),
                "totalTaxWithheld": 0,
                "ppcProductCost": 0,
                "ppcBrandsCost": 0,
                "ppcDisplayCost": 0,
                "ppcStCost": 0,
                "cogs": sanitize_data(metrics.get("total_cogs", 0)),
                "product_cost": sanitize_data(metrics.get("product_cost", 0)),
                "shipping_cost": sanitize_data(metrics.get("shipping_cost", 0)),
            }
        return {
            "summary": summary,
            "netProfitCalculation": {
                "current": net_profit_calc(current),
                "previous": net_profit_calc(previous),
            }
        }
    def to_local_date_string(dt, tz_str):
        local_tz = pytz.timezone(tz_str)
        return dt.astimezone(local_tz).strftime("%Y-%m-%d")
    def create_period_response(label, cur_from, cur_to, prev_from, prev_to, current_metrics, previous_metrics):
        date_ranges = {
            "current": {"from": to_utc_format(cur_from), "to": to_utc_format(cur_to),"from_local":to_local_date_string(cur_from,timezone_str),'to_local':to_local_date_string(cur_to,timezone_str)},
            "previous": {"from": to_utc_format(prev_from), "to": to_utc_format(prev_to,),"from_local":to_local_date_string(prev_from,timezone_str),'to_local':to_local_date_string(prev_to,timezone_str)}
        }
        metrics_response = format_metrics_response(current_metrics, previous_metrics)
        return {
            "dateRanges": date_ranges,
            **metrics_response  
        }
    with ThreadPoolExecutor(max_workers=8) as executor:
        future_today_current = executor.submit(
            calculate_metricss, 
            today_start, today_end, 
            marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,
            timezone_str, False, True
        )
        future_today_previous = executor.submit(
            calculate_metricss, 
            yesterday_start, yesterday_end, 
            marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,
            timezone_str, False, True
        )
        future_yesterday_current = executor.submit(
            calculate_metricss, 
            yesterday_start, yesterday_end, 
            marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,
            timezone_str, False, True
        )
        future_yesterday_previous = executor.submit(
            calculate_metricss, 
            day_before_yesterday_start, day_before_yesterday_end, 
            marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,
            timezone_str, False, True
        )
        future_last7_current = executor.submit(
            calculate_metricss, 
            last7_start, last7_end, 
            marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,
            timezone_str, False, True
        )
        future_last7_previous = executor.submit(
            calculate_metricss, 
            last7_prev_start, last7_prev_end, 
            marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,
            timezone_str, False, True
        )
        future_custom_current = executor.submit(
            calculate_metricss, 
            from_date, to_date, 
            marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,
            timezone_str, False, True
        )
        future_custom_previous = executor.submit(
            calculate_metricss, 
            prev_from, prev_to, 
            marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,
            timezone_str, False, True
        )
        today_current = future_today_current.result()
        today_previous = future_today_previous.result()
        yesterday_current = future_yesterday_current.result()
        yesterday_previous = future_yesterday_previous.result()
        last7_current = future_last7_current.result()
        last7_previous = future_last7_previous.result()
        custom_current = future_custom_current.result()
        custom_previous = future_custom_previous.result()
    response_data = {
        "today": create_period_response(
            "Today", today_start, today_end, yesterday_start, yesterday_end,
            today_current, today_previous
        ),
        "yesterday": create_period_response(
            "Yesterday", yesterday_start, yesterday_end, 
            day_before_yesterday_start, day_before_yesterday_end,
            yesterday_current, yesterday_previous
        ),
        "last7Days": create_period_response(
            "Last 7 Days", last7_start, last7_end, last7_prev_start, last7_prev_end,
            last7_current, last7_previous
        ),
        "custom": create_period_response(
            preset, from_date, to_date, prev_from, prev_to,
            custom_current, custom_previous
        ),
    }
    return JsonResponse(response_data, safe=False)
@csrf_exempt
def allMarketplaceData(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id', [])
    manufacturer_name = json_request.get('manufacturer_name', [])
    fulfillment_channel = json_request.get('fulfillment_channel', None)
    preset = json_request.get('preset')
    timezone_str = 'US/Pacific'
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date:
        from_date, to_date = convertdateTotimezone(start_date, end_date, timezone_str)
    else:
        from_date, to_date = get_date_range(preset, timezone_str)
    marketplace_dict = {
        str(mp.id): mp.name for mp in Marketplace.objects.only("id", "name")
    }
    def fetch_item_details_bulk(item_ids):
        pipeline = [
            {"$match": {"_id": {"$in": item_ids}}},
            {
                "$lookup": {
                    "from": "product",
                    "localField": "ProductDetails.product_id",
                    "foreignField": "_id",
                    "as": "product_ins"
                }
            },
            {"$unwind": {"path": "$product_ins", "preserveNullAndEmptyArrays": True}},
            {
                "$project": {
                    "_id": 1,
                    "price": "$Pricing.ItemPrice.Amount",
                    "tax_price": "$Pricing.ItemTax.Amount",
                    "cogs": {"$ifNull": ["$product_ins.cogs", 0.0]},
                    "sku": "$product_ins.sku",
                    "total_cogs": {"$ifNull": ["$product_ins.total_cogs", 0]},
                    "w_total_cogs": {"$ifNull": ["$product_ins.w_total_cogs", 0]},
                    "vendor_funding": {"$ifNull": ["$product_ins.vendor_funding", 0]},
                }
            }
        ]
        return {
            item["_id"]: item
            for item in OrderItems.objects.aggregate(*pipeline)
        }
    def process_orders(orders):
        all_item_ids = [item_id for order in orders for item_id in order["order_items"]]
        item_map = fetch_item_details_bulk(all_item_ids)
        grouped_orders = defaultdict(list)
        for order in orders:
            key = (order.get("marketplace_id"), order.get("currency"))
            grouped_orders[key].append(order)
        marketplace_metrics = defaultdict(lambda: {"currency_list": []})
        for (mp_id, currency), order_list in grouped_orders.items():
            gross_revenue = 0
            total_cogs = 0
            total_units = 0
            tax_price = 0
            total_product_cost = 0
            temp_price = 0
            vendor_funding = 0
            sku_set = set()
            marketplace_name = marketplace_dict.get(str(mp_id), "")
            for order in order_list:
                gross_revenue += order["order_total"]
                total_units += order['items_order_quantity']
                for item_id in order['order_items']:
                    item_data = item_map.get(item_id)
                    if not item_data:
                        continue
                    temp_price += item_data['price']
                    tax_price += item_data['tax_price']
                    cogs_value = item_data['total_cogs'] if marketplace_name == "Amazon" else item_data['w_total_cogs']
                    total_cogs += cogs_value
                    vendor_funding += item_data['vendor_funding']
                    total_product_cost += item_data['price']
                    if item_data.get('sku'):
                        sku_set.add(item_data['sku'])
            expenses = total_cogs
            net_profit = (temp_price - expenses) + vendor_funding
            roi = (net_profit / expenses) * 100 if expenses > 0 else 0
            margin = (net_profit / gross_revenue) * 100 if gross_revenue > 0 else 0
            currency_data = {
                "currency": currency,
                "grossRevenue": round(gross_revenue, 2),
                "expenses": round(expenses, 2),
                "netProfit": round(net_profit, 2),
                "roi": round(roi, 2),
                "unitsSold": total_units,
                "refunds": 0,
                "skuCount": len(sku_set),
                "margin": round(margin, 2),
                "sessions": 0,
                "pageViews": 0,
                "unitSessionPercentage": 0,
                "seller": "",
                "tax_price": round(tax_price, 2),
                "total_cogs": round(total_cogs, 2),
                "product_cost": round(total_product_cost, 2),
                "shipping_cost": 0
            }
            marketplace_metrics[marketplace_name]["currency_list"].append(currency_data)
        return [
            {
                "image": (
                    "https://i.pinimg.com/originals/01/ca/da/01cada77a0a7d326d85b7969fe26a728.jpg"
                    if mp == "Amazon" else
                    "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcRzjtf8dzq48TtkzeRYx2-_li3gTCkstX2juA&s"
                    if mp == "Walmart" else ""
                ),
                "marketplace": mp,
                "currency_list": data["currency_list"]
            }
            for mp, data in marketplace_metrics.items()
        ]
    def calculate_metrics(orders):
        all_item_ids = [item_id for order in orders for item_id in order["order_items"]]
        item_map = fetch_item_details_bulk(all_item_ids)
        gross_revenue = 0
        total_cogs = 0
        net_profit = 0
        margin = 0
        total_units = 0
        temp_price = 0
        sku_set = set()
        vendor_funding = 0
        tax_price = 0
        for order in orders:
            gross_revenue += order['order_total']
            total_units += order['items_order_quantity']
            for item_id in order['order_items']:
                item_data = item_map.get(item_id)
                if not item_data:
                    continue
                temp_price += item_data['price']
                tax_price += item_data['tax_price']
                marketplace_name = order.get("marketplace_name", "Amazon")
                total_cogs += item_data['total_cogs'] if marketplace_name == "Amazon" else item_data['w_total_cogs']
                vendor_funding += item_data['vendor_funding']
                if item_data.get('sku'):
                    sku_set.add(item_data['sku'])
        net_profit = (temp_price - total_cogs) + vendor_funding
        margin = (net_profit / gross_revenue) * 100 if gross_revenue > 0 else 0
        roi = (net_profit / total_cogs) * 100 if total_cogs > 0 else 0
        return {
            "grossRevenue": round(gross_revenue, 2),
            "expenses": round(total_cogs, 2),
            "netProfit": round(net_profit, 2),
            "roi": round(roi, 2),
            "unitsSold": total_units,
            "refunds": 0,
            "skuCount": len(sku_set),
            "sessions": 0,
            "pageViews": 0,
            "unitSessionPercentage": 0,
            "margin": round(margin, 2),
            "seller": "",
            "tax_price": round(tax_price, 2),
            "total_cogs": round(total_cogs, 2),
            "product_cost": round(gross_revenue, 2),
            "shipping_cost": 0
        }
    def create_period_response(cur_from, cur_to, prev_from, prev_to):
        current_orders = grossRevenue(cur_from, cur_to, marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel, timezone_str)
        previous_orders = grossRevenue(prev_from, prev_to, marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel, timezone_str)
        current = calculate_metrics(current_orders)
        previous = calculate_metrics(previous_orders)
        def with_delta(metric):
            return {
                "current": current[metric],
                "previous": previous[metric],
                "delta": round(current[metric] - previous[metric], 2)
            }
        return {
            "all_marketplace": {
                metric: with_delta(metric) for metric in [
                    "grossRevenue", "netProfit", "expenses", "unitsSold", "refunds",
                    "skuCount", "sessions", "pageViews", "unitSessionPercentage", "margin", "roi"
                ]
            },
            "marketplace_list": process_orders(current_orders)
        }
    custom_duration = to_date - from_date
    prev_from_date = from_date - custom_duration
    prev_to_date = to_date - custom_duration
    response_data = {
        "custom": create_period_response(from_date, to_date, prev_from_date, prev_to_date),
        "from_date": from_date,
        "to_date": to_date
    }
    return JsonResponse(response_data, safe=False)
@csrf_exempt
def allMarketplaceDataxl(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id',[])
    manufacturer_name = json_request.get('manufacturer_name',[])
    fulfillment_channel = json_request.get('fulfillment_channel',None)
    preset = json_request.get('preset')
    timezone_str = 'US/Pacific'
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date != None and start_date != "":
        from_date, to_date = convertdateTotimezone(start_date,end_date,timezone_str)
    else:
        from_date, to_date = get_date_range(preset,timezone_str)
    def grouped_marketplace_metrics(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str):
        orders = grossRevenue(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
        grouped_orders = defaultdict(list)
        for order in orders:
            key = (order.get("marketplace_id"), order.get("currency"))
            grouped_orders[key].append(order)
        marketplace_metrics = defaultdict(lambda: {"currency_list": []})
        for (mp_id, currency), orders in grouped_orders.items():
            gross_revenue = 0
            total_cogs = 0
            total_units = 0
            refund = 0
            tax_price = 0
            other_price = 0
            total_product_cost = 0
            temp_price = 0
            vendor_funding = 0
            sku_set = set()
            m_obj = Marketplace.objects(id=mp_id)
            marketplace = m_obj[0].name if m_obj else ""
            for order in orders:
                gross_revenue += order["order_total"]
                order_total = order["order_total"]
                total_units += order['items_order_quantity']
                tax_price = 0
                for item_id in order['order_items']:
                    item_pipeline = [
                        {"$match": {"_id": item_id}},
                        {
                            "$lookup": {
                                "from": "product",
                                "localField": "ProductDetails.product_id",
                                "foreignField": "_id",
                                "as": "product_ins"
                            }
                        },
                        {"$unwind": {"path": "$product_ins", "preserveNullAndEmptyArrays": True}},
                        {
                            "$project": {
                                "_id": 0,
                                "price": "$Pricing.ItemPrice.Amount",
                                "tax_price": "$Pricing.ItemTax.Amount",
                                "cogs": {"$ifNull": ["$product_ins.cogs", 0.0]},
                                "sku": "$product_ins.sku",
                                "total_cogs" : {"$ifNull":["$product_ins.total_cogs",0]},
                                "w_total_cogs" : {"$ifNull":["$product_ins.w_total_cogs",0]},
                                "vendor_funding" : {"$ifNull":["$product_ins.vendor_funding",0]},
                            }
                        }
                    ]
                    item_result = list(OrderItems.objects.aggregate(*item_pipeline))
                    if item_result:
                        item_data = item_result[0]
                        temp_price += item_data['price']
                        tax_price += item_data['tax_price']
                        if order['marketplace_name'] == "Amazon":
                            total_cogs += item_data['total_cogs'] 
                        else:
                            total_cogs += item_data['w_total_cogs']
                        vendor_funding += item_data['vendor_funding']
                        total_product_cost += item_data['price']
                        if item_data.get('sku'):
                            sku_set.add(item_data['sku'])
            expenses = total_cogs
            net_profit = (temp_price - expenses) + vendor_funding
            roi = (net_profit / expenses) * 100 if expenses > 0 else 0
            margin = (net_profit / gross_revenue) * 100 if gross_revenue > 0 else 0
            currency_data = {
                "Marketplace": marketplace,
                "Currency": currency,
                "Start Date": from_date.date(),
                "End Date": to_date.date(),
                "Gross Revenue": round(gross_revenue, 2),
                "Expenses": round(expenses, 2),
                "COGS": round(total_cogs, 2),
                "Net Profit": round(net_profit, 2),
                "Margin (%)": round(margin, 2),
                "ROI (%)": round(roi, 2),
                "Refunds": refund,
                "Units Sold": total_units,
            }
            marketplace_metrics[marketplace]["currency_list"].append(currency_data)
        rows = []
        for _, data in marketplace_metrics.items():
            for row in data["currency_list"]:
                rows.append(row)
        return rows
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Marketplace Metrics"
    data_rows = grouped_marketplace_metrics(from_date, to_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
    if data_rows:
        headers = list(data_rows[0].keys())
        sheet.append(headers)
        for col in range(1, len(headers) + 1):
            sheet.cell(row=1, column=col).font = Font(bold=True)
        for row in data_rows:
            sheet.append(list(row.values()))
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="marketplace_metrics_{datetime.now().date()}.xlsx"'
    return response
@csrf_exempt
def downloadMarketplaceDataCSV(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id',[])
    manufacturer_name = json_request.get('manufacturer_name',[])
    fulfillment_channel = json_request.get('fulfillment_channel',None)
    preset = json_request.get('preset')
    timezone_str = 'US/Pacific'
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date != None and start_date != "":
        from_date, to_date = convertdateTotimezone(start_date,end_date,timezone_str)
    else:
        from_date, to_date = get_date_range(preset,timezone_str)
    def grouped_marketplace_metrics(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str):
        orders = grossRevenue(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
        grouped_orders = defaultdict(list)
        for order in orders:
            key = (order.get("marketplace_id"), order.get("currency"))
            grouped_orders[key].append(order)
        marketplace_metrics = []
        for (mp_id, currency), orders in grouped_orders.items():
            gross_revenue = 0
            total_cogs = 0
            total_units = 0
            refund = 0
            tax_price = 0
            other_price = 0
            total_product_cost = 0
            vendor_funding = 0
            sku_set = set()
            m_obj = Marketplace.objects(id=mp_id)
            marketplace = m_obj[0].name if m_obj else ""
            for order in orders:
                gross_revenue += order["order_total"]
                order_total = order["order_total"]
                total_units += order['items_order_quantity']
                temp_price = 0
                tax_price = 0
                for item_id in order['order_items']:
                    item_pipeline = [
                        {"$match": {"_id": item_id}},
                        {
                            "$lookup": {
                                "from": "product",
                                "localField": "ProductDetails.product_id",
                                "foreignField": "_id",
                                "as": "product_ins"
                            }
                        },
                        {"$unwind": {"path": "$product_ins", "preserveNullAndEmptyArrays": True}},
                        {
                            "$project": {
                                "_id": 0,
                                "price": "$Pricing.ItemPrice.Amount",
                                "tax_price": "$Pricing.ItemTax.Amount",
                                "cogs": {"$ifNull": ["$product_ins.cogs", 0.0]},
                                "sku": "$product_ins.sku",
                                "total_cogs" : {"$ifNull":["$product_ins.total_cogs",0]},
                                "w_total_cogs" : {"$ifNull":["$product_ins.w_total_cogs",0]},
                                "vendor_funding" : {"$ifNull":["$product_ins.vendor_funding",0]},
                            }
                        }
                    ]
                    item_result = list(OrderItems.objects.aggregate(*item_pipeline))
                    if item_result:
                        item_data = item_result[0]
                        temp_price += item_data['price']
                        tax_price += item_data['tax_price']
                        if order['marketplace_name'] == "Amazon":
                            total_cogs += item_data['total_cogs'] 
                        else:
                            total_cogs += item_data['w_total_cogs']
                        vendor_funding += item_data['vendor_funding']
                        total_product_cost += item_data['price']
                        if item_data.get('sku'):
                            sku_set.add(item_data['sku'])
            expenses = total_cogs 
            net_profit = (total_product_cost - expenses) + vendor_funding
            roi = (net_profit / expenses) * 100 if expenses > 0 else 0
            margin = (net_profit / gross_revenue) * 100 if gross_revenue > 0 else 0
            marketplace_metrics.append({
                "Marketplace": marketplace,
                "Currency": currency,
                "Start Date": from_date.date(),
                "End Date": to_date.date(),
                "Gross Revenue": round(gross_revenue, 2),
                "Expenses": round(expenses, 2),
                "COGS": round(total_cogs, 2),
                "Net Profit": round(net_profit, 2),
                "Margin (%)": round(margin, 2),
                "ROI (%)": round(roi, 2),
                "Refunds": refund,
                "Units Sold": total_units,
            })
        return marketplace_metrics
    metrics = grouped_marketplace_metrics(from_date, to_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="marketplace_metrics.csv"'
    writer = csv.DictWriter(response, fieldnames=metrics[0].keys())
    writer.writeheader()
    for row in metrics:
        writer.writerow(row)
    return response
def sales(orders):
    sku_summary = defaultdict(lambda: {
        "sku": "",
        "product_name": "",
        "images": "",
        "unitsSold": 0,
        "grossRevenue": 0.0,
        "totalCogs": 0.0,
        "netProfit": 0.0,
        "margin": 0.0,
        "vendor_funding" : 0.0
    })
    for order in orders:
        item_ids = order.get("order_items", [])
        fulfillmentChannel = ""
        temp_price = 0.0
        total_cogs = 0.0
        sku_set = set()
        tax_price = 0
        vendor_funding = 0
        for item_id in item_ids:
            item_pipeline = [
                {"$match": {"_id": item_id}},
                {
                    "$lookup": {
                        "from": "product",
                        "localField": "ProductDetails.product_id",
                        "foreignField": "_id",
                        "as": "product_ins"
                    }
                },
                {"$unwind": {"path": "$product_ins", "preserveNullAndEmptyArrays": True}},
                {
                    "$project": {
                        "_id": 0,
                        "id" : "$product_ins._id",
                        "price": "$Pricing.ItemPrice.Amount",
                        "tax_price": "$Pricing.ItemTax.Amount",
                        "QuantityOrdered": {"$ifNull":["$ProductDetails.QuantityOrdered",0]},
                        "cogs": {"$ifNull": ["$product_ins.cogs", 0.0]},
                        "sku": "$product_ins.sku",
                        "product_name": "$product_ins.product_title",
                        "product_id" : {"$ifNull" : ["$product_ins.product_id",2]},
                        "fulfillmentChannel": {
                            "$cond": {
                            "if": {"$eq": ["$product_ins.fullfillment_by_channel", True]},
                            "then": "FBA",
                            "else": "FBM"
                            }
                            },
                        "images": "$product_ins.image_url",
                        "total_cogs" : {"$ifNull":["$product_ins.total_cogs",0]},
                        "w_total_cogs" : {"$ifNull":["$product_ins.w_total_cogs",0]},
                        "vendor_funding" : {"$ifNull":["$product_ins.vendor_funding",0]},
                    }
                }
            ]
            item_result = list(OrderItems.objects.aggregate(*item_pipeline))
            if item_result:
                item_data = item_result[0]
                sku = item_data.get("sku")
                id = item_data.get("id")
                product_name = item_data.get("product_name", "")
                product_id = item_data.get("product_id", "")
                tax_price += item_data['tax_price']
                images = item_data.get("images", [])
                price = item_data.get("price", 0.0)
                if order['marketplace_name'] == "Amazon":
                    cogs = item_data.get("total_cogs", 0.0)
                    temp_units = item_data.get("QuantityOrdered", 0)
                else:
                    cogs = item_data.get("w_total_cogs", 0.0)
                    temp_units = 1
                vendor_funding = item_data.get("vendor_funding", 0.0)
                temp_price += price
                total_cogs += cogs
                fulfillmentChannel = item_data.get("fulfillmentChannel", 0.0)
                if sku:
                    sku_set.add(sku)
                    sku_summary[sku]["id"] = str(id)
                    sku_summary[sku]["sku"] = sku
                    sku_summary[sku]["product_name"] = product_name
                    sku_summary[sku]["asin"] = product_id
                    sku_summary[sku]["images"] = images
                    sku_summary[sku]["unitsSold"] += temp_units
                    sku_summary[sku]["grossRevenue"] += price
                    sku_summary[sku]["totalCogs"] += cogs
                    sku_summary[sku]["fulfillmentChannel"] = fulfillmentChannel
                    sku_summary[sku]["vendor_funding"] += vendor_funding
        for sku in sku_set:
            gross = sku_summary[sku]["grossRevenue"]
            cogs = sku_summary[sku]["totalCogs"]
            vendor_funding = sku_summary[sku]["vendor_funding"]
            net_profit = (gross - cogs) + vendor_funding
            margin = (net_profit / gross) * 100 if gross > 0 else 0
            sku_summary[sku]["netProfit"] = round(net_profit, 2)
            sku_summary[sku]["margin"] = round(margin, 2)
    sorted_skus = sorted(sku_summary.values(), key=lambda x: x["unitsSold"], reverse=True)
    return sorted_skus
@csrf_exempt
def getProductPerformanceSummary(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id',[])
    manufacturer_name = json_request.get('manufacturer_name',[])
    fulfillment_channel = json_request.get('fulfillment_channel',None)
    timezone_str =  'US/Pacific'
    local_tz = pytz.timezone(timezone_str)
    today = datetime.now(local_tz)
    yesterday_start_date = today - timedelta(days=1)
    yesterday_start_date = yesterday_start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_end_date = yesterday_start_date.replace(hour=23, minute=59, second=59)
    previous_day_start_date = yesterday_start_date - timedelta(days=1)
    previous_day_start_date = previous_day_start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    previous_day_end_date = previous_day_start_date.replace(hour=23, minute=59, second=59)
    def fetch_data(start_date, end_date):
        return grossRevenue(start_date, end_date, marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,timezone_str)
    with ThreadPoolExecutor(max_workers=2) as executor:
        future_prev_data = executor.submit(fetch_data, previous_day_start_date, previous_day_end_date)
        future_yes_data = executor.submit(fetch_data, yesterday_start_date, yesterday_end_date)
        prev_data = future_prev_data.result()
        yes_data = future_yes_data.result()
    with ThreadPoolExecutor(max_workers=2) as executor:
        future_yes_data = executor.submit(sales, yes_data)
        future_prev_data = executor.submit(sales, prev_data)
        yes_data = future_yes_data.result()
        prev_data = future_prev_data.result()
    data = get_top_movers(yes_data, prev_data)
    return JsonResponse(data)
@csrf_exempt
def downloadProductPerformanceSummary(request):
    json_request = JSONParser().parse(request)
    action = json_request.get("action", "").lower()
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id',[])
    manufacturer_name = json_request.get('manufacturer_name',[])
    fulfillment_channel = json_request.get('fulfillment_channel',None)
    timezone_str = 'US/Pacific'
    local_tz = pytz.timezone(timezone_str)
    today = datetime.now(local_tz)
    yesterday_start_date = today - timedelta(days=1)
    yesterday_start_date = yesterday_start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_end_date = yesterday_start_date.replace(hour=23, minute=59, second=59)
    previous_day_start_date = yesterday_start_date - timedelta(days=1)
    previous_day_start_date = previous_day_start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    previous_day_end_date = previous_day_start_date.replace(hour=23, minute=59, second=59)
    def fetch_data(start_date, end_date):
        return grossRevenue(start_date, end_date, marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel,timezone_str)
    with ThreadPoolExecutor(max_workers=2) as executor:
        future_prev_data = executor.submit(fetch_data, previous_day_start_date, previous_day_end_date)
        future_yes_data = executor.submit(fetch_data, yesterday_start_date, yesterday_end_date)
        prev_data = future_prev_data.result()
        yes_data = future_yes_data.result()
    with ThreadPoolExecutor(max_workers=2) as executor:
        future_yes_data = executor.submit(sales, yes_data)
        future_prev_data = executor.submit(sales, prev_data)
        yes_data = future_yes_data.result()
        prev_data = future_prev_data.result()
    data = get_top_movers(yes_data, prev_data)
    if action == "top":
        final_summary = data.get('top_3_products', []) if data else []
    elif action == "least":
        final_summary = data.get('least_3_products', []) if data else []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Performance"
    headers = [
         "Product Title","ASIN","SKU","Fulfillment Type","Marketplace" ,"Start Date","End Date","Gross Revenue","Net Profit","Units Sold","Trend"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    for data in final_summary:
        ws.append([
            data["product_name"],
            data["asin"],
            data["sku"],
            data["fulfillmentChannel"],
            yesterday_start_date.date(),
            yesterday_end_date.date(),
            round(data["grossRevenue"], 2),
            round(data["netProfit"], 2),
            data["unitsSold"],
        ])
    for col_num, col in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"Product_Performance_{yesterday_start_date.strftime('%Y-%m-%d')}_{action or 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
@csrf_exempt
def downloadProductPerformanceCSV(request):
    try:
        json_request = JSONParser().parse(request)
        action = json_request.get("action", "").lower()
        marketplace_id = json_request.get('marketplace_id', None)
        brand_id = json_request.get('brand_id', [])
        product_id = json_request.get('product_id', [])
        manufacturer_name = json_request.get('manufacturer_name', [])
        fulfillment_channel = json_request.get('fulfillment_channel', None)
        preset = json_request.get('preset')
        timezone_str = 'US/Pacific'
        local_tz = pytz.timezone(timezone_str)
        today = datetime.now(local_tz)
        logger.info(f"Action: {action}")
        logger.info(f"Marketplace ID: {marketplace_id}")
        logger.info(f"Brand ID: {brand_id}")
        logger.info(f"Product ID: {product_id}")
        yesterday_start_date = today - timedelta(days=1)
        yesterday_start_date = yesterday_start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        yesterday_end_date = yesterday_start_date.replace(hour=23, minute=59, second=59)
        previous_day_start_date = yesterday_start_date - timedelta(days=1)
        previous_day_start_date = previous_day_start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        previous_day_end_date = previous_day_start_date.replace(hour=23, minute=59, second=59)
        logger.info(f"Yesterday range: {yesterday_start_date} to {yesterday_end_date}")
        logger.info(f"Previous day range: {previous_day_start_date} to {previous_day_end_date}")
        def fetch_data(start_date, end_date):
            try:
                result = grossRevenue(start_date, end_date, marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel, timezone_str)
                logger.info(f"Fetched data for {start_date.date()}: {len(result) if result else 0} records")
                return result
            except Exception as e:
                logger.error(f"Error fetching data for {start_date.date()}: {e}")
                return []
        with ThreadPoolExecutor(max_workers=2) as executor:
            future_prev_data = executor.submit(fetch_data, previous_day_start_date, previous_day_end_date)
            future_yes_data = executor.submit(fetch_data, yesterday_start_date, yesterday_end_date)
            prev_data = future_prev_data.result()
            yes_data = future_yes_data.result()
        logger.info(f"Raw data - Yesterday: {len(yes_data) if yes_data else 0}, Previous: {len(prev_data) if prev_data else 0}")
        with ThreadPoolExecutor(max_workers=2) as executor:
            future_yes_data = executor.submit(sales, yes_data)
            future_prev_data = executor.submit(sales, prev_data)
            yes_data = future_yes_data.result()
            prev_data = future_prev_data.result()
        logger.info(f"Processed sales data - Yesterday: {len(yes_data) if yes_data else 0}, Previous: {len(prev_data) if prev_data else 0}")
        data = get_top_movers(yes_data, prev_data)
        logger.info(f"Top movers result: {data}")
        if action == "top":
            limited_summary = data.get('top_3_products', []) if data else []
        elif action == "least":
            limited_summary = data.get('least_3_products', []) if data else []
        limited_summary = []
        if action == "top":
            limited_summary = data.get('top_3_products', []) if data else []
        elif action == "least":
            limited_summary = data.get('least_3_products', []) if data else []
        else:
            logger.warning(f"Invalid action parameter: {action}. Expected 'top' or 'least'")
            limited_summary = []
        logger.info(f"Limited summary count: {len(limited_summary)}")
        if not limited_summary:
            logger.warning("No data found for CSV export")
            if not data:
                logger.error("get_top_movers returned empty data")
            elif action not in ["top", "least"]:
                logger.error(f"Invalid action: {action}")
            else:
                logger.error(f"No data found for action: {action}")
        response = HttpResponse(content_type='text/csv')
        filename = f"Product_Performance_{yesterday_start_date.strftime('%Y-%m-%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow([
            "Product Title", "ASIN", "SKU", "Fulfillment Type", "Marketplace", 
            "Start Date", "End Date", "Gross Revenue", "Net Profit", "Units Sold", "Trend"
        ])
        row_count = 0
        for item in limited_summary:
            try:
                writer.writerow([
                    item.get("product_name", ""),
                    item.get("asin", ""),
                    item.get("sku", ""),
                    item.get("fulfillment_channel", ""),
                    item.get("m_name", ""),
                    yesterday_start_date.date(),
                    yesterday_end_date.date(),
                    round(float(item.get("grossRevenue", 0)), 2),
                    round(float(item.get("netProfit", 0)), 2),
                    int(item.get("unitsSold", 0)),
                    item.get("Trend", ""),
                ])
                row_count += 1
            except Exception as e:
                logger.error(f"Error writing row: {e}, Data: {item}")
                writer.writerow([
                    "ERROR", "ERROR", "ERROR", "ERROR", "ERROR",
                    yesterday_start_date.date(), yesterday_end_date.date(),
                    0.00, 0.00, 0, "ERROR"
                ])
        logger.info(f"Successfully wrote {row_count} rows to CSV")
        return response
    except Exception as e:
        logger.error(f"Error in downloadProductPerformanceCSV: {e}")
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="error.csv"'
        writer = csv.writer(response)
        writer.writerow(["Error", "Message"])
        writer.writerow(["Error occurred", str(e)])
        return response
@csrf_exempt
def CityCSVUploadView(request):
    if request.method != 'POST':
        return JsonResponse({"error": "Only POST method allowed"}, status=405)
    file = request.FILES.get('file')
    if not file:
        return JsonResponse({"error": "No file uploaded"}, status=400)
    try:
        try:
            decoded_file = file.read().decode('utf-8')
        except UnicodeDecodeError:
            decoded_file = file.read().decode('latin1')
    except Exception as e:
        return JsonResponse({"error": f"File decoding failed: {str(e)}"}, status=400)
    io_string = io.StringIO(decoded_file)
    reader = csv.DictReader(io_string)
    city_objects = []
    for row in reader:
        try:
            city_obj = CityDetails(
                city=row['city'],
                city_ascii=row['city_ascii'],
                state_id=row['state_id'],
                state_name=row['state_name'],
                county_fips=row['county_fips'],
                county_name=row['county_name'],
                lat=float(row['lat']),
                lng=float(row['lng']),
                population=int(row['population']),
                density=float(row['density']),
                source=row['source'],
                military=row['military'].strip().upper() == 'TRUE',
                incorporated=row['incorporated'].strip().upper() == 'TRUE',
                timezone=row['timezone'],
                ranking=int(row['ranking']),
                zips=row['zips'],
                uid=int(row['id'])
            )
            city_objects.append(city_obj)
        except Exception as e:
            return JsonResponse({"error": f"Error parsing row: {str(e)}"}, status=400)
    try:
        CityDetails.objects.insert(city_objects, load_bulk=False)
        return JsonResponse({"message": "CSV data uploaded successfully."})
    except Exception as e:
        return JsonResponse({"error": f"Bulk insert failed: {str(e)}"}, status=500)
@csrf_exempt
def getCitywiseSales(request):
    json_request = JSONParser().parse(request)
    level = json_request.get("level", "city").lower()  
    action = json_request.get("action", "all").lower()  
    preset = json_request.get("preset", "Yesterday")
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id',[])
    manufacturer_name = json_request.get('manufacturer_name',[])
    fulfillment_channel = json_request.get('fulfillment_channel',None)
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date != None and start_date != "":
        yesterday_start = datetime.strptime(start_date, '%Y-%m-%d')
        yesterday_end = datetime.strptime(end_date, '%Y-%m-%d')
    else:
        yesterday_start, yesterday_end = get_date_range(preset)
    orders = grossRevenue(yesterday_start, yesterday_end,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel)
    grouped_data = defaultdict(lambda: {"units": 0, "gross": 0.0, "city": "", "state": "", "country": ""})
    for entry in orders:
        address = entry.get("shipping_address") or entry.get("shipping_information", {}).get("postalAddress", {})
        city = address.get("city") or address.get("City")
        state = address.get("state") or address.get("StateOrRegion")
        country =  "USA"
        if level == "city" and city and state and country:
            key = f"{city}|{state}|{country}"
        elif level == "state" and state and country:
            key = f"{state}|{country}"
        elif level == "country" and country:
            key = "USA"
        else:
            continue
        grouped_data[key]["units"] += 1
        grouped_data[key]["gross"] += entry.get("order_total", 0.0)
        grouped_data[key]["city"] = city or ""
        grouped_data[key]["state"] = state or ""
        grouped_data[key]["country"] = "USA"
    geo_lookup = {}
    state_population = defaultdict(int)
    country_population = defaultdict(int)
    if level in ["city", "state", "country"]:
        geo_data = CityDetails.objects.filter()
        for geo in geo_data:
            geo_lookup[geo.city] = geo
            if geo.population:
                if level in ["state"]:
                    key = f"{geo.state_id}|USA"
                    state_population[key] += geo.population
                if level == "country":
                    country_population['USA'] += geo.population
    items = []
    for key, data_ in grouped_data.items():
        city = data_["city"]
        geo = geo_lookup.get(city) if level == "city" else None
        item = {
            "units": data_["units"],
            "gross": round(data_["gross"], 2),
            "country": data_["country"],
            "state_name": ""
        }
        if level in ["city", "state"]:
            item["state"] = data_["state"]
        if level == "city":
            item["city"] = city
            if geo:
                item["lat"] = geo.lat
                item["lon"] = geo.lng
                item["code"] = ""
                item["fips"] = ""
                item["population"] = geo.population
                item["state_name"] = geo.state_name
            else:
                item["lat"] = None
                item["lon"] = None
                item["code"] = ""
                item["fips"] = ""
                item["population"] = 1
                item["state_name"] = ""
        elif level == "state":
            geo_data = CityDetails.objects.filter(state_id=data_["state"]).first()
            item["lat"] = None
            item["lon"] = None
            item["code"] = ""
            item["fips"] = ""
            if geo_data:
                item["state_name"] = geo_data.state_name
            state_key = f"{data_['state']}|USA"
            item["population"] = state_population.get(state_key, 1)
        elif level == "country":
            item["lat"] = None
            item["lon"] = None
            item["code"] = ""
            item["fips"] = ""
            item["population"] = country_population.get('USA', 1)
        if action != "all":
            print(item["population"])
            item['units'] = item['units'] / item["population"] if item["population"] >0 else 1
            item['gross'] = item['gross'] / item["population"] if item["population"] >0 else 1
        items.append(item)
    return JsonResponse({"items": items}, safe=False)
@csrf_exempt
def exportCitywiseSalesExcel(request):
    json_request = JSONParser().parse(request)
    level = json_request.get("level", "city").lower()  
    action = json_request.get("action", "all").lower()  
    preset = json_request.get("preset", "Yesterday")
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id',[])
    manufacturer_name = json_request.get('manufacturer_name',[])
    fulfillment_channel = json_request.get('fulfillment_channel',None)
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date != None and start_date != "":
        yesterday_start = datetime.strptime(start_date, '%Y-%m-%d')
        yesterday_end = datetime.strptime(end_date, '%Y-%m-%d')
    else:
        yesterday_start, yesterday_end = get_date_range(preset)
    orders = grossRevenue(yesterday_start, yesterday_end,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel)
    grouped_data = defaultdict(lambda: {"units": 0, "gross": 0.0, "city": "", "state": "", "country": ""})
    for entry in orders:
        shipping = entry.get("shipping_address") or entry.get("shipping_information", {}).get("postalAddress", {})
        city = shipping.get("city")
        state = shipping.get("state") or shipping.get("StateOrRegion")
        country = shipping.get("country") or shipping.get("CountryCode")
        if level == "city" and city and state and country:
            key = f"{city}|{state}|{country}"
        elif level == "state" and state and country:
            key = f"{state}|{country}"
        elif level == "country" and country:
            key = f"{country}"
        else:
            continue  
        grouped_data[key]["units"] += 1
        grouped_data[key]["gross"] += entry.get("order_total", 0.0)
        grouped_data[key]["city"] = city or ""
        grouped_data[key]["state"] = state or ""
        grouped_data[key]["country"] = country or ""
    geo_lookup = {}
    city_population = defaultdict(int)
    state_population = defaultdict(int)
    country_population = defaultdict(int)
    if level in ["city", "state", "country"]:
        geo_data = CityDetails.objects.filter()
        for geo in geo_data:
            geo_lookup[geo.city] = geo
            if geo.population:
                if level in ["city"]:
                    key = f"{geo.city}|{geo.state_id}|USA"
                    city_population[key] += geo.population
                if level in ["state"]:
                    key = f"{geo.state_id}|USA"
                    state_population[key] += geo.population
                if level == "country":
                    country_population['USA'] += geo.population
    data_rows = []
    headers = ["Date From", "Date To", "Country", "Gross Revenue", "Units Sold"]
    for key, values in grouped_data.items():
        row = [yesterday_end.strftime("%b %d, %Y"), yesterday_start.strftime("%b %d, %Y")]
        if level == "city":
            row.extend([values["country"], values["state"], values["city"]])
            headers = ["Date From", "Date To", "Country", "State", "City", "Gross Revenue", "Units Sold"]
        elif level == "state":
            row.extend([values["country"], values["state"]])
            headers = ["Date From", "Date To", "Country", "State", "Gross Revenue", "Units Sold"]
        else:  
            row.append(values["country"])
            headers = ["Date From", "Date To", "Country", "Gross Revenue", "Units Sold"]
        if action == 'all':
            row.extend([round(values["gross"], 2), values["units"]])
        else:
            if level == "city":
                row.extend([(values["gross"]/city_population.get(values["city"], 1)), values["units"]])
            elif level == "state":
                row.extend([(values["gross"]/state_population.get(values["state"], 1)), values["units"]])
            else:
                row.extend([(values["gross"]/country_population.get("USA", 1)), values["units"]])
        data_rows.append(row)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    for row in data_rows:
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{level.capitalize()}wiseSales.xlsx"'
    wb.save(response)
    return response
@csrf_exempt
def downloadCitywiseSalesCSV(request):
    json_request = JSONParser().parse(request)
    level = json_request.get("level", "city").lower()  
    action = json_request.get("action", "all").lower()  
    preset = json_request.get("preset", "Yesterday")
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id',[])
    manufacturer_name = json_request.get('manufacturer_name',[])
    fulfillment_channel = json_request.get('fulfillment_channel',None)
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date != None and start_date != "":
        yesterday_start = datetime.strptime(start_date, '%Y-%m-%d')
        yesterday_end = datetime.strptime(end_date, '%Y-%m-%d')
    else:
        yesterday_start, yesterday_end = get_date_range(preset)
    orders = grossRevenue(yesterday_start, yesterday_end,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel)
    grouped_data = defaultdict(lambda: {"units": 0, "gross": 0.0, "city": "", "state": "", "country": ""})
    for entry in orders:
        shipping = entry.get("shipping_address") or entry.get("shipping_information", {}).get("postalAddress", {})
        city = shipping.get("city")
        state = shipping.get("state") or shipping.get("StateOrRegion")
        country = shipping.get("country") or shipping.get("CountryCode")
        if level == "city" and city and state and country:
            key = f"{city}|{state}|{country}"
        elif level == "state" and state and country:
            key = f"{state}|{country}"
        elif level == "country" and country:
            key = f"{country}"
        else:
            continue
        grouped_data[key]["units"] += 1
        grouped_data[key]["gross"] += entry.get("order_total", 0.0)
        grouped_data[key]["city"] = city or ""
        grouped_data[key]["state"] = state or ""
        grouped_data[key]["country"] = country or ""
    geo_lookup = {}
    city_population = {}
    state_population = defaultdict(int)  
    country_population = defaultdict(int)
    geo_data = CityDetails.objects.all()
    for geo in geo_data:
        geo_lookup[geo.city] = geo
        if geo.population:
            city_key = f"{geo.city}|{geo.state_id}|USA"
            state_key = f"{geo.state_id}|USA"
            city_population[city_key] = geo.population
            state_population[state_key] += geo.population
            country_population["USA"] += geo.population
    data_rows = []
    for key, values in grouped_data.items():
        row = [yesterday_end.strftime("%b %d, %Y"), yesterday_start.strftime("%b %d, %Y")]
        if level == "city":
            row.extend([values["country"], values["state"], values["city"]])
            headers = ["Date From", "Date To", "Country", "State", "City", "Gross Revenue", "Units Sold"]
            pop_key = f"{values['city']}|{values['state']}|{values['country']}"
            population = city_population.get(pop_key, 1)
        elif level == "state":
            row.extend([values["country"], values["state"]])
            headers = ["Date From", "Date To", "Country", "State", "Gross Revenue", "Units Sold"]
            pop_key = f"{values['state']}|{values['country']}"
            population = state_population.get(pop_key, 1)
        else:  
            row.append(values["country"])
            headers = ["Date From", "Date To", "Country", "Gross Revenue", "Units Sold"]
            population = country_population.get(values['country'], 1)
        if action == 'all':
            row.extend([round(values["gross"], 2), values["units"]])
        else:
            per_capita = round(values["gross"] / population, 4)
            u_p = round(values["units"]/ population,4)
            row.extend([per_capita, u_p])
            headers = headers[:len(headers) - 2] + ["Per Capita Revenue", "Units Sold"]
        data_rows.append(row)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{level.capitalize()}wiseSales.csv"'
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in data_rows:
        writer.writerow(row)
    return response
def generate_monthly_intervals(from_date, to_date):
    intervals = []
    current_date = from_date.replace(day=1)
    while current_date <= to_date:
        intervals.append(current_date.strftime('%Y-%m-%d 00:00:00'))
        if current_date.month == 12:
            current_date = current_date.replace(year=current_date.year + 1, month=1)
        else:
            current_date = current_date.replace(month=current_date.month + 1)
    print(intervals)
    return intervals
def calculate_metrics(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone):
    gross_revenue = 0
    total_cogs = 0
    refund = 0
    net_profit = 0
    margin = 0
    total_units = 0
    sku_set = set()
    product_categories = {}
    product_completeness = {"complete": 0, "incomplete": 0}
    result = grossRevenue(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone)
    order_total = 0
    tax_price = 0
    temp_price = 0
    vendor_funding = 0
    if result:
        for order in result:
            gross_revenue += order['order_total']
            order_total = order['order_total']
            total_units += order['items_order_quantity']
            tax_price = 0
            for item_id in order['order_items']:
                item_pipeline = [
                    { "$match": { "_id": item_id } },
                    {
                        "$lookup": {
                            "from": "product",
                            "localField": "ProductDetails.product_id",
                            "foreignField": "_id",
                            "as": "product_ins"
                        }
                    },
                    { "$unwind": { "path": "$product_ins", "preserveNullAndEmptyArrays": True } },
                    {
                        "$project": {
                            "_id": 0,
                            "price": "$Pricing.ItemPrice.Amount",
                            "tax_price": "$Pricing.ItemTax.Amount",
                            "cogs": { "$ifNull": ["$product_ins.cogs", 0.0] },
                            "sku": "$product_ins.sku",
                            "category": "$product_ins.category",
                            "total_cogs" : {"$ifNull":["$product_ins.total_cogs",0]},
                            "w_total_cogs" : {"$ifNull":["$product_ins.w_total_cogs",0]},
                            "vendor_funding" : {"$ifNull":["$product_ins.vendor_funding",0]},
                        }
                    }
                ]
                item_result = list(OrderItems.objects.aggregate(*item_pipeline))
                if item_result:
                    item_data = item_result[0]
                    temp_price += item_data['price']
                    tax_price += item_data['tax_price']
                    if order['marketplace_name'] == "Amazon":
                        total_cogs += item_data['total_cogs'] 
                    else:
                        total_cogs += item_data['w_total_cogs']
                    vendor_funding += item_data['vendor_funding']
                    if item_data.get('sku'):
                        sku_set.add(item_data['sku'])
                    category = item_data.get('category', 'Unknown')
                    if category in product_categories:
                        product_categories[category] += 1
                    else:
                        product_categories[category] = 1
                    if item_data['price'] and item_data['cogs'] and item_data['sku']:
                        product_completeness["complete"] += 1
                    else:
                        product_completeness["incomplete"] += 1
        net_profit = (temp_price -  total_cogs) + vendor_funding
        margin = (net_profit / gross_revenue) * 100 if gross_revenue > 0 else 0
    return {
        "grossRevenue": round(gross_revenue, 2),
        "expenses": round(total_cogs, 2),
        "netProfit": round(net_profit, 2),
        "roi": round((net_profit / (total_cogs)) * 100, 2) if total_cogs > 0 else 0,
        "unitsSold": total_units,
        "refunds": refund,  
        "skuCount": len(sku_set),
        "sessions": 0,
        "pageViews": 0,
        "unitSessionPercentage": 0,
        "margin": round(margin, 2),
        "seller": "",
        "tax_price": tax_price,
        "total_cogs": total_cogs,
        "product_cost": order_total,
        "shipping_cost": 0,
        "productCategories": product_categories,  
        "productCompleteness": product_completeness  
    }
@csrf_exempt
def getProfitAndLossDetails(request):
    def to_utc_format(dt):
        return dt.strftime("%Y-%m-%dT%H:%M:%SZ")

    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id', [])
    manufacturer_name = json_request.get('manufacturer_name', [])
    fulfillment_channel = json_request.get('fulfillment_channel', None)
    preset = json_request.get('preset')
    timezone = 'US/Pacific'
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)

    if start_date:
        from_date, to_date = convertdateTotimezone(start_date, end_date, timezone)
    else:
        from_date, to_date = get_date_range(preset, timezone)

    def pLcalculate_metrics(start_date, end_date, marketplace_id, brand_id, product_id,
                            manufacturer_name, fulfillment_channel, timezone):
        gross_revenue = total_cogs = refund = net_profit = margin = total_units = 0
        shipping_cost = channel_fee = product_cost = vendor_funding = tax_price = temp_price = 0
        sku_set = set()
        product_categories = {}
        product_completeness = {"complete": 0, "incomplete": 0}

        result = grossRevenue(start_date, end_date, marketplace_id, brand_id, product_id,
                              manufacturer_name, fulfillment_channel, timezone)
        all_item_ids = []
        for order in result:
            all_item_ids.extend(order['order_items'])

        item_pipeline = [
            {"$match": {"_id": {"$in": all_item_ids}}},
            {"$lookup": {
                "from": "product",
                "localField": "ProductDetails.product_id",
                "foreignField": "_id",
                "as": "product_ins"
            }},
            {"$unwind": {"path": "$product_ins", "preserveNullAndEmptyArrays": True}},
            {"$project": {
                "_id": 1,
                "price": "$Pricing.ItemPrice.Amount",
                "tax_price": "$Pricing.ItemTax.Amount",
                "cogs": {"$ifNull": ["$product_ins.cogs", 0.0]},
                "sku": "$product_ins.sku",
                "category": "$product_ins.category",
                "total_cogs": {"$ifNull": ["$product_ins.total_cogs", 0]},
                "w_total_cogs": {"$ifNull": ["$product_ins.w_total_cogs", 0]},
                "vendor_funding": {"$ifNull": ["$product_ins.vendor_funding", 0]},
                "a_shipping_cost": {"$ifNull": ["$product_ins.a_shipping_cost", 0]},
                "w_shiping_cost": {"$ifNull": ["$product_ins.w_shiping_cost", 0]},
                "referral_fee": {"$ifNull": ["$product_ins.referral_fee", 0]},
                "walmart_fee": {"$ifNull": ["$product_ins.walmart_fee", 0]},
                "product_cost": {"$ifNull": ["$product_ins.product_cost", 0]},
                "w_product_cost": {"$ifNull": ["$product_ins.w_product_cost", 0]}
            }}
        ]
        item_results = list(OrderItems.objects.aggregate(*item_pipeline))
        item_lookup = {item['_id']: item for item in item_results}

        for order in result:
            gross_revenue += order['order_total']
            total_units += order['items_order_quantity']
            shipping_cost += order.get('shipping_price', 0) or 0 #added shipping cost

            for item_id in order['order_items']:
                item_data = item_lookup.get(item_id)
                if not item_data:
                    continue
                temp_price += item_data['price']
                tax_price += item_data['tax_price']
                if order['marketplace_name'] == "Amazon":
                    total_cogs += item_data['total_cogs']
                    channel_fee += item_data['referral_fee']
                    product_cost += item_data['product_cost']
                else:
                    total_cogs += item_data['w_total_cogs']
                    channel_fee += item_data['walmart_fee']
                    product_cost += item_data['w_product_cost']
                vendor_funding += item_data['vendor_funding']
                sku_set.add(item_data.get('sku'))
                category = item_data.get('category', 'Unknown')
                product_categories[category] = product_categories.get(category, 0) + 1
                if item_data['price'] and item_data['total_cogs'] and item_data.get('sku'):
                    product_completeness["complete"] += 1
                else:
                    product_completeness["incomplete"] += 1

        net_profit = (temp_price - total_cogs) + vendor_funding
        margin = (net_profit / gross_revenue) * 100 if gross_revenue else 0

        return {
            "grossRevenue": round(gross_revenue, 2),
            "expenses": round(total_cogs, 2),
            "netProfit": round(net_profit, 2),
            "roi": round((net_profit / total_cogs) * 100, 2) if total_cogs else 0,
            "unitsSold": total_units,
            "refunds": refund,
            "skuCount": len(sku_set),
            "sessions": 0,
            "pageViews": 0,
            "unitSessionPercentage": 0,
            "margin": round(margin, 2),
            "seller": "",
            "tax_price": tax_price,
            "total_cogs": total_cogs,
            "product_cost": product_cost,
            "shipping_cost": shipping_cost,
            "productCategories": product_categories,
            "productCompleteness": product_completeness,
            "base_price": temp_price,
            "channel_fee": channel_fee
        }

    def create_period_response(label, cur_from, cur_to, prev_from, prev_to,
                               marketplace_id, brand_id, product_id,
                               manufacturer_name, fulfillment_channel, preset, timezone):
        current = pLcalculate_metrics(cur_from, cur_to, marketplace_id, brand_id, product_id,
                                      manufacturer_name, fulfillment_channel, timezone)
        previous = pLcalculate_metrics(prev_from, prev_to, marketplace_id, brand_id, product_id,
                                       manufacturer_name, fulfillment_channel, timezone)
        def with_delta(metric):
            return {
                "current": current[metric],
                "previous": previous[metric],
                "delta": round(current[metric] - previous[metric], 2)
            }
        summary = {metric: with_delta(metric) for metric in [
            "grossRevenue", "netProfit", "expenses", "unitsSold",
            "refunds", "skuCount", "sessions", "pageViews",
            "unitSessionPercentage", "margin", "roi", "channel_fee"]}
        netProfitCalculation = {
            "current": {
                "gross": current["grossRevenue"],
                "totalCosts": current["expenses"],
                "productRefunds": current["refunds"],
                "totalTax": current["tax_price"],
                "totalTaxWithheld": 0,
                "ppcProductCost": 0,
                "ppcBrandsCost": 0,
                "ppcDisplayCost": 0,
                "ppcStCost": 0,
                "cogs": current["total_cogs"],
                "product_cost": current["product_cost"],
                "base_price": current["base_price"],
                "shipping_cost": current["shipping_cost"],
                "channel_fee": current["channel_fee"]
            },
            "previous": {
                "gross": previous["grossRevenue"],
                "totalCosts": previous["expenses"],
                "productRefunds": previous["refunds"],
                "totalTax": previous["total_cogs"],
                "totalTaxWithheld": 0,
                "ppcProductCost": 0,
                "ppcBrandsCost": 0,
                "ppcDisplayCost": 0,
                "ppcStCost": 0,
                "cogs": previous["total_cogs"],
                "product_cost": previous["product_cost"],
                "base_price": current["base_price"],
                "shipping_cost": previous["shipping_cost"],
                "channel_fee": previous["channel_fee"]
            }
        }
        date_ranges = {
            "current": {"from": to_utc_format(cur_from)},
            "previous": {"from": to_utc_format(prev_from)}
        }
        if preset not in ['Today', 'Yesterday']:
            date_ranges["current"]["to"] = to_utc_format(cur_to)
            date_ranges["previous"]["to"] = to_utc_format(prev_to)
        return {
            "dateRanges": date_ranges,
            "summary": summary,
            "netProfitCalculation": netProfitCalculation,
            "charts": {
                "productDistribution": current["productCategories"],
                "productCompleteness": current["productCompleteness"]
            }
        }

    custom_duration = to_date - from_date
    prev_from_date = from_date - custom_duration
    prev_to_date = to_date - custom_duration
    response_data = {
        "custom": create_period_response("Custom", from_date, to_date, prev_from_date, prev_to_date,
                                          marketplace_id, brand_id, product_id, manufacturer_name,
                                          fulfillment_channel, preset, timezone)
    }
    return JsonResponse(response_data, safe=False)
@csrf_exempt
def profit_loss_chart(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id',[])
    manufacturer_name = json_request.get('manufacturer_name',[])
    fulfillment_channel = json_request.get('fulfillment_channel',None)
    preset = json_request.get('preset')
    timezone = 'US/Pacific'
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    
    if start_date != None and start_date != "":
        from_date, to_date = convertdateTotimezone(start_date,end_date,timezone)
    else:
        from_date, to_date = get_date_range(preset,timezone)
    
    def get_month_range(year, month):
        start_date = datetime(year, month, 1)
        last_day = monthrange(year, month)[1]
        end_date = datetime(year, month, last_day, 23, 59, 59)
        return start_date, end_date
    
    def calculate_metrics_optimized(start_date, end_date, marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel, timezone):
        # Get all orders for the date range
        result = grossRevenue(start_date, end_date, marketplace_id, brand_id, product_id, manufacturer_name, fulfillment_channel, timezone)
        
        if not result:
            return {
                "grossRevenue": 0,
                "expenses": 0,
                "netProfit": 0,
                "roi": 0,
                "unitsSold": 0,
                "refunds": 0,
                "skuCount": 0,
                "margin": 0,
                "tax_price": 0,
                "total_cogs": 0,
                "product_cost": 0,
                "productCategories": {},
                "productCompleteness": {"complete": 0, "incomplete": 0}
            }
        
        # Collect all order item IDs in one go
        all_item_ids = []
        order_data = {}
        
        for order in result:
            order_total = order.get("order_total", 0)
            marketplace_name = order.get('marketplace_name', '')
            units = order.get('items_order_quantity', 0)
            
            order_data[order.get('_id')] = {
                'order_total': order_total,
                'marketplace_name': marketplace_name,
                'units': units
            }
            
            all_item_ids.extend(order.get("order_items", []))
        
        # Single aggregation query to get all item data at once
        item_pipeline = [
            {"$match": {"_id": {"$in": all_item_ids}}},
            {
                "$lookup": {
                    "from": "product",
                    "localField": "ProductDetails.product_id",
                    "foreignField": "_id",
                    "as": "product_ins"
                }
            },
            {"$unwind": {"path": "$product_ins", "preserveNullAndEmptyArrays": True}},
            {
                "$project": {
                    "_id": 1,
                    "price": "$Pricing.ItemPrice.Amount",
                    "tax_price": "$Pricing.ItemTax.Amount",
                    "cogs": {"$ifNull": ["$product_ins.cogs", 0.0]},
                    "sku": "$product_ins.sku",
                    "category": "$product_ins.category",
                    "total_cogs": {"$ifNull": ["$product_ins.total_cogs", 0]},
                    "w_total_cogs": {"$ifNull": ["$product_ins.w_total_cogs", 0]},
                    "vendor_funding": {"$ifNull": ["$product_ins.vendor_funding", 0]},
                }
            }
        ]
        
        # Execute single query for all items
        all_items = list(OrderItems.objects.aggregate(*item_pipeline))
        item_lookup = {str(item['_id']): item for item in all_items}
        
        # Initialize counters
        gross_revenue_amt = 0
        total_cogs = 0
        temp_price = 0
        tax_price = 0
        vendor_funding = 0
        total_units = 0
        sku_set = set()
        product_categories = {}
        product_completeness = {"complete": 0, "incomplete": 0}
        
        # Process orders using pre-fetched data
        for order in result:
            gross_revenue_amt += order.get("order_total", 0)
            total_units += order.get('items_order_quantity', 0)
            marketplace_name = order.get('marketplace_name', '')
            
            for item_id in order.get("order_items", []):
                item = item_lookup.get(str(item_id))
                if item:
                    temp_price += item.get("price", 0)
                    tax_price += item.get("tax_price", 0)
                    
                    if marketplace_name == "Amazon":
                        total_cogs += item.get("total_cogs", 0)
                    else:
                        total_cogs += item.get("w_total_cogs", 0)
                    
                    vendor_funding += item.get("vendor_funding", 0)
                    
                    sku = item.get("sku")
                    if sku:
                        sku_set.add(sku)
                    
                    category = item.get("category", "Unknown")
                    product_categories[category] = product_categories.get(category, 0) + 1
                    
                    if item.get("price") and item.get("total_cogs") and sku:
                        product_completeness["complete"] += 1
                    else:
                        product_completeness["incomplete"] += 1
        
        net_profit = (temp_price - total_cogs) + vendor_funding
        margin = (net_profit / gross_revenue_amt * 100) if gross_revenue_amt else 0
        
        return {
            "grossRevenue": round(gross_revenue_amt, 2),
            "expenses": round(total_cogs, 2),
            "netProfit": round(net_profit, 2),
            "roi": round((net_profit / total_cogs) * 100, 2) if total_cogs else 0,
            "unitsSold": total_units,
            "refunds": 0,  # refund was always 0 in original code
            "skuCount": len(sku_set),
            "margin": round(margin, 2),
            "tax_price": tax_price,
            "total_cogs": total_cogs,
            "product_cost": gross_revenue_amt,  # using gross_revenue_amt as it was order_total
            "productCategories": product_categories,
            "productCompleteness": product_completeness
        }

    def generate_month_keys(start_year, start_month, end_year, end_month):
        months = []
        current = datetime(start_year, start_month, 1)
        end = datetime(end_year, end_month, 1)
        while current <= end:
            months.append(current.strftime("%Y-%m-%d 00:00:00"))
            current += timedelta(days=32)
            current = current.replace(day=1)
        return months

    # Pre-calculate interval configuration
    metrics = ["grossRevenue", "estimatedPayout", "expenses", "netProfit", "units", "ppcSales"]
    values = {metric: {} for metric in metrics}
    
    hourly_presets = ["Today", "Yesterday"]
    daily_presets = ["This Week", "Last Week", "Last 7 days", "Last 14 days", "Last 30 days", 
                    "Last 60 days", "Last 90 days", "Last Month", "This Quarter", "Last Quarter", "Last Year"]

    # --- INTERVAL LOGIC STARTS HERE ---
    from pytz import timezone as pytz_timezone
    pacific_tz = pytz_timezone('US/Pacific')
    current_pacific_time = datetime.now(pacific_tz)

    # Always return hourly intervals if start_date and end_date are the same day
    if start_date and end_date and start_date[:10] == end_date[:10]:
        total_hours = int((to_date - from_date).total_seconds() // 3600) + 1
        interval_keys = []
        for i in range(total_hours):
            interval_time = from_date + timedelta(hours=i)
            if interval_time.tzinfo is None:
                interval_time_pacific = pacific_tz.localize(interval_time)
            else:
                interval_time_pacific = interval_time.astimezone(pacific_tz)
            if interval_time_pacific.replace(minute=0, second=0, microsecond=0) <= current_pacific_time.replace(minute=0, second=0, microsecond=0):
                interval_keys.append(interval_time.strftime("%Y-%m-%d %H:00:00"))
            else:
                break
        interval_type = "hour"
    elif preset in hourly_presets:
        total_hours = int((to_date - from_date).total_seconds() // 3600) + 1
        interval_keys = []
        for i in range(total_hours):
            interval_time = from_date + timedelta(hours=i)
            if interval_time.tzinfo is None:
                interval_time_pacific = pacific_tz.localize(interval_time)
            else:
                interval_time_pacific = interval_time.astimezone(pacific_tz)
            if interval_time_pacific.replace(minute=0, second=0, microsecond=0) <= current_pacific_time.replace(minute=0, second=0, microsecond=0):
                interval_keys.append(interval_time.strftime("%Y-%m-%d %H:00:00"))
            else:
                break
        interval_type = "hour"
    elif preset in daily_presets:
        interval_keys = [(from_date + timedelta(days=i)).strftime("%Y-%m-%d 00:00:00") 
                         for i in range((to_date - from_date).days + 1)]
        interval_type = "day"
    else:
        if start_date and start_date != '':
            interval_keys = [(from_date + timedelta(days=i)).strftime("%Y-%m-%d 00:00:00")
                            for i in range((to_date - from_date).days + 1)]
            interval_type = 'day'   
        else:
            interval_keys = generate_month_keys(
                from_date.year, from_date.month,
                to_date.year, to_date.month
            )
            interval_type = "month"
    # --- INTERVAL LOGIC ENDS HERE ---

    # Process intervals with optimized function
    for key in interval_keys:
        if interval_type == "hour":
            start = datetime.strptime(key, "%Y-%m-%d %H:00:00")
            end = start + timedelta(hours=1) - timedelta(seconds=1)
        elif interval_type == "day":
            start = datetime.strptime(key, "%Y-%m-%d 00:00:00")
            end = start + timedelta(days=1) - timedelta(seconds=1)
        else:  # month
            year, month = int(key[:4]), int(key[5:7])
            start, end = get_month_range(year, month)

        data = calculate_metrics_optimized(start, end, marketplace_id, brand_id, product_id, 
                                         manufacturer_name, fulfillment_channel, timezone)
        
        values["grossRevenue"][key] = data["grossRevenue"]
        values["expenses"][key] = data["expenses"]
        values["netProfit"][key] = data["netProfit"]
        values["units"][key] = data["unitsSold"]

    # Ensure all metrics have all keys (fill missing with 0)
    for metric in metrics:
        for key in interval_keys:
            values[metric].setdefault(key, 0)

    graph = [{"metric": metric, "values": values[metric]} for metric in metrics]
    return JsonResponse({"graph": graph}, safe=False)
@csrf_exempt
def profitLossExportXl(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id',[])
    manufacturer_name = json_request.get('manufacturer_name',[])
    fulfillment_channel = json_request.get('fulfillment_channel',None)
    preset = json_request.get('preset', "Last 30 days")
    timezone_str = 'US/Pacific'
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date and start_date != "":
        local_tz = pytz.timezone(timezone_str)
        naive_from_date = datetime.strptime(start_date, '%Y-%m-%d')
        naive_to_date = datetime.strptime(end_date, '%Y-%m-%d')
        localized_from_date = local_tz.localize(naive_from_date)
        localized_to_date = local_tz.localize(naive_to_date)
        from_date = localized_from_date.astimezone(pytz.UTC)
        to_date = localized_to_date.astimezone(pytz.UTC)
        to_date = to_date.replace(hour=23, minute=59, second=59)
    else:
        from_date, to_date = get_date_range(preset, timezone_str)
    def get_month_range(year, month):
        start_date = datetime(year, month, 1)
        last_day = monthrange(year, month)[1]
        end_date = datetime(year, month, last_day, 23, 59, 59)
        return start_date, end_date
    def calculate_metrics(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str):
        gross_revenue_amt = 0
        total_cogs = 0
        refund = 0
        net_profit = 0
        margin = 0
        total_units = 0
        sku_set = set()
        order_total = 0
        tax_price = 0
        temp_price = 0
        vendor_funding = 0
        m_name = ""
        result = grossRevenue(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
        for order in result:
            gross_revenue_amt += order.get("order_total", 0)
            order_total = order.get("order_total", 0)
            total_units +=order['items_order_quantity']
            tax_price = 0
            marketplace_id = order.get("marketplace_id", "")
            Marketplace_obj = Marketplace.objects.filter(id = marketplace_id).first()
            m_name = ""
            if Marketplace_obj:
                m_name = Marketplace_obj.name
            for item_id in order.get("order_items", []):
                item_pipeline = [
                    {"$match": {"_id": item_id}},
                    {
                        "$lookup": {
                            "from": "product",
                            "localField": "ProductDetails.product_id",
                            "foreignField": "_id",
                            "as": "product_ins"
                        }
                    },
                    {"$unwind": {"path": "$product_ins", "preserveNullAndEmptyArrays": True}},
                    {
                        "$project": {
                            "price": "$Pricing.ItemPrice.Amount",
                            "tax_price": "$Pricing.ItemTax.Amount",
                            "cogs": {"$ifNull": ["$product_ins.cogs", 0.0]},
                            "sku": "$product_ins.sku",
                            "total_cogs" : {"$ifNull":["$product_ins.total_cogs",0]},
                            "w_total_cogs" : {"$ifNull":["$product_ins.w_total_cogs",0]},
                            "vendor_funding" : {"$ifNull":["$product_ins.vendor_funding",0]},
                        }
                    }
                ]
                item_result = list(OrderItems.objects.aggregate(*item_pipeline))
                if item_result:
                    item = item_result[0]
                    temp_price += item.get("price", 0)
                    tax_price += item.get("tax_price", 0)
                    if order['marketplace_name'] == "Amazon":
                        total_cogs += item.get("total_cogs", 0) 
                    else:
                        total_cogs += item.get("w_total_cogs", 0)
                    vendor_funding += item.get("vendor_funding", 0)
                    sku = item.get("sku")
                    if sku:
                        sku_set.add(sku)
        net_profit = (temp_price - total_cogs) + vendor_funding
        margin = (net_profit / gross_revenue_amt * 100) if gross_revenue_amt else 0
        return {
            "Marketplace":m_name,
            "Date and Time":start_date,
            "Gross Revenue": round(gross_revenue_amt, 2),
            "Expenses": round((total_cogs) , 2),
            "Estimated Payout":0,
            "Net Profit": round(net_profit, 2),
            "Units Sold": total_units,
            "PPC Sales": 0 
        }
    def generate_month_keys(start_year, start_month, end_year, end_month):
        months = []
        current = datetime(start_year, start_month, 1)
        end = datetime(end_year, end_month, 1)
        while current <= end:
            months.append(current.strftime("%Y-%m-%d"))
            current += timedelta(days=32)
            current = current.replace(day=1)
        return months
    interval_keys = []
    interval_type = ""
    if preset in ["Today", "Yesterday"]:
        interval_keys = [(from_date + timedelta(hours=i)).strftime("%Y-%m-%d %H:00:00")
                         for i in range(int((to_date - from_date).total_seconds() // 3600) + 1)]
        interval_type = "hour"
    elif preset in ["This Week", "Last Week", "Last 7 days", "Last 14 days", "Last 30 days", "Last 60 days", "Last 90 days"]:
        interval_keys = [(from_date + timedelta(days=i)).strftime("%Y-%m-%d")
                         for i in range((to_date - from_date).days + 1)]
        interval_type = "day"
    else:
        interval_keys = generate_month_keys(from_date.year, from_date.month, to_date.year, to_date.month)
        interval_type = "month"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit Loss Report"
    headers = ["Marketplace", "Date and Time", "Gross Revenue", "Expenses", "Estimated Payout", "Net Profit", "Units", "PPC Sales"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
    for key in interval_keys:
        if interval_type == "hour":
            start = datetime.strptime(key, "%Y-%m-%d %H:00:00")
            end = start + timedelta(hours=1) - timedelta(seconds=1)
            time_label = start.strftime("%Y-%m-%d %H:00:00")
        elif interval_type == "day":
            start = datetime.strptime(key, "%Y-%m-%d")
            end = start + timedelta(days=1) - timedelta(seconds=1)
            time_label = start.strftime("%Y-%m-%d")
        else:
            year, month = int(key[:4]), int(key[5:7])
            start, end = get_month_range(year, month)
            time_label = f"{year}-{month:02d}"
        row_data = calculate_metrics(start, end,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
        ws.append([
            row_data.get("Marketplace", ""),
            time_label,
            row_data.get("Gross Revenue", 0),
            row_data.get("Expenses", 0),
            row_data.get("Estimated Payout", 0),
            row_data.get("Net Profit", 0),
            row_data.get("Units Sold", 0),
            row_data.get("PPC Sales", 0),
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=profit_loss_report.xlsx'
    return response
@csrf_exempt
def profitLossChartCsv(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id', None)
    brand_id = json_request.get('brand_id', [])
    product_id = json_request.get('product_id',[])
    manufacturer_name = json_request.get('manufacturer_name',[])
    fulfillment_channel = json_request.get('fulfillment_channel',None)
    preset = json_request.get('preset',"Last 7 days")
    timezone_str =  'US/Pacific'
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date and start_date != "":
        local_tz = pytz.timezone(timezone_str)
        naive_from_date = datetime.strptime(start_date, '%Y-%m-%d')
        naive_to_date = datetime.strptime(end_date, '%Y-%m-%d')
        localized_from_date = local_tz.localize(naive_from_date)
        localized_to_date = local_tz.localize(naive_to_date)
        from_date = localized_from_date.astimezone(pytz.UTC)
        to_date = localized_to_date.astimezone(pytz.UTC)
        to_date = to_date.replace(hour=23, minute=59, second=59)
    else:
        from_date, to_date = get_date_range(preset, timezone_str)
    def get_month_range(year, month):
        from calendar import monthrange
        start_date = datetime(year, month, 1)
        last_day = monthrange(year, month)[1]
        end_date = datetime(year, month, last_day, 23, 59, 59)
        return start_date, end_date
    def generate_month_keys(start_year, start_month, end_year, end_month):
        months = []
        current = datetime(start_year, start_month, 1)
        end = datetime(end_year, end_month, 1)
        while current <= end:
            months.append(current.strftime("%Y-%m-%d 00:00:00"))
            current += timedelta(days=32)
            current = current.replace(day=1)
        return months
    def dummy_calculate_metrics(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str):
        gross_revenue_amt = 0
        total_cogs = 0
        refund = 0
        net_profit = 0
        margin = 0
        total_units = 0
        sku_set = set()
        order_total = 0
        tax_price = 0
        vendor_funding  =0 
        temp_price = 0
        m_name = ""
        result = grossRevenue(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
        for order in result:
            gross_revenue_amt += order.get("order_total", 0)
            order_total = order.get("order_total", 0)
            total_units +=order['items_order_quantity']
            tax_price = 0
            marketplace_id = order.get("marketplace_id", "")
            Marketplace_obj = Marketplace.objects.filter(id = marketplace_id).first()
            m_name = ""
            if Marketplace_obj:
                m_name = Marketplace_obj.name
            for item_id in order.get("order_items", []):
                item_pipeline = [
                    {"$match": {"_id": item_id}},
                    {
                        "$lookup": {
                            "from": "product",
                            "localField": "ProductDetails.product_id",
                            "foreignField": "_id",
                            "as": "product_ins"
                        }
                    },
                    {"$unwind": {"path": "$product_ins", "preserveNullAndEmptyArrays": True}},
                    {
                        "$project": {
                            "price": "$Pricing.ItemPrice.Amount",
                            "tax_price": "$Pricing.ItemTax.Amount",
                            "cogs": {"$ifNull": ["$product_ins.cogs", 0.0]},
                            "sku": "$product_ins.sku",
                            "total_cogs" : {"$ifNull":["$product_ins.total_cogs",0]},
                            "w_total_cogs" : {"$ifNull":["$product_ins.w_total_cogs",0]},
                            "vendor_funding" : {"$ifNull":["$product_ins.vendor_funding",0]},
                        }
                    }
                ]
                item_result = list(OrderItems.objects.aggregate(*item_pipeline))
                if item_result:
                    item = item_result[0]
                    temp_price += item.get("price", 0)
                    tax_price += item.get("tax_price", 0)
                    if order['marketplace_name'] == "Amazon":
                        total_cogs += item.get("total_cogs", 0) 
                    else:
                        total_cogs += item.get("w_total_cogs", 0)
                    vendor_funding += item.get("vendor_funding", 0)
                    sku = item.get("sku")
                    if sku:
                        sku_set.add(sku)
        net_profit = (temp_price - total_cogs) + vendor_funding
        margin = (net_profit / gross_revenue_amt * 100) if gross_revenue_amt else 0
        return {
            "Marketplace":m_name,
            "Date and Time":start_date,
            "Gross Revenue": round(gross_revenue_amt, 2),
            "Expenses": round((total_cogs) , 2),
            "Estimated Payout":0,
            "Net Profit": round(net_profit, 2),
            "Units Sold": total_units,
            "PPC Sales": 0 
        }
    hourly_presets = ["Today", "Yesterday"]
    daily_presets = ["This Week", "Last Week", "Last 7 days", "Last 14 days", "Last 30 days", "Last 60 days", "Last 90 days","Last Month","This Quarter","Last Quarter","Last Year"]
    if preset in hourly_presets:
        interval_keys = [(from_date + timedelta(hours=i)).strftime("%Y-%m-%d %H:00:00")
                         for i in range(0, int((to_date - from_date).total_seconds() // 3600) + 1)]
        interval_type = "hour"
    elif preset in daily_presets:
        interval_keys = [(from_date + timedelta(days=i)).strftime("%Y-%m-%d 00:00:00")
                         for i in range((to_date - from_date).days + 1)]
        interval_type = "day"
    else:
        interval_keys = generate_month_keys(from_date.year, from_date.month, to_date.year, to_date.month)
        interval_type = "month"
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="profit_loss_{preset.replace(" ", "_").lower()}.csv"'
    writer = csv.writer(response)
    writer.writerow(["Marketplace", "Date and Time", "Gross Revenue", "Expenses", "Estimated Payout", "Net Profit", "Units", "PPC Sales"])
    for key in interval_keys:
        if interval_type == "hour":
            start = datetime.strptime(key, "%Y-%m-%d %H:00:00")
            end = start + timedelta(hours=1) - timedelta(seconds=1)
        elif interval_type == "day":
            start = datetime.strptime(key, "%Y-%m-%d 00:00:00")
            end = start + timedelta(days=1) - timedelta(seconds=1)
        else:
            year, month = int(key[:4]), int(key[5:7])
            start, end = get_month_range(year, month)
        data = dummy_calculate_metrics(start, end,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
        writer.writerow([
            data.get("Marketplace", ""),
            data.get("Date and Time", "").strftime("%Y-%m-%d %H:%M:%S") if isinstance(data.get("Date and Time"), datetime) else data.get("Date and Time"),
            data.get("Gross Revenue", 0),
            data.get("Expenses", 0),
            data.get("Estimated Payout", 0),
            data.get("Net Profit", 0),
            data.get("Units Sold", 0),
            data.get("PPC Sales", 0),
        ])
    return response
from rest_framework.parsers import JSONParser 
@csrf_exempt
def updateChooseMatrix(request):
    json_req = JSONParser().parse(request)
    name = json_req['name']
    if name == "Today Snapshot":
        if 'select_all' in json_req and json_req['select_all'] == True:
            update_fields = {
            'select_all': json_req['select_all'], 'gross_revenue': True,
                'total_cogs': True,
                'profit_margin': True,
                'orders': True,
                'units_sold': True,
                'refund_quantity': True,}
        else:
            update_fields = {
                'select_all': False,
            'gross_revenue': json_req['gross_revenue'],
            'total_cogs': json_req['total_cogs'],
            'profit_margin': json_req['profit_margin'],
            'orders': json_req['orders'],
            'units_sold': json_req['units_sold'],
            'refund_quantity': json_req['refund_quantity'],
            }
    elif name == "Revenue":
        if 'select_all' in json_req and json_req['select_all'] == True:
            update_fields = {
                'select_all': json_req['select_all'], 'gross_revenue': True,
                'units_sold': True,
                'acos': True,
                'tacos': True,
                'refund_quantity': True,
                'net_profit': True,
                'profit_margin': True,
                'refund_amount': True,
                'roas': True,
                'orders': True,
                'ppc_spend': True
            }
        else:
            update_fields = {
            'select_all': False,
            'gross_revenue': json_req['gross_revenue'],
            'units_sold': json_req['units_sold'],
            'acos': json_req['acos'],
            'tacos': json_req['tacos'],
            'refund_quantity': json_req['refund_quantity'],
            'net_profit': json_req['net_profit'],
            'profit_margin': json_req['profit_margin'],
            'refund_amount': json_req['refund_amount'],
            'roas': json_req['roas'],
            'orders': json_req['orders'],
            'ppc_spend': json_req['ppc_spend']
            }
    updated_count = chooseMatrix.objects.filter(name=name).update(**update_fields)
    if updated_count == 0:
        return JsonResponse({'status': 'not found', 'message': f'No entry found with name: {name}'}, status=404)
    return JsonResponse({'status': 'success', 'updated_records': updated_count}, status=200)
def createNotes(self, request):
    try:
        data = JSONParser().parse(request)
        product_id = data.get("product_id")
        user_id = data.get("user_id")
        notes = data.get("notes")
        if not product_id or not user_id or not notes:
            return JsonResponse({"error": "Missing required fields."}, status=400)
        try:
            product = Product.objects.get(id=product_id)
            user_obj = user.objects.get(id=user_id)
        except :
            return JsonResponse({"error": "Product or user not found."}, status=404)
        note = notes_data(product_id=product, user_id=user_obj, notes=notes)
        note.save()
        return JsonResponse({"message": "Note added successfully."}, status=201)
    except :
        return JsonResponse({"error": ""}, status=500)
import re
def ListingOptimizationView(request):
    all_products = Product.objects()
    optimized_count = 0
    total_products = all_products.count()
    def is_optimized(product):
        title = product.product_title or ""
        if len(title) < 100 or re.search(r'(?i)(best|free|deal|offer|discount)', title):
            return False
        bullets = product.features or []
        if len(bullets) < 5:
            return False
        if any(re.search(r'<|>|🔥|👍|😁|[A-Z]{4,}', b) for b in bullets):
            return False
        description = product.product_description or ""
        if len(description) <= 300:
            return False
        words = re.findall(r'\b\w+\b', description)
        if len(words) != len(set(words)):
            return False
        images = product.image_urls or []
        if not images:
            return False
        if any(
            not img.endswith(('.jpg', '.jpeg', '.png')) or 'watermark' in img.lower()
            for img in images
        ):
            return False
        upc = product.upc or ""
        if not re.fullmatch(r'\d{12,14}', upc):
            return False
        category = product.category or ""
        if ">" not in category:
            return False
        return True
    for product in all_products:
        if is_optimized(product):
            optimized_count += 1
    return JsonResponse({
        "total_products": total_products,
        "optimized_products": optimized_count,
        "not_optimized_products": total_products - optimized_count
    })
def obtainChooseMatrix(request):
    name = request.GET.get('name')
    item_pipeline = [
                        { "$match": { "name": name } },
                    ]
    item_result = list(chooseMatrix.objects.aggregate(*item_pipeline))
    if item_result:
        del item_result[0]['_id']
        item_result = item_result[0]
        return JsonResponse(item_result,safe=False)
    return JsonResponse({},safe=False)
def InsightsDashboardView(request):
    all_products = Product.objects.only(
        "id", "product_title", "features", "product_description",
        "image_urls", "upc", "category",
        "asin", "sku", "fullfillment_by_channel"
    ).all()
    count_pipeline = [
        {
            "$count": "total_count"
        }
    ]
    total_count_result = list(Product.objects.aggregate(*(count_pipeline)))
    total_products = total_count_result[0]['total_count'] if total_count_result else 0
    optimized_count = 0
    refund_alerts = []
    fee_alerts = []
    listing_optimization_alerts = []
    product_performance_alerts = []  
    def is_optimized(product):
        title = product.product_title or ""
        if len(title) < 100 or re.search(r'(?i)(best|free|deal|offer|discount)', title):
            return False
        bullets = product.features or []
        if len(bullets) < 5 or any(re.search(r'<|>|🔥|👍|😁|[A-Z]{4,}', b) for b in bullets):
            return False
        description = product.product_description or ""
        if len(description) <= 300:
            return False
        words = re.findall(r'\b\w+\b', description)
        if len(words) != len(set(words)):
            return False
        images = product.image_urls or []
        if not images or any(not img.endswith(('.jpg', '.jpeg', '.png')) or 'watermark' in img.lower() for img in images):
            return False
        upc = product.upc or ""
        if not re.fullmatch(r'\d{12,14}', upc):
            return False
        category = product.category or ""
        if ">" not in category:
            return False
        return True
    def get_image_alerts(product):
        alerts = []
        images = product.image_urls or []
        if not images:
            alerts.append("No main image found. Please upload a clear product image with a white background.")
        else:
            main_image = images[0]
            if not (main_image.endswith('.jpg') or main_image.endswith('.jpeg') or main_image.endswith('.png')):
                alerts.append("Main image format should be JPG or PNG for better clarity and compatibility.")
            if 'white' not in main_image.lower():
                alerts.append("Update your main image background to white. This enhances your product's visual appeal and professionalism while meeting Amazon's requirements.")
            if 'small' in main_image.lower() or 'thumbnail' in main_image.lower():
                alerts.append("Update your main image size so that it is clear and of high quality for your potential customers.")
        return alerts
    Refund_obj = Refund.objects()
    refunded_product_ids = list(set([i.id for i in all_products][:2]))
    for product_id in refunded_product_ids:
        product = Product.objects(id=product_id).first()
        if not product:
            continue
        if is_optimized(product):
            optimized_count += 1
        alerts = get_image_alerts(product)
        if alerts:
            listing_optimization_alerts.append({
                "product_id": str(product.id),
                "title": product.product_title,
                "messages": alerts
            })
        total_orders = 0
        refund_count = Refund.objects(product_id=product.id).count()
        if total_orders > 0:
            refund_rate = (refund_count / total_orders) * 100
        else:
            refund_rate = 0 
        if refund_rate > 6:
            refund_alerts.append({
                "product_id": str(product.id),
                "title": product.product_title,
                "refund_rate": round(refund_rate, 2),
                "message": f"{product.product_title} has exceeded a 6% refund rate. Refund rates are soaring, impacting your profits. Review, analyze, and revise now."
            })
        if refund_rate <= 6:
            product_performance_alerts.append({
                "product_id": str(product.id),
                "title": product.product_title,
                "refund_rate": round(refund_rate, 2),
                "message": f"Refund rates for {product.product_title} have decreased by an impressive 6% or more. Your dedication is driving results, it’s time to take a closer look at your strategy."
            })
    today = datetime.utcnow()
    start_of_this_month = today.replace(day=1)
    start_of_last_month = (start_of_this_month - timedelta(days=1)).replace(day=1)
    this_month_fees = Fee.objects(
        marketplace="amazon.com",
        fee_type="storage",
        date__gte=start_of_this_month,
        date__lt=today
    ).sum('amount') or 0.0
    last_month_fees = Fee.objects(
        marketplace="amazon.com",
        fee_type="storage",
        date__gte=start_of_last_month,
        date__lt=start_of_this_month
    ).sum('amount') or 0.0
    if this_month_fees > last_month_fees:
        increase = round(this_month_fees - last_month_fees, 2)
        fee_alerts.append({
            "marketplace": "amazon.com",
            "increase_amount": increase,
            "message": f"Amazon Storage fees have increased by ${increase} for amazon.com. Storage fees have increased, cutting into your profit margins. Consider optimizing your inventory or fulfillment strategies now."
        })
    inventory_alerts = []
    for product in all_products:
        days_remaining = getattr(product, 'days_of_inventory_remaining', 999)
        if days_remaining <= 45:
            reorder_by_date = datetime.date.today() + datetime.timedelta(days=days_remaining)
            if days_remaining <= 38:
                alert_message = f"Order more inventory now to avoid running out of stock. You have {days_remaining} days of inventory remaining."
            else:
                alert_message = f"Order more inventory by {reorder_by_date.strftime('%B %d, %Y')} to avoid running out of stock. You have {days_remaining} days of inventory remaining."
            inventory_alerts.append({
                "title": getattr(product, 'title', ''),
                "asin": getattr(product, 'asin', ''),
                "sku": getattr(product, 'sku', ''),
                "fulfillment_channel": "FBA" if getattr(product, 'fullfillment_by_channel', False) else "FBM",
                "days_left": days_remaining,
                "reorder_by": reorder_by_date.isoformat(),
                "inventory_alert": alert_message
            })
    return JsonResponse({
    "total_products": total_products,
    "listing_optimization": {
        "optimized_products": optimized_count,
        "not_optimized_products": total_products - optimized_count
    },
    "insights_by_category": {
        "Listing Optimization": len(listing_optimization_alerts),
        "Product Performance": len(refund_alerts) + len(product_performance_alerts),
        "Inventory": len(fee_alerts) + len(inventory_alerts),
        "Refunds": len(refund_alerts),
        "Keyword": 42  
    },
    "alerts_feed": [  
        *[
            {
                "type": "Listing Optimization",
                "title": alert["title"],
                "date": datetime.today(),  
                "message": msg
            }
            for alert in listing_optimization_alerts
            for msg in alert["messages"]
        ],
        *[
            {
                "type": "Refunds",
                "title": alert["title"],
                "date":  datetime.today(),
                "message": alert["message"]
            }
            for alert in refund_alerts
        ],
        *[
            {
                "type": "Product Performance",
                "title": alert["title"],
                "date":  datetime.today(),
                "message": alert["message"]
            }
            for alert in product_performance_alerts
        ],
        *[
            {
                "type": "Inventory",
                "title": alert["title"],
                "date":  datetime.today(),
                "message": alert["inventory_alert"]
            }
            for alert in inventory_alerts
        ],
        *[
            {
                "type": "Inventory",
                "title": "Storage Fee Alert",
                "date":  datetime.today(),
                "message": alert["message"]
            }
            for alert in fee_alerts
        ],
        *[
        ]
    ]
})
def productsDetailsPageSummary(request):
    product_id = request.GET.get('product_id')
    pipeline = [
        {
            "$match": {
                "_id": ObjectId(product_id)
            }
        },
        {
            "$project": {
                "_id": 0,
                "sku": "$sku",
                "asin": {"$ifNull": ["$product_id", ""]},
                "product_title": {"$ifNull": ["$product_title", ""]},
                "image_url": {"$ifNull": ["$image_url", ""]},
                "price": {"$ifNull": ["$price", 0.0]},
                "stock": {"$ifNull": ["$quantity", 0]},
                "review_count": {"$ifNull": ["$review_count", 0]},
                "age": {
                    "$cond": {
                        "if": {"$ne": ["$product_created_date", None]},
                        "then": {
                            "$dateDiff": {
                                "startDate": "$product_created_date",
                                "endDate": "$$NOW",
                                "unit": "month"
                            }
                        },
                        "else": 0
                    }
                },
                "listing_quality_score": {"$ifNull": ["$listing_quality_score", 0.0]},
                "currency": {"$ifNull": ["$currency", ""]},
            }
        }
    ]
    item_result = list(Product.objects.aggregate(*pipeline))
    if item_result:
        item_result = item_result[0]
        return item_result
    return {}
def format_date_label(preset, start_date, end_date):
    if preset == "Today":
        return start_date.strftime("%B %d, %Y")
    elif preset == "Yesterday":
        return start_date.strftime("%B %d, %Y")
    elif preset in ["Last 7 Days", "Last 30 Days"]:
        return f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
    else:
        return f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
def getdaywiseproductssold_dict(start_date, end_date, product_id, is_hourly=False):
    results = getdaywiseproductssold(start_date, end_date, product_id, is_hourly)
    return {item["date"]: item for item in results}
def get_val_from_dict(date_obj, data_dict):
    date_str = date_obj.strftime("%Y-%m-%d")
    entry = data_dict.get(date_str)
    if entry:
        return entry["total_quantity"], float(entry["total_price"])
    return 0, 0.0
def sum_period_from_dict(start_day, end_day, data_dict):
    qty, price = 0, 0.0
    day = start_day
    while day <= end_day:
        q, p = get_val_from_dict(day, data_dict)
        qty += q
        price += p
        day += timedelta(days=1)
    return qty, round(price, 2)
def calc_diff_trend(current, previous):
    diff = round(current - previous, 2)
    trend = "up" if diff > 0 else "down" if diff < 0 else "neutral"
    return diff, trend
def productsSalesOverview(request):
    from django.utils import timezone
    product_id = request.GET.get("product_id")
    preset = request.GET.get("preset", "")
    timezone_str = request.GET.get('timezone', 'US/Pacific')
    local_tz = pytz.timezone(timezone_str)
    now = datetime.now(local_tz)
    is_hourly = False
    login_date = now.date()
    yesterday = login_date - timedelta(days=1)
    prev_day = yesterday - timedelta(days=1)
    last_7days_start = login_date - timedelta(days=7)
    last_7days_end = login_date - timedelta(days=1)
    prev_7days_start = last_7days_start - timedelta(days=7)
    prev_7days_end = last_7days_start - timedelta(days=1)
    label = None
    filled_graph = []
    stats_data_dict = getdaywiseproductssold_dict(
        datetime.combine(login_date - timedelta(days=15), datetime.min.time()),
        datetime.combine(login_date , datetime.max.time()),
        product_id,
        is_hourly=False
    )
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
    else:
        is_hourly = preset in ["Today", "Yesterday"]
        start_date, end_date = get_date_range(preset,timezone_str)
    if start_date and end_date:
        label = format_date_label(preset, start_date, end_date)
        graph_data = getdaywiseproductssold(start_date, end_date, product_id, is_hourly)
        for item in graph_data:
            raw_date = item.get("date")
            try:
                dt = datetime.strptime(raw_date, "%Y-%m-%d %H:00") if is_hourly else datetime.strptime(raw_date, "%Y-%m-%d")
                item["date"] = dt.strftime("%Y-%m-%d %H:00:00") if is_hourly else dt.strftime("%Y-%m-%d")
            except Exception:
                continue
        sales_dict = {item["date"]: item for item in graph_data}
        if is_hourly:
            base_date = start_date.strftime("%Y-%m-%d")
            hour_range = range(0, 25 if preset == "Today" else 24)
            for hour in hour_range:
                time_str = f"{base_date} {hour:02d}:00:00"
                filled_graph.append(sales_dict.get(time_str, {
                    "date": time_str,
                    "total_quantity": 0,
                    "total_price": 0.0
                }))
        else:
            current = start_date.date()
            while current <= end_date.date():
                date_str = current.strftime("%Y-%m-%d")
                filled_graph.append(sales_dict.get(date_str, {
                    "date": date_str,
                    "total_quantity": 0,
                    "total_price": 0.0
                }))
                current += timedelta(days=1)
    local_tz = pytz.timezone(timezone_str)
    today = datetime.now(local_tz).date()
    yesterday = today - timedelta(days=1)
    prev_day = yesterday - timedelta(days=1)
    prev_prev_day = prev_day - timedelta(days=1)
    last_7days_start = today - timedelta(days=7)
    last_7days_end = yesterday
    prev_7days_start = last_7days_start - timedelta(days=7)
    prev_7days_end = last_7days_start - timedelta(days=1)
    y_qty, y_price = get_val_from_dict(yesterday, stats_data_dict)
    p_qty, p_price = get_val_from_dict(prev_day, stats_data_dict)
    p_p_qty, p_p_price = get_val_from_dict(prev_prev_day, stats_data_dict)
    curr_qty, curr_price = sum_period_from_dict(last_7days_start, last_7days_end, stats_data_dict)
    prev_qty, prev_price = sum_period_from_dict(prev_7days_start, prev_7days_end, stats_data_dict)
    today_qty,today_price=get_val_from_dict(today,stats_data_dict)
    units = {
        "yesterday": {
            "value": y_qty,
            "difference": calc_diff_trend(y_qty, p_qty)[0],
            "trend": calc_diff_trend(y_qty, p_qty)[1],
        },
        "today":{
            "value":today_qty,
            'difference':calc_diff_trend(today_qty,y_qty)[0],
            "trend":calc_diff_trend(today_qty,y_qty)[1],
        },
        "previous_day": {
            "value": p_qty,
            "difference": calc_diff_trend(p_qty, p_p_qty)[0],
            "trend": calc_diff_trend(p_qty, p_p_qty)[1],
        },
        "last_7_days": {
            "value": curr_qty,
            "difference": calc_diff_trend(curr_qty, prev_qty)[0],
            "trend": calc_diff_trend(curr_qty, prev_qty)[1],
        }
    }
    sales = {
        "yesterday": {
            "value": y_price,
            "difference": calc_diff_trend(y_price, p_price)[0],
            "trend": calc_diff_trend(y_price, p_price)[1],
        },
        "today":{
            "value":today_price,
            "difference":calc_diff_trend(today_price,y_price)[0],
            "trend":calc_diff_trend(today_price,y_price)[1]
        },
        "previous_day": {
            "value": p_price,
            "difference": calc_diff_trend(p_price, p_p_price)[0],
            "trend": calc_diff_trend(p_price, p_p_price)[1],
        },
        "last_7_days": {
            "value": curr_price,
            "difference": calc_diff_trend(curr_price, prev_price)[0],
            "trend": calc_diff_trend(curr_price, prev_price)[1],
        }
    }
    return {
        "label": label,
        "units": units,
        "sales": sales,
        "graph": filled_graph
    }
def productsListingQualityScore(request):
    product_id = request.GET.get('product_id')
    product_doc = DatabaseModel.get_document(Product.objects,{"id" : ObjectId(product_id)})
    product_dict = product_doc.to_mongo().to_dict()
    listing_data = calculate_listing_score(product_dict)
    DatabaseModel.update_documents(Product.objects,{"id" : ObjectId(product_id)},{"listing_quality_score" : listing_data['final_score']})
    scoreData = {
        "asin": product_dict.get("product_id",""),
        "imageUrl": product_dict.get("image_url",""),
        "title": product_dict.get("product_title",""),  
        "productUrl": product_dict.get("product_url",""),
        "metricData": {
            "titleStrangeSymbols": {
                "metric": "titleStrangeSymbols",
                "metricTitle": "Title does not contain symbols or emojis",
                "metricTooltip": "Emojis and symbols can hamper readability",
                "passed": listing_data['rules_checks'][0]
            },
            "titleLength": {
                "metric": "titleLength",
                "metricTitle": "Title contains 150+ characters",
                "metricTooltip": "Maximized number of relevant keywords in your title improves discoverability",
                "passed": listing_data['rules_checks'][1]
            },
            "qtyBullets": {
                "metric": "qtyBullets",
                "metricTitle": "5+ bullet points",
                "metricTooltip": "Maximized number of bullet points can help improve discoverability",
                "passed": listing_data['rules_checks'][2]
            },
            "lengthBullets": {
                "metric": "lengthBullets",
                "metricTitle": "150+ characters in each bullet point",
                "metricTooltip": "Maximized number of relevant keywords in bullet points helps improve discoverability",
                "passed": listing_data['rules_checks'][3]
            },
            "capitalizedBullets": {
                "metric": "capitalizedBullets",
                "metricTitle": "First letter of bullet points is capitalized",
                "metricTooltip": "Capitalized first letter of first word in each bullet point improves readability",
                "passed": listing_data['rules_checks'][4]
            },
            "allCapsBullets": {
                "metric": "allCapsBullets",
                "metricTitle": "Bullet points are not in all caps",
                "metricTooltip": "Amazon TOS discourages using all caps",
                "passed": listing_data['rules_checks'][5]
            },
            "ebcAndDescription": {
                "metric": "ebcAndDescription",
                "metricTitle": "1,000+ characters in description or A+ content",
                "metricTooltip": "Maximized number of relevant keywords in description helps improve discoverability",
                "passed": listing_data['rules_checks'][6]
            },
            "imageResolution": {
                "metric": "imageResolution",
                "metricTitle": "1000 x 1000 px +",
                "metricTooltip": "Images at least 1000 x 1000px enable zoom feature",
                "passed": listing_data['rules_checks'][7]
            },
            "imageBackground": {
                "metric": "imageBackground",
                "metricTitle": "Main image is on a white background",
                "metricTooltip": "Amazon TOS requires main image to be on a white background",
                "passed": listing_data['rules_checks'][8]
            },
            "imagesQty": {
                "metric": "imagesQty",
                "metricTitle": "7+ images",
                "metricTooltip": "Increased number of images can help drive conversions",
                "passed": listing_data['rules_checks'][9]
            },
            "videosQty": {
                "metric": "videosQty",
                "metricTitle": "Includes video",
                "metricTooltip": "Videos can help users learn more about the product and increase conversions",
                "passed": listing_data['rules_checks'][10]
            },
            "reviewQty": {
                "metric": "reviewQty",
                "metricTitle": "20+ reviews",
                "metricTooltip": "Increased number of reviews can increase product credibility with shoppers",
                "passed": listing_data['rules_checks'][11]
            },
            "reviewRating": {
                "metric": "reviewRating",
                "metricTitle": "4+ average star ratings",
                "metricTooltip": "Increased number of positive, 4+ star ratings can increase product credibility with shoppers",
                "passed": listing_data['rules_checks'][12]
            }
        },
        "totalScore": listing_data['final_score']
    }
    return scoreData
def productsTrafficandConversions(request):
    data = dict()
    preset = request.GET.get('preset')
    product_id = request.GET.get('product_id')
    product_obj = DatabaseModel.get_document(Product.objects,{"id" : ObjectId(product_id)},['product_id'])
    data['asin'] = product_obj.product_id
    timezone_str = request.GET.get('timezone', 'US/Pacific')
    start_date = request.GET.get("start_date", None)
    end_date = request.GET.get("end_date", None)
    if start_date != None and start_date != "":
        start_date, end_date = convertdateTotimezone(start_date,end_date,timezone_str)
    else:
        start_date, end_date = get_date_range(preset,timezone_str)
    data['date'] = start_date.strftime("%b %d, %Y") + " - " + end_date.strftime("%b %d, %Y")
    if timezone_str != 'UTC':
        start_date,end_date = convertLocalTimeToUTC(start_date, end_date, timezone_str)
    daily_sales = getdaywiseproductssold(start_date, end_date, product_id)
    view_and_sales = pageViewsandSessionCount(start_date,end_date,product_id)
    data['total_units_sold'] = sum(item['total_quantity'] for item in daily_sales)
    data['average_units_sold'] = 0
    units_sold_graph = []
    for item in daily_sales:
        units_sold_graph.append({
            "date": item['date'],
            "units": item['total_quantity'],
            "average" : 0
        })
    data['units_sold_graph'] = units_sold_graph
    data['total_sessions'] = sum(item['session_count'] for item in view_and_sales)
    data['average_sessions'] = 0
    sessions_graph = []
    for item in view_and_sales:
        sessions_graph.append({
            "date": item['date'],
            "sessions": item['session_count'],
            "average" : 0
        })
    data['sessions_graph'] = sessions_graph
    data['total_page_views'] = sum(item['page_views'] for item in view_and_sales)
    data['average_page_views'] = 0
    page_views_graph = []
    for item in view_and_sales:
        page_views_graph.append({
            "date": item['date'],
            "page_views": item['page_views'],
            "average" : 0
        })
    data['page_views_graph'] = page_views_graph
    return data
@csrf_exempt
def getSKUlist(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    search_query = json_request.get('search_query')
    brand_id = json_request.get('brand_id')
    manufacturer_name = json_request.get('manufacturer_name')
    match =dict()
    pipeline = []
    if search_query != None and search_query != "":
        search_query = re.escape(search_query.strip())
        match["sku"] = {"$regex": search_query, "$options": "i"}
    if marketplace_id != None and marketplace_id != "" and marketplace_id != "all" and marketplace_id != "custom":
        match['marketplace_id'] = ObjectId(marketplace_id)
    if brand_id != None and brand_id != "" and brand_id != [] and brand_id != "custom":
        brand_list = [ObjectId(i) for i in brand_id]
        match['brand_id'] = {"$in":brand_list}
    if manufacturer_name != None and manufacturer_name != "" and manufacturer_name != [] and manufacturer_name != "custom":
        match['manufacturer_name'] = {"$in":manufacturer_name}
    if match != {}:
        pipeline.append({"$match": match})
    pipeline.extend([
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "sku": "$sku",
            }
        },{
            "$sort":{
                "sku":1}
        }
    ])
    sku_list = list(Product.objects.aggregate(*pipeline))
    return sku_list
@csrf_exempt
def getproductIdlist(request):
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    brand_id = json_request.get('brand_id')
    search_query = json_request.get('search_query')
    manufacturer_name = json_request.get('manufacturer_name')
    match = dict()
    pipeline = []
    if brand_id and isinstance(brand_id, str):
        brand_id = [brand_id]
    if manufacturer_name and isinstance(manufacturer_name, str):
        manufacturer_name = [manufacturer_name]
    if brand_id and brand_id not in ["", [], "custom"]:
        brand_list = [ObjectId(i) for i in brand_id]
        brand_objs = list(Brand.objects.filter(id__in=brand_list))
        brand_names = [b.name for b in brand_objs]
        match['brand_id'] = {"$in": brand_list}
    else:
        brand_names = []
    if search_query:
        search_query = re.escape(search_query.strip())
        match["product_id"] = {"$regex": search_query, "$options": "i"}
    if marketplace_id and marketplace_id not in ["", "all", "custom"]:
        match['marketplace_id'] = ObjectId(marketplace_id)
    if manufacturer_name and manufacturer_name not in ["", [], "custom"]:
        match['manufacturer_name'] = {"$in": manufacturer_name}
    if match:
        pipeline.append({"$match": match})
    else:
        pipeline.append({"$sample": {"size": 10}})  
    pipeline.append({
        "$project": {
            "_id": 0,
            "id": {"$toString": "$_id"},
            "Asin": "$product_id",
            "product_title": "$product_title"  
        }
    })
    asin_list = list(Product.objects.aggregate(*pipeline))
    return asin_list
def getBrandListforfilter(request):
    data = dict()
    marketplace_id = request.GET.get('marketplace_id')
    search_query = request.GET.get('search_query')
    skip = int(request.GET.get('skip', 1))
    product_ids=request.GET.get('product_id[]')
    query = {}
    if marketplace_id and marketplace_id not in ["", "all", "custom"]:
        query['marketplace_id'] = ObjectId(marketplace_id)
    if search_query and search_query.strip():
        search_query = re.escape(search_query.strip())
        query["name"] = {"$regex": search_query, "$options": "i"}
    if not query:
        brand_cursor = Brand.objects.only('name').order_by('name')
    else:
        brand_cursor = Brand.objects.filter(**query).only('name').order_by('name')
    brand_list = [
        {
            "id": str(brand.id),
            "name": brand.name
        }
        for brand in brand_cursor
    ]
    data['brand_list'] = brand_list
    return data
def obtainManufactureNames(request):
    marketplace_id = request.GET.get('marketplace_id')
    search_query = request.GET.get('search_query')
    match = {}
    if search_query:
        search_query = re.escape(search_query.strip())
        match["manufacturer_name"] = {"$regex": search_query, "$options": "i"}
    if marketplace_id and marketplace_id not in ["", "all", "custom"]:
        match['marketplace_id'] = ObjectId(marketplace_id)
    match['manufacturer_name'] = match.get('manufacturer_name', {})
    match['manufacturer_name']["$ne"] = ""
    match['manufacturer_name']["$ne"] = None
    pipeline = []
    if match:
        pipeline.append({"$match": match})
    pipeline.extend([
        {
            "$group": {
                "_id": "$manufacturer_name"
            }
        },
        {
            "$project": {
                "_id": 0,
                "manufacturer_name": "$_id"
            }
        },
        {
            "$sort": {
                "manufacturer_name": 1
            }
        }
    ])
    Product_list = list(Product.objects.aggregate(*pipeline))
    names = [p["manufacturer_name"] for p in Product_list if p["manufacturer_name"] not in ["", None]]
    data = {"manufacturer_name_list": names}
    return JsonResponse(data, safe=False)
def InsightsProductWise(request):
    product_id = request.GET.get('product_id')
    optimized_count = 0
    refund_alerts = []
    fee_alerts = []
    listing_optimization_alerts = []
    product_performance_alerts = []  
    def is_optimized(product):
        title = product.product_title or ""
        if len(title) < 100 or re.search(r'(?i)(best|free|deal|offer|discount)', title):
            return False
        bullets = product.features or []
        if len(bullets) < 5 or any(re.search(r'<|>|🔥|👍|😁|[A-Z]{4,}', b) for b in bullets):
            return False
        description = product.product_description or ""
        if len(description) <= 300:
            return False
        words = re.findall(r'\b\w+\b', description)
        if len(words) != len(set(words)):
            return False
        images = product.image_urls or []
        if not images or any(not img.endswith(('.jpg', '.jpeg', '.png')) or 'watermark' in img.lower() for img in images):
            return False
        upc = product.upc or ""
        if not re.fullmatch(r'\d{12,14}', upc):
            return False
        category = product.category or ""
        if ">" not in category:
            return False
        return True
    def get_image_alerts(product):
        alerts = []
        images = product.image_urls or []
        if not images:
            alerts.append("No main image found. Please upload a clear product image with a white background.")
        else:
            main_image = images[0]
            if not (main_image.endswith('.jpg') or main_image.endswith('.jpeg') or main_image.endswith('.png')):
                alerts.append("Main image format should be JPG or PNG for better clarity and compatibility.")
            if 'white' not in main_image.lower():
                alerts.append("Update your main image background to white. This enhances your product's visual appeal and professionalism while meeting Amazon's requirements.")
            if 'small' in main_image.lower() or 'thumbnail' in main_image.lower():
                alerts.append("Update your main image size so that it is clear and of high quality for your potential customers.")
        return alerts
    Refund_obj = Refund.objects(product_id=product_id)
    refunded_product_ids = list(set([i.product_id.id for i in Refund_obj]))
    for product_id in refunded_product_ids:
        product = Product.objects(id=product_id).first()
        if not product:
            continue
        if is_optimized(product):
            optimized_count += 1
        alerts = get_image_alerts(product)
        if alerts:
            listing_optimization_alerts.append({
                "product_id": str(product.id),
                "title": product.product_title,
                "messages": alerts
            })
        pipeline = [
            {
                "$lookup": {
                    "from": "order_items",
                    "localField": "order_items",
                    "foreignField": "_id",
                    "as": "order_items"
                }
            },
            {"$unwind": "$order_items"},
            {
                "$match": {
                    "order_items.ProductDetails.product_id": ObjectId(str(product.id))
                }
            }
        ]
        orders = list(Order.objects.aggregate(*pipeline))
        total_orders = len(orders)
        refund_count = Refund.objects(product_id=product.id).count()
        if total_orders > 0:
            refund_rate = (refund_count / total_orders) * 100
            if refund_rate > 6:
                refund_alerts.append({
                    "product_id": str(product.id),
                    "title": product.product_title,
                    "refund_rate": round(refund_rate, 2),
                    "message": f"{product.product_title} has exceeded a 6% refund rate. Refund rates are soaring, impacting your profits. Review, analyze, and revise now."
                })
            if refund_rate <= 6:
                product_performance_alerts.append({
                    "product_id": str(product.id),
                    "title": product.product_title,
                    "refund_rate": round(refund_rate, 2),
                    "message": f"Refund rates for {product.product_title} have decreased by an impressive 6% or more. Your dedication is driving results, it’s time to take a closer look at your strategy."
                })
    today = datetime.utcnow()
    start_of_this_month = today.replace(day=1)
    start_of_last_month = (start_of_this_month - timedelta(days=1)).replace(day=1)
    this_month_fees = Fee.objects(
        marketplace="amazon.com",
        fee_type="storage",
        date__gte=start_of_this_month,
        date__lt=today
    ).sum('amount') or 0.0
    last_month_fees = Fee.objects(
        marketplace="amazon.com",
        fee_type="storage",
        date__gte=start_of_last_month,
        date__lt=start_of_this_month
    ).sum('amount') or 0.0
    if this_month_fees > last_month_fees:
        increase = round(this_month_fees - last_month_fees, 2)
        fee_alerts.append({
            "marketplace": "amazon.com",
            "increase_amount": increase,
            "message": f"Amazon Storage fees have increased by ${increase} for amazon.com. Storage fees have increased, cutting into your profit margins. Consider optimizing your inventory or fulfillment strategies now."
        })
    inventory_alerts = []
    product_obj = Product.objects(id=product_id).first()
    days_remaining = getattr(product_obj, 'days_of_inventory_remaining', 999)
    if days_remaining <= 45:
        reorder_by_date = datetime.date.today() + datetime.timedelta(days=days_remaining)
        if days_remaining <= 38:
            alert_message = f"Order more inventory now to avoid running out of stock. You have {days_remaining} days of inventory remaining."
        else:
            alert_message = f"Order more inventory by {reorder_by_date.strftime('%B %d, %Y')} to avoid running out of stock. You have {days_remaining} days of inventory remaining."
        inventory_alerts.append({
            "title": getattr(product_obj, 'title', ''),
            "asin": getattr(product_obj, 'asin', ''),
            "sku": getattr(product_obj, 'sku', ''),
            "fulfillment_channel": getattr(product_obj, 'fulfillment_channel', ''),
            "days_left": days_remaining,
            "reorder_by": reorder_by_date.isoformat(),
            "inventory_alert": alert_message
        })
    return JsonResponse({
    "alerts_feed": [
        *[
            {
                "type": "Listing Optimization",
                "title": alert["title"],
                "date": datetime.today(),
                "message": msg
            }
            for alert in listing_optimization_alerts
            for msg in alert["messages"]
        ],
        *[
            {
                "type": "Refunds",
                "title": alert["title"],
                "date":  datetime.today(),
                "message": alert["message"]
            }
            for alert in refund_alerts
        ],
        *[
            {
                "type": "Product Performance",
                "title": alert["title"],
                "date":  datetime.today(),
                "message": alert["message"]
            }
            for alert in product_performance_alerts
        ],
        *[
            {
                "type": "Inventory",
                "title": alert["title"],
                "date":  datetime.today(),
                "message": alert["inventory_alert"]
            }
            for alert in inventory_alerts
        ],
        *[
            {
                "type": "Inventory",
                "title": "Storage Fee Alert",
                "date":  datetime.today(),
                "message": alert["message"]
            }
            for alert in fee_alerts
        ],
        *[
        ]
    ]
})
@csrf_exempt
def getProfitAndLossDetailsForProduct(request):
    def to_utc_format(dt):
        return dt.strftime("%Y-%m-%dT%H:%M:%SZ")
    json_request = JSONParser().parse(request)
    product_id = json_request.get('product_id')
    preset = json_request.get('preset')
    preset = json_request.get('preset')
    timezone_str = json_request.get('timezone', 'US/Pacific')
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date != None and start_date != "":
        start_date, end_date = convertdateTotimezone(start_date,end_date,timezone_str)
    else:
        start_date, end_date = get_date_range(preset,timezone_str)
    def calculate_metrics(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str):
        gross_revenue = 0
        total_cogs = 0
        refund = 0
        net_profit = 0
        margin = 0
        total_units = 0
        shipping_cost = 0
        channel_fee = 0
        sku_set = set()
        product_categories = {}
        product_completeness = {"complete": 0, "incomplete": 0}
        result = grossRevenue(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
        order_total = 0
        tax_price = 0
        temp_price = 0
        vendor_funding = 0
        if result:
            for order in result:
                gross_revenue += order['order_total']
                order_total = order['order_total']
                for item_id in order['order_items']:
                    item_pipeline = [
                        { "$match": { "_id": item_id } },
                        {
                            "$lookup": {
                                "from": "product",
                                "localField": "ProductDetails.product_id",
                                "foreignField": "_id",
                                "as": "product_ins"
                            }
                        },
                        { "$unwind": { "path": "$product_ins", "preserveNullAndEmptyArrays": True } },
                        {
                            "$project": {
                                "_id": 0,
                                "price": "$Pricing.ItemPrice.Amount",
                                "tax_price": "$Pricing.ItemTax.Amount",
                                "cogs": { "$ifNull": ["$product_ins.cogs", 0.0] },
                                "sku": "$product_ins.sku",
                                "category": "$product_ins.category",
                                "total_cogs" : {"$ifNull":["$product_ins.total_cogs",0]},
                            "w_total_cogs" : {"$ifNull":["$product_ins.w_total_cogs",0]},
                            "vendor_funding" : {"$ifNull":["$product_ins.vendor_funding",0]},
                            "a_shipping_cost" : {"$ifNull":["$product_ins.a_shipping_cost",0]},
                            "w_shiping_cost" : {"$ifNull":["$product_ins.w_shiping_cost",0]},
                            "referral_fee" : {"$ifNull":["$product_ins.referral_fee",0]},
                            "walmart_fee" : {"$ifNull":["$product_ins.walmart_fee",0]},
                            }
                        }
                    ]
                    item_result = list(OrderItems.objects.aggregate(*item_pipeline))
                    if item_result:
                        item_data = item_result[0]
                        temp_price += item_data['price']
                        tax_price += item_data['tax_price']
                        if order['marketplace_name'] == "Amazon":
                            total_cogs += item_data['total_cogs']
                            shipping_cost += item_data['a_shipping_cost']
                            channel_fee += item_data['referral_fee']
                        else:
                            total_cogs += item_data['w_total_cogs']
                            shipping_cost += item_data['w_shiping_cost']
                            channel_fee += item_data['walmart_fee']
                        vendor_funding += item_data['vendor_funding']
                        total_units += 1
                        if item_data.get('sku'):
                            sku_set.add(item_data['sku'])
                        category = item_data.get('category', 'Unknown')
                        if category in product_categories:
                            product_categories[category] += 1
                        else:
                            product_categories[category] = 1
                        if item_data['price'] and item_data['total_cogs'] and item_data['sku']:
                            product_completeness["complete"] += 1
                        else:
                            product_completeness["incomplete"] += 1
            net_profit = (temp_price - total_cogs) + vendor_funding
            margin = (net_profit / gross_revenue) * 100 if gross_revenue > 0 else 0
        return {
            "grossRevenue": round(gross_revenue, 2),
            "expenses": round((total_cogs) , 2),
            "netProfit": round(net_profit, 2),
            "roi": round((net_profit / (total_cogs)) * 100, 2) if total_cogs > 0 else 0,
            "unitsSold": total_units,
            "refunds": refund,   
            "skuCount": len(sku_set),
            "sessions": 0,
            "pageViews": 0,
            "unitSessionPercentage": 0,
            "margin": round(margin, 2),
            "seller":"",
            "tax_price":tax_price,
            "total_cogs":total_cogs,
            "product_cost":order_total,
            "shipping_cost":shipping_cost,
            "productCategories": product_categories,  
            "productCompleteness": product_completeness,  
            'base_price':temp_price,
            'channel_fee' : channel_fee
        }
    def create_period_response(label, cur_from, cur_to, prev_from, prev_to,marketplace_id=None,brand_id=[],product_id=[],manufacturer_name=[],fulfillment_channel=[],preset=None,timezone_str="UTC"):
        current = calculate_metrics(cur_from, cur_to,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
        previous = calculate_metrics(prev_from, prev_to,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
        def with_delta(metric):
            return {
                "current": current[metric],
                "previous": previous[metric],
                "delta": round(current[metric] - previous[metric], 2)
            }
        if preset in ['Today', 'Yesterday']:
            return {
                "dateRanges": {
                    "current": {"from": to_utc_format(cur_from),},
                    "previous": {"from": to_utc_format(prev_from),}
                },
                "summary": {
                    "grossRevenue": with_delta("grossRevenue"),
                    "netProfit": with_delta("netProfit"),
                    "expenses": with_delta("expenses"),
                    "unitsSold": with_delta("unitsSold"),
                    "refunds": with_delta("refunds"),
                    "skuCount": with_delta("skuCount"),
                    "sessions": with_delta("sessions"),
                    "pageViews": with_delta("pageViews"),
                    "unitSessionPercentage": with_delta("unitSessionPercentage"),
                    "margin": with_delta("margin"),
                    "roi": with_delta("roi")
                },
                "netProfitCalculation": {
                    "current": {
                        "gross": current["grossRevenue"],
                        "totalCosts": current["expenses"],
                        "productRefunds": current["refunds"],
                        "totalTax": current["tax_price"] if 'tax_price' in current else 0,
                        "totalTaxWithheld": 0,
                        "ppcProductCost": 0,
                        "ppcBrandsCost": 0,
                        "ppcDisplayCost": 0,
                        "ppcStCost": 0,
                        "cogs": current["total_cogs"] if 'total_cogs' in current else 0,
                        "product_cost": current["product_cost"],
                        "base_price": current["base_price"],
                        "shipping_cost": current["shipping_cost"],
                        "channel_fee" : current["channel_fee"]
                    },
                    "previous": {
                        "gross": previous["grossRevenue"],
                        "totalCosts": previous["expenses"],
                        "productRefunds": previous["refunds"],
                        "totalTax": previous["total_cogs"] if 'total_cogs' in previous else 0,
                        "totalTaxWithheld": 0,
                        "ppcProductCost": 0,
                        "ppcBrandsCost": 0,
                        "ppcDisplayCost": 0,
                        "ppcStCost": 0,
                        "cogs": previous["total_cogs"] if 'total_cogs' in previous else 0,
                        "product_cost": previous["product_cost"],
                        "base_price": current["base_price"],
                        "shipping_cost": previous["shipping_cost"],
                        "channel_fee" : previous["channel_fee"]
                    }
                },
                "charts": {
                    "productDistribution": current["productCategories"],  
                    "productCompleteness": current["productCompleteness"]  
                }
            }
        else:
            return {
                "dateRanges": {
                    "current": {"from": to_utc_format(cur_from),"to": to_utc_format(cur_to)},
                    "previous": {"from": to_utc_format(prev_from),"to": to_utc_format(prev_to)}
                },
                "summary": {
                    "grossRevenue": with_delta("grossRevenue"),
                    "netProfit": with_delta("netProfit"),
                    "expenses": with_delta("expenses"),
                    "unitsSold": with_delta("unitsSold"),
                    "refunds": with_delta("refunds"),
                    "skuCount": with_delta("skuCount"),
                    "sessions": with_delta("sessions"),
                    "pageViews": with_delta("pageViews"),
                    "unitSessionPercentage": with_delta("unitSessionPercentage"),
                    "margin": with_delta("margin"),
                    "roi": with_delta("roi")
                },
                "netProfitCalculation": {
                    "current": {
                        "gross": current["grossRevenue"],
                        "totalCosts": current["expenses"],
                        "productRefunds": current["refunds"],
                        "totalTax": current["tax_price"] if 'tax_price' in current else 0,
                        "totalTaxWithheld": 0,
                        "ppcProductCost": 0,
                        "ppcBrandsCost": 0,
                        "ppcDisplayCost": 0,
                        "ppcStCost": 0,
                        "cogs": current["total_cogs"] if 'total_cogs' in current else 0,
                        "product_cost": current["product_cost"],
                        "base_price": current["base_price"],
                        "shipping_cost": current["shipping_cost"],
                    },
                    "previous": {
                        "gross": previous["grossRevenue"],
                        "totalCosts": previous["expenses"],
                        "productRefunds": previous["refunds"],
                        "totalTax": previous["total_cogs"] if 'total_cogs' in previous else 0,
                        "totalTaxWithheld": 0,
                        "ppcProductCost": 0,
                        "ppcBrandsCost": 0,
                        "ppcDisplayCost": 0,
                        "ppcStCost": 0,
                        "cogs": previous["total_cogs"] if 'total_cogs' in previous else 0,
                        "product_cost": previous["product_cost"],
                        "base_price": current["base_price"],
                        "shipping_cost": previous["shipping_cost"],
                    }
                },
                "charts": {
                    "productDistribution": current["productCategories"],  
                    "productCompleteness": current["productCompleteness"]  
                }
            }
    custom_duration = end_date - start_date
    prev_from_date = start_date - custom_duration
    prev_to_date = end_date - custom_duration
    response_data = {
        "custom": create_period_response("Custom", start_date, end_date, prev_from_date, prev_to_date,None,[],[product_id],[],[],preset),
    }
    return response_data
@csrf_exempt
def profitlosschartForProduct(request):
    json_request = JSONParser().parse(request)
    product_id = json_request.get('product_id')
    preset = json_request.get('preset')
    timezone_str = json_request.get('timezone', 'US/Pacific')
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date != None and start_date != "":
        start_date, end_date = convertdateTotimezone(start_date,end_date,timezone_str)
    else:
        start_date, end_date = get_date_range(preset,timezone_str)
    def get_month_range(year, month):
        start_date = datetime(year, month, 1)
        last_day = monthrange(year, month)[1]
        end_date = datetime(year, month, last_day, 23, 59, 59)
        return start_date, end_date
    def calculate_metrics(start_date, end_date,marketplace_id=None,brand_id=[],product_id=[],manufacturer_name=[],fulfillment_channel=[],timezone_str="UTC"):
        gross_revenue_amt = 0
        total_cogs = 0
        refund = 0
        net_profit = 0
        margin = 0
        total_units = 0
        sku_set = set()
        product_categories = {}
        product_completeness = {"complete": 0, "incomplete": 0}
        order_total = 0
        tax_price = 0
        temp_price = 0
        vendor_funding = 0
        result = grossRevenue(start_date, end_date,marketplace_id,brand_id,product_id,manufacturer_name,fulfillment_channel,timezone_str)
        for order in result:
            gross_revenue_amt += order.get("order_total", 0)
            order_total = order.get("order_total", 0)
            tax_price = 0
            for item_id in order.get("order_items", []):
                item_pipeline = [
                    {"$match": {"_id": item_id}},
                    {
                        "$lookup": {
                            "from": "product",
                            "localField": "ProductDetails.product_id",
                            "foreignField": "_id",
                            "as": "product_ins"
                        }
                    },
                    {"$unwind": {"path": "$product_ins", "preserveNullAndEmptyArrays": True}},
                    {
                        "$project": {
                            "price": "$Pricing.ItemPrice.Amount",
                            "tax_price": "$Pricing.ItemTax.Amount",
                            "cogs": {"$ifNull": ["$product_ins.cogs", 0.0]},
                            "sku": "$product_ins.sku",
                            "category": "$product_ins.category",
                            "total_cogs" : {"$ifNull":["$product_ins.total_cogs",0]},
                            "w_total_cogs" : {"$ifNull":["$product_ins.w_total_cogs",0]},
                            "vendor_funding" : {"$ifNull":["$product_ins.vendor_funding",0]},
                        }
                    }
                ]
                item_result = list(OrderItems.objects.aggregate(*item_pipeline))
                if item_result:
                    item = item_result[0]
                    temp_price += item.get("price", 0)
                    tax_price += item.get("tax_price", 0)
                    if order['marketplace_name'] == "Amazon":
                        total_cogs += item.get("total_cogs", 0) 
                    else:
                        total_cogs += item.get("w_total_cogs", 0)
                    vendor_funding += item.get("vendor_funding", 0)
                    total_units += 1
                    sku = item.get("sku")
                    if sku:
                        sku_set.add(sku)
                    category = item.get("category", "Unknown")
                    product_categories[category] = product_categories.get(category, 0) + 1
                    if item.get("price") and item.get("total_cogs") and sku:
                        product_completeness["complete"] += 1
                    else:
                        product_completeness["incomplete"] += 1
        net_profit = (temp_price - total_cogs) + vendor_funding
        margin = (net_profit / gross_revenue_amt * 100) if gross_revenue_amt else 0
        return {
            "grossRevenue": round(gross_revenue_amt, 2),
            "expenses": round((total_cogs) , 2),
            "netProfit": round(net_profit, 2),
            "roi": round((net_profit / (total_cogs)) * 100, 2) if (total_cogs) else 0,
            "unitsSold": total_units,
            "refunds": refund,
            "skuCount": len(sku_set),
            "margin": round(margin, 2),
            "tax_price": tax_price,
            "total_cogs": total_cogs,
            "product_cost": order_total,
            "productCategories": product_categories,
            "productCompleteness": product_completeness
        }
    def generate_month_keys(start_year, start_month, end_year, end_month):
        months = []
        current = datetime(start_year, start_month, 1)
        end = datetime(end_year, end_month, 1)
        while current <= end:
            months.append(current.strftime("%Y-%m-%d 00:00:00"))
            current += timedelta(days=32)
            current = current.replace(day=1)
        return months
    metrics = ["grossRevenue", "estimatedPayout", "expenses", "netProfit", "units", "ppcSales"]
    values = {metric: {} for metric in metrics}
    hourly_presets = ["Today", "Yesterday"]
    daily_presets = daily_presets = ["This Week", "Last Week", "Last 7 days", "Last 14 days", "Last 30 days", "Last 60 days", "Last 90 days","Last Month","This Quarter","Last Quarter","Last Year"]
    if preset in hourly_presets:
        interval_keys = [(start_date + timedelta(hours=i)).strftime("%Y-%m-%d %H:00:00") 
                         for i in range(0, int((end_date - start_date).total_seconds() // 3600) + 1)]
        interval_type = "hour"
    elif preset in daily_presets:
        interval_keys = [(start_date + timedelta(days=i)).strftime("%Y-%m-%d 00:00:00") 
                         for i in range((end_date - start_date).days + 1)]
        interval_type = "day"
    else:
        interval_keys = generate_month_keys(
            start_date.year, start_date.month,
            end_date.year, end_date.month
        )
        interval_type = "month"
    for key in interval_keys:
        if interval_type == "hour":
            start = datetime.strptime(key, "%Y-%m-%d %H:00:00")
            end = start + timedelta(hours=1) - timedelta(seconds=1)
        elif interval_type == "day":
            start = datetime.strptime(key, "%Y-%m-%d 00:00:00")
            end = start + timedelta(days=1) - timedelta(seconds=1)
        else:
            year, month = int(key[:4]), int(key[5:7])
            start, end = get_month_range(year, month)
        data = calculate_metrics(start, end,None,[],[product_id],[],None)
        values["grossRevenue"][key] = data["grossRevenue"]
        values["expenses"][key] = data["expenses"]
        values["netProfit"][key] = data["netProfit"]
        values["units"][key] = data["unitsSold"]
    for metric in metrics:
        for key in interval_keys:
            values[metric].setdefault(key, 0)
    graph = [{"metric": metric, "values": values[metric]} for metric in metrics]
    data ={
       "graph": graph 
    }
    return data
@csrf_exempt
def getrevenuedetailsForProduct(request):
    json_request = JSONParser().parse(request)
    preset = json_request.get("preset", None)
    product_id = json_request.get("product_id", None)
    timezone_str = json_request.get('timezone', 'US/Pacific')
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date != None and start_date != "":
        start_date, end_date = convertdateTotimezone(start_date,end_date,timezone_str)
    else:
        start_date, end_date = get_date_range(preset,timezone_str)
    if preset in ['Today', 'Yesterday']:
        date_range_label = f"{start_date.strftime('%b %d, %Y')} - {start_date.strftime('%b %d, %Y')}"
    else:
        date_range_label = f"{start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"
    def get_previous_date_range(start_date, end_date):
        duration = end_date - start_date
        previous_start_date = start_date - duration - timedelta(days=1)
        previous_end_date = start_date - timedelta(days=1)
        return previous_start_date.strftime("%Y-%m-%d"), previous_end_date.strftime("%Y-%m-%d")
    compare_startdate, compare_enddate = get_previous_date_range(start_date, end_date)
    def fetch_total():
        return totalRevenueCalculationForProduct(start_date, end_date,None, [], [product_id], [], None,timezone_str)
    def fetch_compare_total():
        return totalRevenueCalculation(compare_startdate, compare_enddate, None, [], [product_id],[], None,timezone_str)
    with ThreadPoolExecutor(max_workers=4) as executor:
        future_total = executor.submit(fetch_total)
        compare_total = None
        if compare_startdate != None and compare_startdate != "":
            compare_startdate = datetime.strptime(compare_startdate, "%Y-%m-%d").replace(hour=0, minute=0, second=0, microsecond=0)
            compare_enddate = datetime.strptime(compare_enddate, "%Y-%m-%d").replace(hour=23, minute=59, second=59, microsecond=0)
            future_compare_total = executor.submit(fetch_compare_total)
    total = future_total.result()
    if compare_startdate != None and compare_startdate != "":
        compare_total = future_compare_total.result()
    data = {
        "date_range_label" : date_range_label,
        "total": total,
    }
    difference = {
        "gross_revenue": round(total["gross_revenue"] - compare_total["gross_revenue"],2),
        "net_profit": round(total["net_profit"] - compare_total["net_profit"],2),
        "units_sold": total["units_sold"] - compare_total["units_sold"]
    }
    data['compare_total'] = difference
    return data
@csrf_exempt
def getInventryLogForProductdaywise(request):
    json_request = JSONParser().parse(request)
    preset = json_request.get("preset", "Today")
    product_id = json_request.get("product_id")
    timezone_str = json_request.get('timezone', 'US/Pacific')
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date != None and start_date != "":
        start_date, end_date = convertdateTotimezone(start_date,end_date,timezone_str)
    else:
        start_date, end_date = get_date_range(preset,timezone_str)
    if preset in ['Today', 'Yesterday']:
        date_range_label = f"{start_date.strftime('%b %d, %Y')} - {start_date.strftime('%b %d, %Y')}"
    else:
        date_range_label = f"{start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}"
    if timezone_str != 'UTC':
        start_date,end_date = convertLocalTimeToUTC(start_date, end_date, timezone_str)
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]
    logs = inventry_log.objects.filter(product_id=product_id, date__gte=start_date, date__lte=end_date)
    for i in logs:
        print(i.product_id.product_title)
    log_dict = {log.date.date(): log for log in logs}
    response_data = []
    for date in date_range:
        log = log_dict.get(date.date())
        response_data.append({
            "date": date.strftime('%b %d'),
            "Available": log.available if log else 0,
            "Reserved": log.reserved if log else 0,
        })
    data = {
        "date_range_label" : date_range_label,
        "response_data" : response_data
    }
    return data
def getProductInformation(request):
    product_id = request.GET.get('product_id')
    product_obj = DatabaseModel.get_document(Product.objects,{"id" : product_id})
    marketplace_ids = product_obj.marketplace_ids
    marketplaces = Marketplace.objects.filter(id__in=[ins.id for ins in marketplace_ids])
    has_amazon = len(marketplaces.filter(name__iexact="Amazon")) > 0
    has_walmart = len(marketplaces.filter(name__iexact="Walmart")) > 0
    response_data = {
        "selling_status": "Active",
        "asin/wpid": product_obj.product_id,
        "SKU": product_obj.sku,
        "Brand": product_obj.brand_name if product_obj.brand_name else "N/A",
        "date_range": product_obj.product_created_date.strftime("%b %d, %Y") + " - Current" if product_obj.product_created_date else "N/A - Current",
        "product_title": product_obj.product_title if product_obj.product_title else "N/A",
        "marketplaces": [],
        "ASIN" : "",
        "WPID" : ""
    }
    marketplace_name = getattr(product_obj.marketplace_id, 'name', None)
    if marketplace_name == "Amazon":
        response_data['ASIN'] = product_obj.product_id
    else:
        response_data['WPID'] = product_obj.product_id
    if has_amazon:
        response_data["marketplaces"].append({
        "name": "Amazon",
        "product_cost": round(product_obj.product_cost, 2) if product_obj.product_cost else 0,
        "shipping_cost": round(product_obj.a_shipping_cost, 2) if product_obj.a_shipping_cost else 0,
        "amazon_fee": round(product_obj.referral_fee, 2) if product_obj.referral_fee else 0,
        "total_cogs": round(product_obj.total_cogs, 2) if product_obj.total_cogs else 0,
        })
    if has_walmart:
        response_data["marketplaces"].append({
        "name": "Walmart",
        "product_cost": round(product_obj.w_product_cost, 2) if product_obj.w_product_cost else 0,
        "shipping_cost": round(product_obj.w_shiping_cost, 2) if product_obj.w_shiping_cost else 0,
        "walmart_fee": round(product_obj.walmart_fee, 2) if product_obj.walmart_fee else 0,
        "total_cogs": round(product_obj.w_total_cogs, 2) if product_obj.w_total_cogs else 0,
        })
    return response_data
@csrf_exempt
def updateProductDetails(request):
    json_request = JSONParser().parse(request)
    product_id = json_request.get('product_id')
    update_obj = json_request.get('update_obj', {})
    DatabaseModel.update_documents(Product.objects,{"id": product_id}, update_obj)
    return True
@csrf_exempt
def productUnitProfitability(request):
    json_request = JSONParser().parse(request)
    product_id = json_request.get("product_id")
    reponse_list =[]
    product_obj = DatabaseModel.get_document(Product.objects,{"id" : product_id})
    marketplaces = [ins.name for ins in product_obj.marketplace_ids]
    has_amazon = "Amazon" in marketplaces
    has_walmart = "Walmart" in marketplaces
    vendor_funding = round(product_obj.vendor_funding, 2) if product_obj.vendor_funding else 0
    if has_amazon:
        price = round(product_obj.price, 2) if product_obj.price else 0
        p_cost = round(product_obj.product_cost, 2) if product_obj.product_cost else 0
        s_cost = round(product_obj.a_shipping_cost, 2) if product_obj.a_shipping_cost else 0
        fee = round(product_obj.referral_fee, 2) if product_obj.referral_fee else 0
        reponse_list.append({
            "asin" : product_obj.product_id,
            "channel" : "Amazon",
            "channel_image" :  DatabaseModel.get_document(Marketplace.objects,{"name" : "Amazon"},['image_url']).image_url,
            "base_price" : price,
            "product_cost" : p_cost,
            "shipping_cost" : s_cost,
            "cogs" : round(p_cost + s_cost,2),
            "gross_profit" : round(price - (p_cost + s_cost), 2),
            "amazon_fee" : fee,
            "net_profit" : round((price - (p_cost + s_cost + fee)) + vendor_funding, 2)
        })
    if has_walmart:
        price = round(product_obj.price, 2) if product_obj.price else 0
        p_cost = round(product_obj.w_product_cost, 2) if product_obj.w_product_cost else 0
        s_cost = round(product_obj.w_shiping_cost, 2) if product_obj.w_shiping_cost else 0
        fee = round(product_obj.walmart_fee, 2) if product_obj.walmart_fee else 0
        reponse_list.append({
            "wpid" : product_obj.product_id,
            "channel" : "Walmart",
            "channel_image" : DatabaseModel.get_document(Marketplace.objects,{"name" : "Walmart"},['image_url']).image_url,
            "base_price" : price,
            "product_cost" : p_cost,
            "shipping_cost" : s_cost,
            "cogs" : round(p_cost + s_cost,2),
            "gross_profit" : round(price - (p_cost + s_cost), 2),
            "walmart_fee" : fee,
            "net_profit" : round((price - (p_cost + s_cost + fee)) + vendor_funding, 2)
        })
    return reponse_list
@csrf_exempt
def productNetprofit(request):
    json_request = JSONParser().parse(request)
    preset = json_request.get("preset", "Today")
    product_id = json_request.get("product_id")
    timezone_str = json_request.get('timezone', 'US/Pacific')
    start_date = json_request.get("start_date", None)
    end_date = json_request.get("end_date", None)
    if start_date != None and start_date != "":
        start_date, end_date = convertdateTotimezone(start_date,end_date,timezone_str)
    else:
        start_date, end_date = get_date_range(preset,timezone_str)
    if timezone_str != 'UTC':
        start_date,end_date = convertLocalTimeToUTC(start_date, end_date, timezone_str)
    product_obj = DatabaseModel.get_document(Product.objects,{"id" : ObjectId(product_id)})
    amazon_p_cost = round(product_obj.product_cost, 2) if product_obj.product_cost else 0
    amazon_s_cost = round(product_obj.a_shipping_cost, 2) if product_obj.a_shipping_cost else 0
    amazon_fee = round(product_obj.referral_fee, 2) if product_obj.referral_fee else 0
    walmart_p_cost = round(product_obj.w_product_cost, 2) if product_obj.w_product_cost else 0
    walmart_s_cost = round(product_obj.w_shiping_cost, 2) if product_obj.w_shiping_cost else 0
    walmart_fee = round(product_obj.walmart_fee, 2) if product_obj.walmart_fee else 0
    t_vendor_funding = round(product_obj.vendor_funding, 2) if product_obj.vendor_funding else 0
    def calculate_product_net_profit(product_id, start_date, end_date):
        gross_revenue = 0
        channel_fee = 0
        shipping_cost = 0
        product_cost = 0
        base_price = 0
        tax_price = 0
        vendor_funding = 0
        orders = grossRevenue(start_date, end_date, None, [], [product_id], [], None,timezone_str)
        for order in orders:
            gross_revenue += order.get("order_total", 0)
            total_units = order.get("items_order_quantity", 0)
            for item_id in order.get("order_items", []):
                item_pipeline = [
                    {"$match": {"_id": item_id}},
                    {
                        "$project": {
                            "price": { "$ifNull":["$Pricing.ItemPrice.Amount",0]},
                            "tax_price": { "$ifNull":["$Pricing.ItemTax.Amount",0]}
                        }
                    }
                ]
                item_result = list(OrderItems.objects.aggregate(*item_pipeline))
                if item_result:
                    item = item_result[0]
                    base_price += item.get("price", 0)
                    tax_price += item.get("tax_price", 0)
            if order['marketplace_name'] == "Amazon":
                shipping_cost += (amazon_s_cost * total_units)
                product_cost += (amazon_p_cost * total_units)
                channel_fee += (amazon_fee * total_units)
            else:
                shipping_cost += (walmart_s_cost * total_units)
                product_cost += (walmart_p_cost * total_units)
                channel_fee += (walmart_fee * total_units)
            vendor_funding += (t_vendor_funding * total_units)
        return {
            "gross_revenue" : round(gross_revenue,2),
            "tax_price" : round(tax_price, 2),
            "base_price": round(base_price, 2),
            "product_cost": round(product_cost, 2),
            "shipping_cost": round(shipping_cost, 2),
            "channel_fee": round(channel_fee, 2),
            "cogs": round(product_cost, 2) + round(shipping_cost, 2),
            "gross_profit": round(base_price - (product_cost + shipping_cost), 2),
            "net_profit": round((base_price - (product_cost + shipping_cost + channel_fee)) + vendor_funding, 2)
        }
    data = calculate_product_net_profit(product_id, start_date, end_date)
    return data
def cogsGraph(request):
    product_id = request.GET.get('product_id')
    product_obj = DatabaseModel.get_document(Product.objects,{"id" : product_id})
    marketplace_ids = product_obj.marketplace_ids
    response_list = [{
        "date_range": product_obj.product_created_date.strftime("%b %d, %Y") + " - Current" if product_obj.product_created_date else "N/A - Current",
        "product_cost": round(product_obj.product_cost, 2) if product_obj.product_cost else 0,
        "shipping_cost": round(product_obj.a_shipping_cost, 2) if product_obj.a_shipping_cost else 0,
        "amazon_fee": round(product_obj.referral_fee, 2) if product_obj.referral_fee else 0,
        "total_cogs": round(product_obj.total_cogs, 2) if product_obj.total_cogs else 0,
    }]
    return response_list

def priceGraph(request):
    product_id = request.GET.get('product_id')
    preset = request.GET.get('preset', 'Today')
    timezone_str = request.GET.get('timezone', 'US/Pacific')
    start_date, end_date = get_date_range(preset, timezone_str)
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]
    price_changes = productPriceChange.objects.filter(
        product_id=product_id,
        change_date__gte=start_date,
        change_date__lte=end_date
    ).order_by('change_date')
    price_change_dict = {change.change_date.date(): change.old_price for change in price_changes}
    product_obj = Product.objects.filter(id=product_id).first()
    current_price = product_obj.price if product_obj else 0.0
    response_data = []
    last_known_price = current_price  
    for date in date_range:
        price = price_change_dict.get(date.date(), last_known_price)
        response_data.append({
            "date": date.strftime('%b %d'),
            "price": round(price, 2)
        })
    return response_data
async def get_orders_by_brand_and_date(brands, start_date, end_date):
    from datetime import datetime
    from bson import ObjectId
    import pytz
    try:
        start_datetime, end_datetime = None, None
        pacific_tz = pytz.timezone("US/Pacific")
        if start_date:
            dt = datetime.strptime(start_date, '%Y-%m-%d')
            start_datetime = pacific_tz.localize(dt.replace(hour=0, minute=0, second=0))
        if end_date:
            dt = datetime.strptime(end_date, '%Y-%m-%d')
            end_datetime = pacific_tz.localize(dt.replace(hour=23, minute=59, second=59))

        match_query = {}

        date_conditions = {}
        if start_datetime:
            date_conditions["$gte"] = start_datetime
        if end_datetime:
            date_conditions["$lte"] = end_datetime
        if date_conditions:
            match_query["order_date"] = date_conditions

        # Part B: Add brand conditions (NOT nested)
        if brands:
            brand_object_ids = [ObjectId(b) for b in brands if len(b) == 24]
            products_in_brands = Product.objects(brand_id__in=brand_object_ids).only('id')
            product_ids_from_brands = [p.id for p in products_in_brands]
            order_items_with_products = OrderItems.objects(ProductDetails__product_id__in=product_ids_from_brands).only('id')
            order_item_ids = [oi.id for oi in order_items_with_products]
            match_query["order_items"] = {"$in": order_item_ids}

        pipeline = []
        
        if match_query:
            pipeline.append({"$match": match_query})

        pipeline.extend([
            {"$unwind": {"path": "$order_items", "preserveNullAndEmptyArrays": True}},
            {"$lookup": {"from": "order_items", "localField": "order_items", "foreignField": "_id", "as": "order_item_details"}},
            {"$unwind": {"path": "$order_item_details", "preserveNullAndEmptyArrays": True}},
            {"$lookup": {"from": "product", "localField": "order_item_details.ProductDetails.product_id", "foreignField": "_id", "as": "product_info"}},
            {"$unwind": {"path": "$product_info", "preserveNullAndEmptyArrays": True}},
            {"$lookup": {"from": "brand", "localField": "product_info.brand_id", "foreignField": "_id", "as": "brand_info"}},
            {"$unwind": {"path": "$brand_info", "preserveNullAndEmptyArrays": True}},
            {"$group": {
                "_id": "$_id",
                "purchase_order_id": {"$first": "$purchase_order_id"},
                "order_date": {"$first": "$order_date"},
                "order_status": {"$first": "$order_status"},
                "marketplace_id": {"$first": "$marketplace_id"},
                 "shipping_price":    {"$first": "$shipping_price"},
                "shipping_tax":      {"$first": "$shipping_tax"},
                "brand_names": {"$addToSet": "$brand_info.name"},
                "skus":{"$addToSet":'$order_item_details.ProductDetails.SKU'},
                'total_quantity':{"$sum":"$order_item_details.ProductDetails.QuantityOrdered"},
                "order_total": {"$first": "$order_total"},
                "unit_prices":{"$addToSet":"$order_item_details.Pricing.ItemPrice.Amount"},
                "item_taxes": {"$sum": {"$ifNull": ["$order_item_details.Pricing.ItemTax.Amount", 0]}},
                "subtotal": {"$sum": {"$multiply": ["$order_item_details.ProductDetails.QuantityOrdered", "$order_item_details.Pricing.ItemPrice.Amount"]}}
            }},
            {"$lookup": {"from": "marketplace", "localField": "marketplace_id", "foreignField": "_id", "as": "marketplace_info"}},
            {"$project": {
                "_id": 0,
                "order_id": {"$toString": "$_id"},
                "purchase_order_id": "$purchase_order_id",
                "order_date": "$order_date",
                
                "order_status": "$order_status",
                "marketplace_name": {"$ifNull": [{"$arrayElemAt": ["$marketplace_info.name", 0]}, "Unknown"]},
                "brand_name": {
                    "$reduce": {
                        "input": "$brand_names", "initialValue": "", "in": {
                            "$cond": [ {"$eq": ["$$value", ""]}, "$$this", {"$concat": ["$$value", ", ", "$$this"]} ]
                        }
                    }
                },
                "sku": {
    "$reduce": {
        "input": "$skus", "initialValue": "", "in": {
            "$cond": [ {"$eq": ["$$value", ""]}, "$$this", {"$concat": ["$$value", ", ", "$$this"]} ]
        }
    }
},
"total_quantity": "$total_quantity",
"unit_price": {
    "$reduce": {
        "input": "$unit_prices", "initialValue": "", "in": {
            "$cond": [ {"$eq": ["$$value", ""]}, {"$toString": "$$this"}, {"$concat": ["$$value", ", ", {"$toString": "$$this"}]} ]
        }
    }
},
"subtotal":"$subtotal",
'tax':"$item_taxes",
"shipping_price":{"$ifNull":["$shipping_price",0]},
'shipping_tax':{"$ifNull":["$shipping_tax",0]},
"order_total": "$order_total",
            }},
            {"$sort": {"order_date": -1}}
        ])
        
        orders = list(Order.objects.aggregate(*pipeline))
        pacific_tz = pytz.timezone("US/Pacific")
        for order in orders:
            if order.get('order_date'):
                try:
                    if order['order_date'].tzinfo is None:
                        order['order_date'] = pytz.utc.localize(order['order_date'])
                    order['order_date'] = order['order_date'].astimezone(pacific_tz)
                    order['order_date'] = order['order_date'].strftime('%Y-%m-%d %H:%M:%S %Z')
                except Exception as e:
                    print(f"Error converting date for order {order.get('order_id')}: {e}")
                    order['order_date'] = str(order['order_date']) if order['order_date'] else ""
            if order.get('expected_delivery_date'):
                try:
                    if order['expected_delivery_date'].tzinfo is None:
                        order['expected_delivery_date'] = pytz.utc.localize(order['expected_delivery_date'])
                    order['expected_delivery_date'] = order['expected_delivery_date'].astimezone(pacific_tz)
                    order['expected_delivery_date'] = order['expected_delivery_date'].strftime('%Y-%m-%d %H:%M:%S %Z')
                except Exception as e:
                    print(f"Error converting expected delivery date for order {order.get('order_id')}: {e}")
                    order['expected_delivery_date'] = str(order['expected_delivery_date']) if order['expected_delivery_date'] else ""
        return orders
    except Exception as e:
        print(f"Error in get_orders_by_brand_and_date: {e}")
        return []
    
async def get_all_orders_by_brand_and_date(brands, start_date, end_date, include_custom=False):
    regular_orders = await get_orders_by_brand_and_date(brands, start_date, end_date)
    if not include_custom:
        return regular_orders
    try:
        from datetime import datetime
        import pytz
        pipeline = []
        if start_date or end_date:
            date_match = {}
            pacific_tz = pytz.timezone("US/Pacific")
            if start_date:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
                start_datetime = pacific_tz.localize(start_datetime.replace(hour=0, minute=0, second=0))
                date_match["$gte"] = start_datetime
            if end_date:
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d')
                end_datetime = pacific_tz.localize(end_datetime.replace(hour=23, minute=59, second=59))
                date_match["$lte"] = end_datetime
            pipeline.append({
                "$match": {
                    "purchase_order_date": date_match
                }
            })
        pipeline.append({
            "$project": {
                "_id": 0,
                "order_id": {"$toString": "$_id"},
                "purchase_order_id": "$order_id",  
                "customer_name": "$customer_name",
                "customer_email": {"$ifNull": ["$customer_email", ""]},
                "shipping_address": "$shipping_address",
                "order_date": "$purchase_order_date",
                "order_status": "$order_status",
                
                "currency": {"$ifNull": ["$currency", "USD"]},
                "total_quantity": "$total_quantity",
                "items_order_quantity": {"$ifNull": ["$items_order_quantity", "$total_quantity"]},
                "marketplace_name": "Custom Order",
                "brand_name": "Custom",
                "expected_delivery_date": "$expected_delivery_date",
                "tracking_number": {"$ifNull": ["$tracking_number", ""]},
                "sku": {
    "$reduce": {
        "input": "$ordered_products.sku",
        "initialValue": "",
        "in": {
            "$cond": [
                {"$eq": ["$$value", ""]},
                "$$this",
                {"$concat": ["$$value", ", ", "$$this"]}
            ]
        }
    },
    "unit_price": {
    "$reduce": {
        "input": "$ordered_products.unit_price",
        "initialValue": "",
        "in": {
            "$cond": [
                {"$eq": ["$$value", ""]},
                {"$toString": "$$this"},
                {"$concat": ["$$value", ", ", {"$toString": "$$this"}]}
            ]
        }
    }
},
"subtotal":{"$sum":"$ordered_products.quantity_price"},
"tax":"$tax_amount",
"shipping_price":"$shipment_cost",
"shipping_tax":0,
"order_total": "$total_price",
}
            }
        })
        pipeline.append({
            "$sort": {
                "order_date": -1
            }
        })
        custom_orders = list(custom_orders.objects.aggregate(*pipeline))
        pacific_tz = pytz.timezone("US/Pacific")
        for order in custom_orders:
            if order.get('order_date'):
                try:
                    if order['order_date'].tzinfo is None:
                        order['order_date'] = pytz.utc.localize(order['order_date'])
                    order['order_date'] = order['order_date'].astimezone(pacific_tz)
                    order['order_date'] = order['order_date'].strftime('%Y-%m-%d %H:%M:%S %Z')
                except Exception as e:
                    print(f"Error converting date for custom order {order.get('order_id')}: {e}")
                    order['order_date'] = str(order['order_date']) if order['order_date'] else ""
        all_orders = regular_orders + custom_orders
        all_orders.sort(key=lambda x: x.get('order_date', ''), reverse=True)
        return all_orders
    except Exception as e:
        print(f"Error getting custom orders: {e}")
        return regular_orders  
@csrf_exempt
def downloadOrders(request):
    try:
        data = json.loads(request.body)
        brands = data.get('brands', [])
        start_date = data.get('start_date')
        end_date = data.get("end_date")
        file_format = data.get("format", 'csv')
        orders = asyncio.run(get_all_orders_by_brand_and_date(brands, start_date, end_date))
        if not orders:
            return {'error': "No orders for the given filters"}
        df = pd.DataFrame(orders)
        if file_format == 'csv':
            output = io.StringIO()
            df.to_csv(output, index=False)
            response = HttpResponse(output.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename=orders.csv'
            return response 
        elif file_format == 'xlsx':
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, sheet_name='Orders')
            output.seek(0)
            response = HttpResponse(
                output.read(), 
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
            return response 
        elif file_format=='txt':
            output=io.StringIO()
            df.to_csv(output,index=False,sep='\t')
            response=HttpResponse(output.getvalue(),content_type='text/plain')
            response['Content-Disposition'] = 'attachment; filename=orders.txt'
            return response 
        else:
            return {'error': "Invalid format"}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'error': 'An internal server error occurred.'}