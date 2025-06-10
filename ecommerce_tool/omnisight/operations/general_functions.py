import requests
import uuid
import xml.etree.ElementTree as ET  # Import XML parser
from omnisight.operations.walmart_utils import getAccesstoken
from omnisight.models import *
import pandas as pd
from ecommerce_tool.crud import DatabaseModel
from bson import ObjectId
import json
from rest_framework.parsers import JSONParser
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from omnisight.operations.helium_utils import get_date_range




def getMarketplaceList(request):
    pipeline = [
        {
            "$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "name" : 1,
                "image_url" : 1,
            }
        }
    ]
    marketplace_list = list(Marketplace.objects.aggregate(*(pipeline)))
    for marketplace_ins  in marketplace_list:
        if marketplace_ins['name'] == "Amazon":
            marketplace_ins['fulfillment_channel'] = [{'FBA' : "AFN"},{'FBM' : "MFN"}]
        elif marketplace_ins['name'] == "Walmart":
            marketplace_ins['fulfillment_channel'] = [{'FBM' : "SellerFulfilled"}]
    return marketplace_list

#---------------------------------------------PRODUCT APIS---------------------------------------------------
@csrf_exempt
def getProductList(request):
    pipeline = [
            {
                "$project" : {
                    "_id" : 1,
                    "name" : 1,
                    "image_url" : 1,
                }
            }
        ]
    marketplace_list = list(Marketplace.objects.aggregate(*(pipeline)))
    data = dict()
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    skip = int(json_request.get('skip'))
    limit = int(json_request.get('limit'))
    search_query = json_request.get('search_query')   
    marketplace = json_request.get('marketplace')
    category_name = json_request.get('category_name')
    brand_id_list = json_request.get('brand_id_list')
    sort_by = json_request.get('sort_by')
    sort_by_value = json_request.get('sort_by_value')
    pipeline = []
    count_pipeline = []
    match = {}
    if marketplace_id != None and marketplace_id != "":
        match['marketplace_ids'] = {"$in":[ObjectId(marketplace_id)]}
    if category_name != None and category_name != "" and category_name != []:
        match['category'] = {"$in":category_name}
    if brand_id_list != None and brand_id_list != "" and brand_id_list != []:
        match['brand_id'] = {"$in":[ObjectId(brand_id) for brand_id in brand_id_list]}
    if search_query != None and search_query != "":
        search_query = search_query.strip() 
        match["$or"] = [
            {"product_title": {"$regex": search_query, "$options": "i"}},
            {"sku": {"$regex": search_query, "$options": "i"}},
        ]
    if match != {}:
        match_pipeline = {
            "$match" : match}
        pipeline.append(match_pipeline)
        count_pipeline.append(match_pipeline)
    pipeline.extend([
        {
            "$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "product_title" : 1,
                "product_id" : 1,
                "sku" : 1,
                "asin" : {"$ifNull" : ["$asin",""]},  # If asin is null, replace with empty string
                "price" : 1,
                "quantity" : 1,
                "published_status" : 1,
                "category" : {"$ifNull" : ["$category",""]},  # If category is null, replace with empty string
                "image_url" : {"$ifNull" : ["$image_url",""]},  # If image_url is null, replace with empty string
                "marketplace_ids": {"$ifNull": ["$marketplace_ids", []]},  # If marketplace_ids is null, replace with empty list
                "marketplace_ins": [],
                "marketplace_image_url": [] 
              
            }
        },
        {
            "$skip" : skip
        },
        {
            "$limit" : limit  # Ensure limit is applied correctly
        }
    ])
    if sort_by != None and sort_by != "":
        sort = {
            "$sort" : {
                sort_by : int(sort_by_value)
            }
        }
        pipeline.append(sort)
    product_list = list(Product.objects.aggregate(*(pipeline)))
    for ins in product_list:
        marketplace_ids = ins.get('marketplace_ids', [])
        ins['marketplace_details'] = []
        for marketplace in marketplace_list:
            if marketplace['_id'] in marketplace_ids:
                ins['marketplace_ins'].append(marketplace['name'])
                ins['marketplace_image_url'].append(marketplace['image_url'])
        del ins['marketplace_ids']
    count_pipeline.extend([
        {
            "$count": "total_count"
        }
    ])
    total_count_result = list(Product.objects.aggregate(*(count_pipeline)))
    total_count = total_count_result[0]['total_count'] if total_count_result else 0

    data['total_count'] = total_count
    data['product_list'] = product_list
    return data



def getProductCategoryList(request):
    data = dict()
    marketplace_id = request.GET.get('marketplace_id')
    match = {}
    if marketplace_id != None and marketplace_id != "":
        match['marketplace_id'] = ObjectId(marketplace_id)
    match['end_level'] = True
    pipeline = [
        {
            "$match": match    
        },
        {
            "$lookup": {
                "from": "product",
                "localField": "name",
                "foreignField": "category",
                "as": "products"
            }
        },
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "name": 1,
                "product_count": {"$size": "$products"}
            }
        },
        {
            "$match": {
                "product_count": {"$gt": 0}
            }
        }
    ]
    category_list = list(Category.objects.aggregate(*(pipeline)))
    data['category_list'] = category_list
    return data



def getBrandList(request):
    data = dict()
    marketplace_id = request.GET.get('marketplace_id')
    pipeline = []
    if marketplace_id != None and marketplace_id != "":
        match = {
            "$match" : {
                "marketplace_id" : ObjectId(marketplace_id)
            }
        }
        pipeline.append(match)
    pipeline.extend([
        {
            "$lookup": {
                "from": "product",
                "localField": "_id",
                "foreignField": "brand_id",
                "as": "products"
            }
        },
        {
            "$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "name" : 1,
                "product_count": {"$size": "$products"}
            }
        }
    ])
    brand_list = list(Brand.objects.aggregate(*(pipeline)))
    data['brand_list'] = brand_list
    return data


def fetchProductDetails(request):
    data = dict()
    product_id = request.GET.get('product_id')
    pipeline = [
        {
            "$match": {
                "_id": ObjectId(product_id)
            }
        },
        {
            "$lookup": {
                "from": "marketplace",
                "localField": "marketplace_ids",
                "foreignField": "_id",
                "as": "marketplace_ins"
            }
        },
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "product_title": {"$ifNull": ["$product_title", ""]},
                "product_description": {"$ifNull": ["$product_description", ""]},
                "product_id": {"$ifNull": ["$product_id", ""]},
                "product_id_type": {"$ifNull": ["$product_id_type", ""]},
                "sku": {"$ifNull": ["$sku", ""]},
                "price": {"$round": [{"$ifNull": ["$price", 0]}, 2]},
                "currency": {"$ifNull": ["$currency", ""]},
                "quantity": {"$ifNull": ["$quantity", 0]},
                "published_status": 1,
                "marketplace_ins": {
                    "$map": {
                        "input": "$marketplace_ins",
                        "as": "marketplace",
                        "in": "$$marketplace.name"
                    }
                },
                "marketplace_image_url": {
                    "$map": {
                        "input": "$marketplace_ins",
                        "as": "marketplace",
                        "in": "$$marketplace.image_url"
                    }
                },
                "item_condition": {"$ifNull": ["$item_condition", ""]},
                "item_note": {"$ifNull": ["$item_note", ""]},
                "listing_id": {"$ifNull": ["$listing_id", ""]},
                "upc": {"$ifNull": ["$upc", ""]},
                "gtin": {"$ifNull": ["$gtin", ""]},
                "asin": {"$ifNull": ["$asin", ""]},
                "model_number": {"$ifNull": ["$model_number", ""]},
                "category": {"$ifNull": ["$category", ""]},
                "brand_name": {"$ifNull": ["$brand_name", ""]},
                "manufacturer_name": {"$ifNull": ["$manufacturer_name", ""]},
                "attributes": {"$ifNull": ["$attributes", {}]},
                "features": {"$ifNull": ["$features", []]},
                "shelf_path": {"$ifNull": ["$shelf_path", ""]},
                "image_url": {"$ifNull": ["$image_url", ""]},
                "image_urls": {"$ifNull": ["$image_urls", []]},
                "vendor_funding": {"$round": [{"$ifNull": ["$vendor_funding", 0.0]}, 2]},
                "veddor_discount": {"$round": [{"$ifNull": ["$vendor_discount", 0.0]}, 2]},
                "product_cost": {"$round": [{"$ifNull": ["$product_cost", 0]}, 2]},
                "referral_fee": {"$round": [{"$ifNull": ["$referral_fee", 0]}, 2]},
                "a_shipping_cost": {"$round": [{"$ifNull": ["$a_shipping_cost", 0]}, 2]},
                "total_cogs": {"$round": [{"$ifNull": ["$total_cogs", 0]}, 2]},
                "w_product_cost": {"$round": [{"$ifNull": ["$w_product_cost", 0]}, 2]},
                "walmart_fee": {"$round": [{"$ifNull": ["$walmart_fee", 0]}, 2]},
                "w_shiping_cost": {"$round": [{"$ifNull": ["$w_shiping_cost", 0]}, 2]},
                "w_total_cogs": {"$round": [{"$ifNull": ["$w_total_cogs", 0]}, 2]},
                "pack_size": {"$ifNull": ["$pack_size", ""]},
            }
        }
    ]
    product_details = list(Product.objects.aggregate(*(pipeline)))
    if len(product_details):
        data = product_details[0]
    return data


def getOrdersBasedOnProduct(request):
    data = dict()
    product_id = request.GET.get('product_id')
    skip = int(request.GET.get('skip',0))  # Default skip = 0 if not provided
    limit = int(request.GET.get('limit', 10))  # Default limit = 100 if not provided

    count_pipeline = [
        {
            "$match": {
                "order_items": {"$exists": True, "$ne": []}
            }
        },
        {
            "$lookup": {
                "from": "order_items",
                "localField": "order_items",
                "foreignField": "_id",
                "as": "order_items"
            }
        },
        {
            "$unwind": "$order_items"
        },
        {
            "$match": {
                "order_items.ProductDetails.product_id": ObjectId(product_id)
            }
        },
        {
            "$count": "total_count"
        }
    ]
    total_count_result = list(Order.objects.aggregate(*count_pipeline))
    total_count = total_count_result[0]['total_count'] if total_count_result else 0

    pipeline = [
        {
            "$match": {
                "order_items": {"$exists": True, "$ne": []}
            }
        },
        {
            "$lookup": {
                "from": "order_items",
                "localField": "order_items",
                "foreignField": "_id",
                "as": "order_items_ins"
            }
        },
        {
            "$unwind": "$order_items_ins"
        },
        {
            "$match": {
                "order_items_ins.ProductDetails.product_id": ObjectId(product_id)
            }
        },
        {
            "$sort": {
                "order_date": -1
            }
        },
        {
            "$skip": skip
        },
        {
            "$limit": limit
        },
        {
            "$group": {
                "_id": "$_id",
                "purchase_order_id": {"$first": "$purchase_order_id"},
                "order_date": {"$first": "$order_date"},
                "order_status": {"$first": "$order_status"},
                "order_total": {"$first": "$order_total"},
                "currency": {"$first": "$currency"},
                "marketplace_id": {"$first": "$marketplace_id"}
            }
        },
        {
            "$lookup": {
                "from": "marketplace",
                "localField": "marketplace_id",
                "foreignField": "_id",
                "as": "marketplace_ins"
            }
        },
        {
            "$unwind": {
                "path": "$marketplace_ins",
                "preserveNullAndEmptyArrays": True
            }
        },
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "purchase_order_id": 1,
                "order_date": {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": "$order_date",
                    }
                },
                "order_status": 1,
                "order_total": {"$round": ["$order_total", 2]},
                "currency": 1,
                "marketplace_name": "$marketplace_ins.name"
            }
        }
    ]
    orders = list(Order.objects.aggregate(*pipeline, allowDiskUse=True))

    data['total_count'] = total_count
    data['orders'] = orders
    custom_pipeline = [
        {
            "$match": {
                "ordered_products.product_id" : ObjectId(product_id)
            }
        },
        {
            "$project": {
                "_id": 0,
                "id": {"$toString" : "$_id"},
                "purchase_order_id": "$order_id",
                "order_date": "$purchase_order_date",
                "order_status": 1,
                "order_total": "$total_price",
                "currency": {"$ifNull" : ["$currency","USD"]},
                "marketplace_name": "custom"
            }
        }
    ]
    custom_orders = list(custom_order.objects.aggregate(*custom_pipeline))
    orders.extend(custom_orders)

    data['orders'] = orders
    return data



def getOrdersBasedOnProduct(request):
    data = {'total_count': 0, 'orders': []}
    product_id = request.GET.get('product_id')
    skip = int(request.GET.get('skip', 0))
    limit = int(request.GET.get('limit', 10))

    # Step 1: Get relevant order_item IDs based on product_id
    matching_order_items = list(OrderItems.objects.filter(ProductDetails__product_id=ObjectId(product_id)).only('id'))

    if not matching_order_items:
        return data

    # Step 2: Count total relevant orders
    total_count = Order.objects.filter(order_items__in=matching_order_items).count()

    # Step 3: Fetch orders with pagination
    orders_queryset = Order.objects.filter(order_items__in=matching_order_items)\
        .order_by("-order_date")\
        .skip(skip).limit(limit - skip)

    # Step 4: Convert to dictionary and fetch marketplace in one go
    order_list = list(orders_queryset)
    marketplace_ids = [o.marketplace_id.id for o in order_list if o.marketplace_id]
    marketplaces = {
        m.id: m.name for m in Marketplace.objects.filter(id__in=marketplace_ids)
    }

    # Step 5: Format response
    orders = []
    for o in order_list:
        orders.append({
            "id": str(o.id),
            "purchase_order_id": o.purchase_order_id,
            "order_date": o.order_date.strftime('%Y-%m-%dT%H:%M:%S.%fZ')[:-3],
            "order_status": o.order_status,
            "order_total": round(o.order_total, 2) if o.order_total else 0.0,
            "currency": o.currency,
            "marketplace_name": marketplaces.get(o.marketplace_id.id, None),
        })

    # Step 6: Append custom orders if needed
    custom_orders = list(custom_order.objects.filter(
        ordered_products__product_id =  ObjectId(product_id)
    ).only("order_id", "purchase_order_date", "order_status", "total_price", "currency"))

    for c in custom_orders:
        orders.append({
            "id": str(c.id),
            "purchase_order_id": c.order_id,
            "order_date": c.purchase_order_date.strftime('%Y-%m-%dT%H:%M:%S.%fZ')[:-3],
            "order_status": c.order_status,
            "order_total": round(c.total_price, 2) if c.total_price else 0.0,
            "currency": c.currency or "USD",
            "marketplace_name": "custom"
        })
    data = {
        "total_count": total_count,
        "orders": orders
    }
    return data



#---------------------------------ORDER APIS-------------------------------------------------------------------


@csrf_exempt
def fetchAllorders(request):
    data = dict()
    orders = []
    pipeline = []
    count_pipeline = []

    json_request = JSONParser().parse(request)
    user_id = json_request.get('user_id')
    limit = int(json_request.get('limit', 100))  # Default limit = 100 if not provided
    skip = int(json_request.get('skip', 0))  # Default skip = 0 if not provided
    market_place_id = json_request.get('marketplace_id')
    sort_by = json_request.get('sort_by')
    sort_by_value = json_request.get('sort_by_value')
    search_query = json_request.get('search_query')
        
    if market_place_id != None and market_place_id != "" and market_place_id != "all" and market_place_id == "custom":
        search_query = search_query.strip() 
        match = { "$match" : 
                    {"order_id": {"$regex": search_query, "$options": "i"}}}
        pipeline.append(match)
        pipeline = [
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "order_id": {"$ifNull": ["$order_id", ""]},
                "customer_name": {"$ifNull": ["$customer_name", ""]},
                "shipping_address": {"$ifNull": ["$shipping_address", ""]},
                "total_quantity": {"$ifNull": ["$total_quantity", 0]},
                "total_price": {"$ifNull": [{"$round": ["$total_price", 2]}, 0.0]},
                # "taxes": {"$ifNull": ["$taxes", 0.0]},
                "purchase_order_date": {"$ifNull": ["$purchase_order_date", None]},
                "expected_delivery_date": {"$ifNull": ["$expected_delivery_date", None]},
                "order_status" : "$order_status",
                "currency" : {"$ifNull" : ["$currency","USD"]}
            }
        }
        ]
        if sort_by != None and sort_by != "":
            sort = {
                "$sort" : {
                    sort_by : int(sort_by_value)
                }
            }
        else:
            sort = {
                "$sort" : {
                    "id" : -1
                }
            }
        pipeline.append(sort)
        pipeline.extend([
            {
            "$skip": skip
        },
        {
            "$limit": limit
        }
        ])

        manual_orders = list(custom_order.objects.aggregate(*pipeline))
        count_pipeline = [
            {
                "$count": "total_count"
            }
        ]
        total_count_result = list(custom_order.objects.aggregate(*(count_pipeline)))
        total_count = total_count_result[0]['total_count'] if total_count_result else 0
        data['total_count'] = total_count
        data['manual_orders'] = manual_orders
        data['status'] = "custom"

    elif market_place_id != None and market_place_id != "" and market_place_id != "all" and market_place_id != "custom":
        match = {
            "$match": {
                "marketplace_id": ObjectId(market_place_id)
            }
        }
        pipeline.append(match)
        count_pipeline.append(match)
    if search_query != None and search_query != "":
        search_query = search_query.strip() 
        match = { "$match" : 
                    {"purchase_order_id": {"$regex": search_query, "$options": "i"}}}
            # {"sku": {"$regex": search_query, "$options": "i"}},
        pipeline.append(match)
        count_pipeline.append(match)
    if market_place_id != "custom":
        if sort_by != None and sort_by != "":
            sort = {
                "$sort" : {
                    sort_by : int(sort_by_value)
                }
            }
        else:
            sort =  {
                "$sort" : {
                    "order_date" : -1
                }
            }
        pipeline.append(sort)
        pipeline.extend([
            {
            "$skip": skip
        },
        {
            "$limit": limit
        }
        ])
        pipeline.extend([

            {
            "$lookup": {
                "from": "marketplace",
                "localField": "marketplace_id",
                "foreignField": "_id",
                "as": "marketplace_ins"
            }
            },
            {
            "$unwind": "$marketplace_ins"
            },
            {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "purchase_order_id": "$purchase_order_id",
                "order_date": "$order_date",
                "order_status": "$order_status",
                "order_total": "$order_total",
                "currency": "$currency",
                "marketplace_name": "$marketplace_ins.name",
                "items_order_quantity": "$items_order_quantity"
            }
            }
        ])
        
        orders = list(Order.objects.aggregate(*(pipeline)))
        count_pipeline.extend([
            {
                "$count": "total_count"
            }
        ])
        total_count_result = list(Order.objects.aggregate(*(count_pipeline)))
        total_count = total_count_result[0]['total_count'] if total_count_result else 0
        
        data['orders'] = orders
        data['total_count'] = total_count
        data['status'] = ""

    pipeline = [
            {
                "$project" : {
                    "_id" : 0,
                    "id" : {"$toString" : "$_id"},
                    "name" : 1,
                    "image_url" : 1,
                }
            }
        ]
    data['marketplace_list'] = list(Marketplace.objects.aggregate(*(pipeline)))
    return data



def fetchOrderDetails(request):
    data = {}
    user_id = request.GET.get('user_id')
    order_id = request.GET.get('order_id')

    pipeline = [
        {
            "$match": {
                "_id": ObjectId(order_id)
            }
        },
        {
            "$lookup": {
                "from": "marketplace",
                "localField": "marketplace_id",
                "foreignField": "_id",
                "as": "marketplace_ins"
            }
        },
        {
            "$unwind": "$marketplace_ins"
        },
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "purchase_order_id": {"$ifNull": ["$purchase_order_id", ""]},
                "customer_order_id": {"$ifNull": ["$customer_order_id", ""]},
                "seller_order_id": {"$ifNull": ["$seller_order_id", ""]},
                "customer_email_id": {"$ifNull": ["$customer_email_id", ""]},
                "order_date": {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": {"$ifNull": ["$order_date", None]},
                    }
                },
                "earliest_ship_date": {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": {"$ifNull": ["$earliest_ship_date", None]},
                    }
                },
                "latest_ship_date": {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": {"$ifNull": ["$latest_ship_date", None]},
                    }
                },
                "last_update_date": {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": {"$ifNull": ["$last_update_date", None]},
                    }
                },
                "shipping_information": {"$ifNull": ["$shipping_information", {}]},
                "ship_service_level": {"$ifNull": ["$ship_service_level", ""]},
                "shipment_service_level_category": {"$ifNull": ["$shipment_service_level_category", ""]},
                "automated_shipping_settings": {"$ifNull": ["$automated_shipping_settings", {}]},
                "order_status": {"$ifNull": ["$order_status", ""]},
                "number_of_items_shipped": {"$ifNull": ["$number_of_items_shipped", 0]},
                "number_of_items_unshipped": {"$ifNull": ["$number_of_items_unshipped", 0]},
                "fulfillment_channel": {"$ifNull": ["$fulfillment_channel", ""]},
                "sales_channel": {"$ifNull": ["$sales_channel", ""]},
                "order_type": {"$ifNull": ["$order_type", ""]},
                "is_premium_order": {"$ifNull": ["$is_premium_order", False]},
                "is_prime": {"$ifNull": ["$is_prime", False]},
                "has_regulated_items": {"$ifNull": ["$has_regulated_items", False]},
                "is_replacement_order": {"$ifNull": ["$is_replacement_order", False]},
                "is_sold_by_ab": {"$ifNull": ["$is_sold_by_ab", False]},
                "is_ispu": {"$ifNull": ["$is_ispu", False]},
                "is_access_point_order": {"$ifNull": ["$is_access_point_order", False]},
                "is_business_order": {"$ifNull": ["$is_business_order", False]},
                "marketplace_name": {"$ifNull": ["$marketplace_ins.name", ""]},
                "payment_method": {"$ifNull": ["$payment_method", ""]},
                "payment_method_details": {"$ifNull": ["$payment_method_details", []]},
                "order_total": {"$ifNull": [{"$round": ["$order_total", 2]}, 0]},
                "currency": {"$ifNull": ["$currency", ""]},
                "is_global_express_enabled": {"$ifNull": ["$is_global_express_enabled", False]},
                "order_items": {"$ifNull": ["$order_items", []]}
            }
        }
    ]

    order_details = list(Order.objects.aggregate(*pipeline))

    if order_details:
        
        data = order_details[0]
        # Fetch OrderItems details using the IDs
        order_items_ids = data.get("order_items", [])
        if order_items_ids:
            order_items = DatabaseModel.list_documents(OrderItems.objects,{"id__in" : order_items_ids})
            
            serialized_items = []
            for item in order_items:
                serialized_item = {
                    "id": str(item.id),
                    "OrderId": item.OrderId,
                    "Platform": item.Platform,
                    "ProductDetails": {
                        "product_id": str(item.ProductDetails.product_id.id) if item.ProductDetails and item.ProductDetails.product_id else None,
                        "Title": item.ProductDetails.Title if item.ProductDetails else None,
                        "SKU": item.ProductDetails.SKU if item.ProductDetails else None,
                        "Condition": item.ProductDetails.Condition if item.ProductDetails else None,
                        "QuantityOrdered": item.ProductDetails.QuantityOrdered if item.ProductDetails else 0,
                        "QuantityShipped": item.ProductDetails.QuantityShipped if item.ProductDetails else 0
                    },
                    "Pricing": item.Pricing.to_mongo() if hasattr(item.Pricing, "to_mongo") else (dict(item.Pricing) if item.Pricing else {}),
                    "Fulfillment": {
                        "FulfillmentOption": item.Fulfillment.FulfillmentOption if item.Fulfillment else None,
                        "ShipMethod": item.Fulfillment.ShipMethod if item.Fulfillment else None,
                        "Carrier": item.Fulfillment.Carrier if item.Fulfillment else None,
                        "TrackingNumber": item.Fulfillment.TrackingNumber if item.Fulfillment else None,
                        "TrackingURL": item.Fulfillment.TrackingURL if item.Fulfillment else None,
                        "ShipDateTime": item.Fulfillment.ShipDateTime.isoformat() if item.Fulfillment and item.Fulfillment.ShipDateTime else None
                    },
                    "OrderStatus": {
                        "Status": item.OrderStatus.Status if item.OrderStatus else None,
                        "StatusDate": item.OrderStatus.StatusDate.isoformat() if item.OrderStatus and item.OrderStatus.StatusDate else None
                    },
                    "TaxCollection": item.TaxCollection.to_mongo() if hasattr(item.TaxCollection, "to_mongo") else (dict(item.TaxCollection) if item.TaxCollection else {}),
                    "IsGift": item.IsGift
                }
                serialized_items.append(serialized_item)
            
            data["order_items"] = serialized_items
        else:
            data["order_items"] = []

    return data


#----------------------ORDER CREATION---------------------------------------------------------------------------


@csrf_exempt
def getProductListForOrdercreation(request):
    data = dict()
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    skip = int(json_request.get('skip',0))
    limit = int(json_request.get('limit',50))
    search_query = json_request.get('search_query')
    pipeline = []
    count_pipeline = []
    match = {}
    if marketplace_id != None and marketplace_id != "":
        match['marketplace_id'] = ObjectId(marketplace_id)
    if search_query != None and search_query != "":
        match['product_title'] = {"$regex": search_query, "$options": "i"}
    if match != {}:
        match_pipeline = {
            "$match" : match}
        pipeline.append(match_pipeline)
    pipeline.extend([
        {
            "$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "product_title" : 1,
                "sku" : 1,
                "price" : 1,
            }
        },
        {
            "$skip" : skip
        },
        {
            "$limit" : limit
        }
    ])
    product_list = list(Product.objects.aggregate(*(pipeline)))
    count_pipeline.extend([
        {
            "$count": "total_count"
        }
    ])
    total_count_result = list(Product.objects.aggregate(*(count_pipeline)))
    total_count = total_count_result[0]['total_count'] if total_count_result else 0

    data['total_count'] = total_count
    
    return product_list


@csrf_exempt
def createManualOrder(request):
    data = dict()
    json_request = JSONParser().parse(request)
    product_detail = list()
    ordered_products = json_request.get('ordered_products')
    custom_product_obj = json_request.get('custom_product_obj')
    user_id = json_request.get('user_id')
    if user_id:
        custom_product_obj['user_id'] = ObjectId(user_id)
    
    custom_product_obj['order_id'] = str(''.join([str(uuid.uuid4().int)[:13]]))
     
    custom_product_obj['customer_order_id'] = datetime.now().strftime('%Y%m%d') + str(uuid.uuid4().int)[:5]
    # Handle purchase_order_date
    try:
        custom_product_obj['purchase_order_date'] = datetime.strptime(
            json_request.get(custom_product_obj['purchase_order_date']), '%Y-%m-%d'
        )
    except:
        pass
    
    expected_delivery_date = custom_product_obj.get('expected_delivery_date')
    if expected_delivery_date:
        custom_product_obj['expected_delivery_date'] = datetime.strptime(expected_delivery_date, '%Y-%m-%d')
    else:
        custom_product_obj['expected_delivery_date'] = None
  

    # Process ordered products
    for ins in ordered_products:
        product_detail_dict = product_details(
            product_id=ObjectId(ins.get('product_id')),
            title=ins.get('title'),
            sku=ins.get('sku'),
            unit_price=float(ins.get('unit_price', 0)),
            quantity=int(ins.get('quantity', 0)),
            quantity_price=float(ins.get('quantity_price', 0))
        )
        product_detail.append(product_detail_dict)

    custom_product_obj['ordered_products'] = product_detail

    # Calculate total_price with discount and tax
    total_price = float(custom_product_obj.get('total_price', 0))
    discount = float(custom_product_obj.get('discount', 0))  # Discount in percentage
    tax = float(custom_product_obj.get('tax', 0))  # Tax in percentage
    shipment_cost = float(custom_product_obj.get('shipment_cost', 0))


    if discount > 0:
        custom_product_obj['discount_amount'] = (total_price * discount / 100)  # Subtract discount
        total_price -= custom_product_obj['discount_amount']
    if tax > 0:
        custom_product_obj['tax_amount'] = (total_price * tax / 100)  # Add tax
        total_price += custom_product_obj['tax_amount']

    custom_product_obj['total_price'] = round(total_price + shipment_cost, 2)  # Round to 2 decimal places

    # Save the manual order
    manual_order = DatabaseModel.save_documents(custom_order, custom_product_obj)
    data['message'] = "Manual order created successfully."
    data['order_id'] = str(manual_order.id)
    return data

@csrf_exempt
def listManualOrders(request):
    data = dict()
    json_request = JSONParser().parse(request)
    limit = int(json_request.get('limit', 100))  # Default limit = 100 if not provided
    skip = int(json_request.get('skip', 0))  # Default skip = 0 if not provided
    sort_by = json_request.get('sort_by')
    sort_by_value = json_request.get('sort_by_value')
    pipeline = [
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "order_id": {"$ifNull": ["$order_id", ""]},
                "customer_name": {"$ifNull": ["$customer_name", ""]},
                "shipping_address": {"$ifNull": ["$shipping_address", ""]},
                "total_quantity": {"$ifNull": ["$total_quantity", 0]},
                "total_price": {"$ifNull": [{"$round": ["$total_price", 0.0]}, 0.0]},
                # "taxes": {"$ifNull": ["$taxes", 0.0]},
                "purchase_order_date": {"$ifNull": ["$purchase_order_date", None]},
                "expected_delivery_date": {"$ifNull": ["$expected_delivery_date", None]},
            }
        },
        {
            "$skip": skip
        },
        {
            "$limit": limit + skip
        }
    ]
    if sort_by != None and sort_by != "":
        sort = {
            "$sort" : {
                sort_by : int(sort_by_value)
            }
        }
        pipeline.append(sort)

    manual_orders = list(custom_order.objects.aggregate(*pipeline))
    count_pipeline = [
        {
            "$count": "total_count"
        }
    ]
    total_count_result = list(custom_order.objects.aggregate(*(count_pipeline)))
    total_count = total_count_result[0]['total_count'] if total_count_result else 0
    data['total_count'] = total_count
    data['manual_orders'] = manual_orders
    return data

@csrf_exempt
def updateManualOrder(request):
    data = dict()
    json_request = JSONParser().parse(request)
    order_id = json_request.get('order_id')
    ordered_products = json_request.get('ordered_products',[])
    custom_product_obj = json_request.get('custom_product_obj')

    # Update basic fields
  
    try:
        purchase_order_date = custom_product_obj.get('purchase_order_date')
        if purchase_order_date:
            custom_product_obj['purchase_order_date'] = datetime.strptime(purchase_order_date, '%Y-%m-%d')
    except:
        pass

    expected_delivery_date = custom_product_obj.get('expected_delivery_date')
    if expected_delivery_date:
        custom_product_obj['expected_delivery_date'] = datetime.strptime(expected_delivery_date, '%Y-%m-%d')
    else:
        custom_product_obj['expected_delivery_date'] = None

    # Update ordered products
    product_detail = []
    for ins in ordered_products:
        product_detail_dict = product_details(
            product_id=ObjectId(ins.get('product_id')),
            title=ins.get('title'),
            sku=ins.get('sku'),
            unit_price=float(ins.get('unit_price', 0)),
            quantity=int(ins.get('quantity', 0)),
            quantity_price=float(ins.get('quantity_price', 0))
        )
        product_detail.append(product_detail_dict)

    custom_product_obj['ordered_products'] = product_detail

    # Recalculate total_price with discount and tax
    total_price = float(custom_product_obj.get('total_price', 0))
    discount = float(custom_product_obj.get('discount', 0))  # Discount in percentage
    tax = float(custom_product_obj.get('tax', 0))  # Tax in percentage
    shipment_cost = float(custom_product_obj.get('shipment_cost', 0))

    if discount > 0:
        custom_product_obj['discount_amount'] = (total_price * discount / 100)  # Subtract discount
        total_price -= custom_product_obj['discount_amount']
    if tax > 0:
        custom_product_obj['tax_amount'] = (total_price * tax / 100)  # Add tax
        total_price += custom_product_obj['tax_amount']

    custom_product_obj['total_price'] = round(total_price + shipment_cost, 2)  # Round to 2 decimal places

    # Save the updated manual order
    DatabaseModel.update_documents(custom_order.objects,{"id" : order_id},custom_product_obj)

    data['message'] = "Manual order updated successfully."
    return data


def fetchManualOrderDetails(request):
    data = dict()
    order_id = request.GET.get('order_id')
    pipeline = [
        {
            "$match": {
                "_id": ObjectId(order_id)
            }
        },
        {
            "$lookup": {
                "from": "product",
                "localField": "ordered_products.product_id",
                "foreignField": "_id",
                "as": "product_details"
            }
        },
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "order_id": {"$ifNull": ["$order_id", ""]},
                "customer_name": {"$ifNull": ["$customer_name", ""]},
                "shipping_address": {"$ifNull": ["$shipping_address", ""]},
                "total_quantity": {"$ifNull": ["$total_quantity", 0]},
                "total_price": {"$ifNull": ["$total_price", 0.0]},
                "shipment_type": {"$ifNull": ["$shipment_type", ""]},
                "channel": {"$ifNull": ["$channel", ""]},
                "order_status": {"$ifNull": ["$order_status", "Pending"]},
                "payment_status": {"$ifNull": ["$payment_status", "Pending"]},
                "payment_mode": {"$ifNull": ["$payment_mode", ""]},
                "invoice": {"$ifNull": ["$invoice", ""]},
                "transaction_id": {"$ifNull": ["$transaction_id", ""]},
                "tax": {"$ifNull": ["$tax", 0.0]},
                "tax_amount": {"$ifNull": [{"$round":["$tax_amount",2]}, 0.0]},
                "discount_amount": {"$ifNull": [{"$round":["$discount_amount",2]}, 0.0]},
                "discount": {"$ifNull": ["$discount", 0.0]},
                "supplier_name": {"$ifNull": ["$supplier_name", ""]},
                "mail": {"$ifNull": ["$mail", ""]},
                "contact_number": {"$ifNull": ["$contact_number", ""]},
                "customer_note": {"$ifNull": ["$customer_note", ""]},
                "tags": {"$ifNull": ["$tags", ""]},
                "package_dimensions": {"$ifNull": ["$package_dimensions", ""]},
                "weight": {"$ifNull": ["$weight", 0.0]},
                "shipment_cost": {"$ifNull": ["$shipment_cost", 0.0]},
                "shipment_speed": {"$ifNull": ["$shipment_speed", ""]},
                "shipment_mode": {"$ifNull": ["$shipment_mode", ""]},
                "carrier": {"$ifNull": ["$carrier", ""]},
                "tracking_number": {"$ifNull": ["$tracking_number", ""]},
                "shipping_label": {"$ifNull": ["$shipping_label", ""]},
                "shipping_label_preview": {"$ifNull": ["$shipping_label_preview", ""]},
                "shipping_label_print": {"$ifNull": ["$shipping_label_print", ""]},
                "channel_name": {"$ifNull": ["$channel_name", ""]},
                "channel_order_id": {"$ifNull": ["$channel_order_id", ""]},
                "fulfillment_type": {"$ifNull": ["$fulfillment_type", ""]},
                "purchase_order_date": {"$ifNull": ["$purchase_order_date", None]},
                "expected_delivery_date": {"$ifNull": ["$expected_delivery_date", None]},
                "created_at": {"$ifNull": ["$created_at", None]},
                "updated_at": {"$ifNull": ["$updated_at", None]},
                "customer_order_id" : {"$ifNull" : ["$customer_order_id",""]},
                "weight_value" : {"$ifNull" : ['$weight_value',""]},
                "currency" : {"$ifNull" : ['$currency',"USD"]},
                "ordered_products": {
                    "$map": {
                        "input": "$ordered_products",
                        "as": "product",
                        "in": {
                            "product_id": {"$toString": "$$product.product_id"},
                            "product_title": {"$ifNull": ["$$product.title", ""]},
                            "sku": {"$ifNull": ["$$product.sku", ""]},
                            "price": {"$ifNull": ["$$product.unit_price", 0.0]},
                            "quantity": {"$ifNull": ["$$product.quantity", 0]},
                            "quantity_price": {"$ifNull": ["$$product.quantity_price", 0.0]},
                        }
                    }
                }
            }
        }
    ]
    manual_order_details = list(custom_order.objects.aggregate(*pipeline))
    if manual_order_details:
        for i in manual_order_details[0]['ordered_products']:
            pipeline = [
            {"$match": {"_id" : ObjectId(i['product_id'])}},
            {
                "$project": {
                    "_id": None,
                    "image_url": {"$ifNull": ["$image_url", ""]}
                }
            }
            ]
            product_image_obj = list(Product.objects.aggregate(*(pipeline)))
            try:
                i['product_image'] = product_image_obj[0].get('image_url')
            except:
                i['product_image'] = ""
        data['order_details'] = manual_order_details[0]
    else:
        data['error'] = "Manual order not found."
    return data


#-------------------------------------DASH BOARD APIS-------------------------------------------------------------------------------------------------
def ordersCountForDashboard(request):
    data = dict()
    marketplace_id = request.GET.get('marketplace_id')
    start_date = request.GET.get('start_date')  # Custom start date
    end_date = request.GET.get('end_date')  # Custom end date
    preset = request.GET.get("preset", "Today")

    match_conditions = {}
    if start_date != None and start_date != "":
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    else:
        start_date, end_date = get_date_range(preset)

    match_conditions["order_date"] = {"$gte": start_date, "$lte": end_date}


    if marketplace_id == "all":
        # Count for Order collection
        pipeline = [
            {"$match": match_conditions},
            {
                "$group": {
                    "_id": None,
                    "count": {"$sum": 1}
                }
            }
        ]
        order_status_count = list(Order.objects.aggregate(*(pipeline)))
        total_order_count = order_status_count[0].get('count', 0) if order_status_count else 0

        # Count for custom_order collection
        custom_pipeline = [
            {"$match": match_conditions},
            {
                "$group": {
                    "_id": None,
                    "count": {"$sum": 1}
                }
            }
        ]
        custom_order_status_count = list(custom_order.objects.aggregate(*(custom_pipeline)))
        custom_order_count = custom_order_status_count[0].get('count', 0) if custom_order_status_count else 0

        # Combine counts
        total_order_count += custom_order_count

        data['total_order_count'] = {
            "value": total_order_count,
            "percentage": f"{100.00}%"
        }

        pipeline = [
            {
                "$project": {
                    "_id": 1,
                    "name": 1,
                }
            }
        ]
        marketplace_list = list(Marketplace.objects.aggregate(*(pipeline)))
        for ins in marketplace_list:
            # Count for Order collection per marketplace
            pipeline = [
                {"$match": {**match_conditions, "marketplace_id": ins['_id']}},
                {
                    "$group": {
                        "_id": None,
                        "count": {"$sum": 1},
                        "order_value": {"$sum": "$order_total"}

                    }
                }
            ]
            order_status_count = list(Order.objects.aggregate(*(pipeline)))
            order_count = order_status_count[0].get('count', 0) if order_status_count else 0
            order_value = order_status_count[0].get('order_value', 0) if order_status_count else 0
            percentage = round((order_count / total_order_count) * 100, 2) if total_order_count else 0
            data[ins['name']] = {
                "count": order_count,
                "percentage": f"{percentage}%",
                "order_value" : round(order_value,2)
            }
        # custom_pipeline = [
        #         {"$match": {**match_conditions}},
        #         {
        #             "$group": {
        #                 "_id": None,
        #                 "count": {"$sum": 1},
        #                 "order_value": {"$sum": "$total_price"}
        #             }
        #         }
        #     ]
        # custom_order_status_count = list(custom_order.objects.aggregate(*(custom_pipeline)))
        # custom_order_count = custom_order_status_count[0].get('count', 0) if custom_order_status_count else 0
        # order_value = custom_order_status_count[0].get('order_value', 0) if custom_order_status_count else 0
        # percentage = round((custom_order_count / total_order_count) * 100, 2) if total_order_count else 0
        # data['custom'] = {
        #     "value": custom_order_count,
        #     "percentage": f"{percentage}%",
        #     "order_value" : round(order_value,2)
        # }
    elif marketplace_id != "all" and marketplace_id != "custom":
        # Count for Order collection
        pipeline = [
            {"$match": {**match_conditions, "marketplace_id": ObjectId(marketplace_id)}},
            {
                "$group": {
                    "_id": None,
                    "count": {"$sum": 1},
                    "order_value": {"$sum": "$order_total"}

                }
            }
        ]
        order_status_count = list(Order.objects.aggregate(*(pipeline)))
        order_count = order_status_count[0].get('count', 0) if order_status_count else 0
        order_value = order_status_count[0].get('order_value', 0) if order_status_count else 0
        marketplace_name = DatabaseModel.get_document(Marketplace.objects, {"id": marketplace_id}, ['name']).name
        data[marketplace_name] = {
            "value": order_count,
            "percentage": f"{100.00}%",
            "order_value" : round(order_value,2)
        }
        data['total_order_count'] = {
            "value": order_count,
            "percentage": f"{100.00}%"
        }
    elif marketplace_id == "custom":
        custom_pipeline = [
                {"$match": {**match_conditions}},
                {
                    "$group": {
                        "_id": None,
                        "count": {"$sum": 1},
                        "order_value": {"$sum": "$total_price"}
                    }
                }
            ]
        custom_order_status_count = list(custom_order.objects.aggregate(*(custom_pipeline)))
        custom_order_count = custom_order_status_count[0].get('count', 0) if custom_order_status_count else 0
        order_value = custom_order_status_count[0].get('order_value', 0) if custom_order_status_count else 0
        data['custom'] = {
            "value": custom_order_count,
            "percentage": f"{100.00}%",
            "order_value" : round(order_value,2)
        }
        data['total_order_count'] = {
            "value": custom_order_count,
            "percentage": f"{100.00}%"
        }

    return data


def totalSalesAmount(request):
    data = dict()
    marketplace_id = request.GET.get('marketplace_id')
    pipeline = []
    if marketplace_id != None and marketplace_id != "":
        match = {
            "$match" : {
                "marketplace_id" : ObjectId(marketplace_id)
            }
        }
        pipeline.append(match)
    pipeline.extend([
        {
            "$group": {
                "_id": None,
                "total_sales": {"$sum": "$order_total"}
            }
        }
    ])
    total_sales = list(Order.objects.aggregate(*(pipeline)))
    data['total_sales'] = total_sales[0]['total_sales'] if total_sales else 0
    return data


@csrf_exempt
def salesAnalytics(request):
    data = dict()
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')  # Optional marketplace filter
    date_range = json_request.get('date_range', 'all')  # 'week', 'month', 'year', or 'all'
    start_date = json_request.get('start_date')  # Optional custom start date
    end_date = json_request.get('end_date')  # Optional custom end date

    preset = json_request.get("preset", "Today")        
    if start_date != None and start_date != "":
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
    else:
        start_date, end_date = get_date_range(preset)

    # Match conditions
    match_conditions = {}
    if start_date:
        match_conditions["order_date"] = {"$gte": start_date, "$lte": end_date}
    if marketplace_id != None and marketplace_id != "all" and marketplace_id != "custom":
        match_conditions["marketplace_id"] = ObjectId(marketplace_id)

    # Pipeline for total sales amount (Order collection)
    total_sales_pipeline = [
        {"$match": match_conditions},
        {"$group": {"_id": None, "total_sales": {"$sum": "$order_total"}}}
    ]
    total_sales_result = list(Order.objects.aggregate(*total_sales_pipeline))
    total_sales = total_sales_result[0]['total_sales'] if total_sales_result else 0

    # Pipeline for total sales amount (custom_order collection)
    custom_match_conditions = {}
    if start_date:
        custom_match_conditions["purchase_order_date"] = {"$gte": start_date, "$lte": end_date}

    custom_total_sales_pipeline = [
        {"$match": custom_match_conditions},
        {"$group": {"_id": None, "total_sales": {"$sum": "$total_price"}}}
    ]
    custom_total_sales_result = list(custom_order.objects.aggregate(*custom_total_sales_pipeline))
    custom_total_sales = custom_total_sales_result[0]['total_sales'] if custom_total_sales_result else 0

    if marketplace_id == "custom":
        data['total_sales'] = custom_total_sales
    else:
        # Combine total sales from both collections
        data['total_sales'] = total_sales + custom_total_sales

    # Pipeline for order count and value by order days (Order collection)
    order_days_pipeline = [
        {"$match": match_conditions},
        {
            "$group": {
                "_id": {
                    "year": {"$year": "$order_date"},
                    "month": {"$month": "$order_date"},
                    "day": {"$dayOfMonth": "$order_date"}
                },
                "order_count": {"$sum": 1},
                "order_value": {"$sum": "$order_total"}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    order_days_data = list(Order.objects.aggregate(*order_days_pipeline))

    # Pipeline for order count and value by order days (custom_order collection)
    custom_order_days_pipeline = [
        {"$match": custom_match_conditions},
        {
            "$group": {
                "_id": {
                    "year": {"$year": "$purchase_order_date"},
                    "month": {"$month": "$purchase_order_date"},
                    "day": {"$dayOfMonth": "$purchase_order_date"}
                },
                "order_count": {"$sum": 1},
                "order_value": {"$sum": "$total_price"}
            }
        },
        {"$sort": {"_id": 1}}
    ]
    custom_order_days_data = list(custom_order.objects.aggregate(*custom_order_days_pipeline))

    # Combine and format order days data
    combined_order_days = {}
    if marketplace_id != "custom":
        for day in order_days_data:
            date_key = f"{day['_id']['year']}-{day['_id']['month']:02d}-{day['_id']['day']:02d}"
            combined_order_days[date_key] = {
                "order_count": day["order_count"],
                "order_value": round(day["order_value"], 2)
            }

    for day in custom_order_days_data:
        date_key = f"{day['_id']['year']}-{day['_id']['month']:02d}-{day['_id']['day']:02d}"
        if date_key in combined_order_days:
            combined_order_days[date_key]["order_count"] += day["order_count"]
            combined_order_days[date_key]["order_value"] += round(day["order_value"], 2)
        else:
            combined_order_days[date_key] = {
                "order_count": day["order_count"],
                "order_value": round(day["order_value"], 2)
            }

    # Sort combined data by date
    formatted_order_days = [
        {"date": date, "order_count": data["order_count"], "order_value": data["order_value"]}
        for date, data in sorted(combined_order_days.items())
    ]

    data['order_days'] = formatted_order_days
    return data

@csrf_exempt
def mostSellingProducts(request):
    data = dict()
    pipeline = list()
    marketPlaceId = request.GET.get('marketPlaceId')
    start_date = request.GET.get('start_date')  # Optional custom start date
    end_date = request.GET.get('end_date')  # Optional custom end date

    # Match conditions for date range
    match_conditions = {}
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        match_conditions["order_data.order_date"] = {"$gte": start_date, "$lte": end_date}

    # Pipeline to filter orders based on date range and marketplace
    pipeline.extend([
        {
            "$lookup": {
                "from": "order",
                "localField": "_id",
                "foreignField": "order_items",
                "as": "order_data"
            }
        },
        {
            "$unwind": "$order_data"
        },
        {
            "$match": match_conditions
        },
        {
            "$lookup": {
                "from": "product",
                "localField": "ProductDetails.product_id",
                "foreignField": "_id",
                "as": "product_ins"
            }
        },
        {
            "$unwind": "$product_ins"
        },
        {
            "$lookup": {
                "from": "marketplace",
                "localField": "product_ins.marketplace_id",
                "foreignField": "_id",
                "as": "marketplace_ins"
            }
        },
        {
            "$unwind": "$marketplace_ins"
        },
    ])

    if marketPlaceId and marketPlaceId != "all" and marketPlaceId != "custom":
        match = {
            "$match": {
                "product_ins.marketplace_id": ObjectId(marketPlaceId)
            }
        }
        pipeline.append(match)

    pipeline2 = [
        {
            "$group": {
                "_id": {
                    "product_id": "$ProductDetails.product_id",
                    "product_title": "$product_ins.product_title",
                    "sku": "$product_ins.sku",
                    "image_url": "$product_ins.image_url",
                    "price": "$product_ins.price",
                    "channel_name" : "$marketplace_ins.name"
                },
                "total_quantity_sold": {"$sum": "$ProductDetails.QuantityOrdered"},
                "total_revenue": {"$sum": {"$multiply": ["$ProductDetails.QuantityOrdered", "$product_ins.price"]}}
            }
        },
        {
            "$project": {
                "_id": 0,
                "product_id": {"$toString": "$_id.product_id"},
                "product_title": "$_id.product_title",
                "sku": "$_id.sku",
                "product_image": "$_id.image_url",
                "price": "$_id.price",
                "channel_name" : "$_id.channel_name",
                "sales_count": "$total_quantity_sold",
                "revenue": {"$round": ["$total_revenue", 2]}
            }
        },
        {
            "$sort": {"sales_count": -1}
        },
        {
            "$limit": 7
        },
    ]
    pipeline.extend(pipeline2)
    top_products = list(OrderItems.objects.aggregate(*pipeline))

    # Add custom_order count and value
    custom_match_conditions = {}
    if start_date and end_date:
        custom_match_conditions["purchase_order_date"] = {"$gte": start_date, "$lte": end_date}

    custom_pipeline = [
        {
            "$match": custom_match_conditions
        },
        {
            "$unwind": "$ordered_products"
        },
        {
            "$group": {
                "_id": {
                    "product_id": "$ordered_products.product_id",
                    "product_title": "$ordered_products.title",
                    "sku": "$ordered_products.sku",
                    "price": "$ordered_products.unit_price"
                },
                "total_quantity_sold": {"$sum": "$ordered_products.quantity"},
                "total_revenue": {"$sum": "$ordered_products.quantity_price"}
            }
        },
        {
            "$project": {
                "_id": 0,
                "product_id": {"$toString": "$_id.product_id"},
                "product_title": "$_id.product_title",
                "sku": "$_id.sku",
                "price": "$_id.price",
                "sales_count": "$total_quantity_sold",
                "revenue": {"$round": ["$total_revenue", 2]}
            }
        },
        {
            "$sort": {"sales_count": -1}
        },
        {
            "$limit": 7
        }
    ]
    custom_top_products = list(custom_order.objects.aggregate(*custom_pipeline))
    for i in custom_top_products:
        pipeline = [
        {"$match": {"_id" : ObjectId(i['product_id'])}},
        {
            "$project": {
                "_id": None,
                "image_url": {"$ifNull": ["$image_url", ""]}
            }
        }
        ]
        product_image_obj = list(Product.objects.aggregate(*(pipeline)))
        try:
            i['product_image'] = product_image_obj[0].get('image_url')
        except:
            print(i['product_id'])
            i['product_image'] = ""
        i['channel_name'] = "custom"
    if marketPlaceId == "custom":
        data['top_products'] = custom_top_products
        return data
    if marketPlaceId == "all":
        top_products.extend(custom_top_products)
        top_products = sorted(top_products, key=lambda x: x['sales_count'], reverse=True)[:7]

    data['top_products'] = top_products
    return data

def change_sign(value):
    """
    Change the sign of a given value only if it's negative.
    
    Args:
    value (number): The input value to potentially change the sign of.
    
    Returns:
    number: The input value with its sign changed if it was negative, otherwise unchanged.
    """
    if value < 0:
        return -value
    else:
        return value
    
    
@csrf_exempt
def getSalesTrendPercentage(request):
    data = dict()
    json_request = JSONParser().parse(request)
    range_type = json_request.get('range_type', 'month')  # 'day', 'week', 'month', 'year'
    marketplace_id = json_request.get('marketplace_id')  # Marketplace filter (e.g., 'amazon', 'walmart')

    # Determine date ranges based on range_type
    now = datetime.now()
    if range_type == 'day':
        current_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        previous_start = current_start - timedelta(days=1)
        previous_end = current_start
    elif range_type == 'week':
        current_start = now - timedelta(days=now.weekday())
        current_start = current_start.replace(hour=0, minute=0, second=0, microsecond=0)
        previous_start = current_start - timedelta(weeks=1)
        previous_end = current_start
    elif range_type == 'month':
        current_start = datetime(now.year, now.month, 1)
        previous_month_end = current_start - timedelta(days=1)
        previous_start = datetime(previous_month_end.year, previous_month_end.month, 1)
        previous_end = current_start
    elif range_type == 'year':
        current_start = datetime(now.year, 1, 1)
        previous_start = datetime(now.year - 1, 1, 1)
        previous_end = current_start
    else:
        data['error'] = "Invalid range_type provided."
        return data

    # Match pipeline for current and previous ranges
    match_pipeline = [
        {
            "$facet": {
                "current_range": [
                    {
                        "$match": {
                            "order_date": {
                                "$gte": current_start,
                                "$lt": now
                            },
                            **({"marketplace_id": ObjectId(marketplace_id)} if marketplace_id and marketplace_id != "custom" and marketplace_id != "all" else {})
                        }
                    },
                    {
                        "$group": {
                            "_id": "$marketplace_id",
                            "sales_value": {"$sum": "$order_total"}
                        }
                    }
                ],
                "previous_range": [
                    {
                        "$match": {
                            "order_date": {
                                "$gte": previous_start,
                                "$lt": previous_end
                            },
                            **({"marketplace_id": ObjectId(marketplace_id)} if marketplace_id and marketplace_id != "custom" and marketplace_id != "all" else {})
                        }
                    },
                    {
                        "$group": {
                            "_id": "$marketplace_id",
                            "sales_value": {"$sum": "$order_total"}
                        }
                    }
                ]
            }
        }
    ]

    custom_match_pipeline = [
        {
            "$facet": {
                "current_range": [
                    {
                        "$match": {
                            "purchase_order_date": {
                                "$gte": current_start,
                                "$lt": now
                            }
                        }
                    },
                    {
                        "$group": {
                            "_id": None,
                            "sales_value": {"$sum": "$total_price"}
                        }
                    }
                ],
                "previous_range": [
                    {
                        "$match": {
                            "purchase_order_date": {
                                "$gte": previous_start,
                                "$lt": previous_end
                            }
                        }
                    },
                    {
                        "$group": {
                            "_id": None,
                            "sales_value": {"$sum": "$total_price"}
                        }
                    }
                ]
            }
        }
    ]

    if marketplace_id == "custom":
        trend_data = list(custom_order.objects.aggregate(*custom_match_pipeline))
    else:
        trend_data = list(Order.objects.aggregate(*match_pipeline))
        custom_trend_data = list(custom_order.objects.aggregate(*custom_match_pipeline))

        # Combine custom_order data with Order data
        if custom_trend_data:
            for key in ["current_range", "previous_range"]:
                for item in custom_trend_data[0][key]:
                    trend_data[0][key].append({"_id": "custom", "sales_value": item["sales_value"]})

    if trend_data:
        current_range_data = {item["_id"]: item["sales_value"] for item in trend_data[0]["current_range"]}
        previous_range_data = {item["_id"]: item["sales_value"] for item in trend_data[0]["previous_range"]}

        if marketplace_id == "all":  # Combine all marketplaces
            current_total = sum(current_range_data.values())
            previous_total = sum(previous_range_data.values())
            percentage_change = ((current_total - previous_total) / previous_total * 100) if previous_total != 0 else (100 if current_total > 0 else 0)
            current_percentage = (current_total / previous_total * 100) if previous_total != 0 else (100 if current_total > 0 else 0)
            data['trend_percentage'] = [{
                "id": "All Channels",
                "current_range_sales": current_total,
                "previous_range_sales": previous_total,
                "trend_percentage": round(percentage_change, 2),
                "current_percentage" : round(current_percentage,2)
            }]
        elif marketplace_id == "custom":  # Only custom_order data
            current_total = current_range_data.get(None, 0)
            previous_total = previous_range_data.get(None, 0)
            percentage_change = ((current_total - previous_total) / previous_total * 100) if previous_total != 0 else (100 if current_total > 0 else 0)
            current_percentage = (current_total / previous_total * 100) if previous_total != 0 else (100 if current_total > 0 else 0)
            data['trend_percentage'] = [{
                "id": "Custom Orders",
                "current_range_sales": current_total,
                "previous_range_sales": previous_total,
                "trend_percentage": (round(percentage_change, 2)),
                "current_percentage" : round(current_percentage,2)
            }]
        else:  # Specific marketplace
            trend_percentage = []
            marketplace_name = DatabaseModel.get_document(Marketplace.objects, {"id": marketplace_id}, ['name']).name
            for key in set(current_range_data.keys()).union(previous_range_data.keys()):
                current_value = current_range_data.get(key, 0)
                previous_value = previous_range_data.get(key, 0)
                percentage_change = ((current_value - previous_value) / previous_value * 100) if previous_value != 0 else (100 if current_value > 0 else 0)
                current_percentage = (current_value / previous_value * 100) if previous_value != 0 else (100 if current_value > 0 else 0)
                trend_percentage.append({
                    "id": str(marketplace_name),
                    "current_range_sales": current_value,
                    "previous_range_sales": previous_value,
                    "trend_percentage": (round(percentage_change, 2)),
                    "current_percentage" : round(current_percentage,2)
                })
            data['trend_percentage'] = trend_percentage
    else:
        data['trend_percentage'] = []

    return data

@csrf_exempt
def fetchSalesSummary(request):
    data = {}
    total_sales_pipeline = []
    pipeline = []
    match = {}

    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    start_date = json_request.get('start_date')  # Optional custom start date
    end_date = json_request.get('end_date')  # Optional custom end date

    # Add date range filter if provided
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        match["order_date"] = {"$gte": start_date, "$lte": end_date}

    if marketplace_id != None and marketplace_id != "all" and marketplace_id != "" and marketplace_id != "custom":
        match["marketplace_id"] = ObjectId(marketplace_id)
    if marketplace_id == "all" or marketplace_id != "custom":
        if match:
            match_stage = {"$match": match}
            total_sales_pipeline.append(match_stage)
            pipeline.append(match_stage)

        # Pipeline to calculate total units sold, total sold product count, and total sales
        total_sales_pipeline.extend([
            {
                "$group": {
                    "_id": None,  # Grouping by None to get a single summary document
                    "total_sales": {"$sum": "$order_total"}
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "total_sales": {"$round": ["$total_sales", 2]}
                }
            }
        ])

        summary = list(Order.objects.aggregate(*total_sales_pipeline))
        if summary:
            data['total_sales'] = summary[0].get('total_sales', 0)
        else:
            data['total_sales'] = 0

        pipeline.extend([
            {
                "$unwind": "$order_items"  # Unwind the order_items list to process each item individually
            },
            {
                "$group": {
                    "_id": None,  # Grouping by None to get a single summary document
                    "ids": {"$addToSet": "$order_items"}  # Collect unique order_items
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "ids": 1
                }
            },
        ])
        summary1 = list(Order.objects.aggregate(*pipeline))
        if summary1:
            s_pipeline = [
                {"$match": {
                    "_id": {"$in": summary1[0]['ids']}
                }},
                {
                    "$group": {
                        "_id": None,
                        "total_units_sold": {"$sum": "$ProductDetails.QuantityOrdered"},
                        "unique_product_ids": {"$addToSet": "$ProductDetails.product_id"},
                    }
                },
                {
                    "$project": {
                        "_id": 0,
                        "total_units_sold": 1,
                        "total_sold_product_count": {"$size": "$unique_product_ids"},
                    }
                }
            ]
            summary2 = list(OrderItems.objects.aggregate(*s_pipeline))
            if summary2:
                data['total_units_sold'] = summary2[0].get('total_units_sold', 0)
                data['total_sold_product_count'] = summary2[0].get('total_sold_product_count', 0)
        else:
            data['total_units_sold'] = 0
            data['total_sold_product_count'] = 0

    # Add custom_order data
    custom_match = {}
    if start_date and end_date:
        custom_match["purchase_order_date"] = {"$gte": start_date, "$lte": end_date}

    custom_pipeline = [
        {"$match": custom_match},
        {"$unwind": "$ordered_products"},
        {
            "$group": {
                "_id": None,
                "total_custom_sales": {"$sum": "$ordered_products.quantity_price"},
                "total_custom_units_sold": {"$sum": "$ordered_products.quantity"},
                "unique_custom_product_ids": {"$addToSet": "$ordered_products.product_id"}
            }
        },
        {
            "$project": {
                "_id": 0,
                "total_custom_sales": {"$round": ["$total_custom_sales", 2]},
                "total_custom_units_sold": 1,
                "total_custom_sold_product_count": {"$size": "$unique_custom_product_ids"}
            }
        }
    ]

    custom_summary = list(custom_order.objects.aggregate(*custom_pipeline))
    if custom_summary:
        total_custom_sales = custom_summary[0].get('total_custom_sales', 0)
        total_custom_units_sold = custom_summary[0].get('total_custom_units_sold', 0)
        total_custom_sold_product_count = custom_summary[0].get('total_custom_sold_product_count', 0)
    else:
        total_custom_sales = 0
        total_custom_units_sold = 0
        total_custom_sold_product_count = 0

    if marketplace_id == "custom":
        data['total_sales'] = total_custom_sales
        data['total_units_sold'] = total_custom_units_sold
        data['total_sold_product_count'] = total_custom_sold_product_count

    if marketplace_id == "": 
        # Combine totals
        data['total_sales'] += total_custom_sales
        data['total_units_sold'] += total_custom_units_sold
        data['total_sold_product_count'] += total_custom_sold_product_count

    return data

@csrf_exempt
def fetchTopSellingCategories(request):
    data = dict()
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')  # Optional marketplace filter
    limit = int(json_request.get('limit', 15))  # Default limit = 15 if not provided
    start_date = json_request.get('start_date')  # Optional custom start date
    end_date = json_request.get('end_date')  # Optional custom end date

    # Match conditions
    match_conditions = {}
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        match_conditions["order_data.order_date"] = {"$gte": start_date, "$lte": end_date}
    if marketplace_id != None and marketplace_id != "all" and marketplace_id != "custom":
        match_conditions["order_data.marketplace_id"] = ObjectId(marketplace_id)

    if marketplace_id != "custom":
        # Pipeline to fetch top selling categories based on order value
        pipeline = [
            {
                "$lookup": {
                    "from": "order",
                    "localField": "_id",
                    "foreignField": "order_items",
                    "as": "order_data"
                }
            },
            {
                "$unwind": "$order_data"
            },
            {
                "$match": match_conditions
            },
            {
                "$lookup": {
                    "from": "product",
                    "localField": "ProductDetails.product_id",
                    "foreignField": "_id",
                    "as": "product_ins"
                }
            },
            {
                "$unwind": "$product_ins"
            },
            {
                "$lookup": {
                    "from": "category",
                    "localField": "product_ins.category",
                    "foreignField": "name",
                    "as": "category_ins"
                }
            },
            {
                "$unwind": "$category_ins"
            },
            {
                "$group": {
                    "_id": {
                        "category_id": "$category_ins._id",
                        "category_name": "$category_ins.name"
                    },
                    "total_order_value": {"$sum": {"$multiply": ["$ProductDetails.QuantityOrdered", "$product_ins.price"]}}
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "category_id": {"$toString": "$_id.category_id"},
                    "category_name": "$_id.category_name",
                    "total_order_value": {"$round": ["$total_order_value", 2]}
                }
            },
            {
                "$sort": {"total_order_value": -1}
            },
            {
                "$limit": limit
            }
        ]

        top_categories = list(OrderItems.objects.aggregate(*pipeline))

    # Add custom_order product value for categories
    custom_match_conditions = {}
    if start_date and end_date:
        custom_match_conditions["purchase_order_date"] = {"$gte": start_date, "$lte": end_date}

    custom_pipeline = [
        {"$match": custom_match_conditions},
        {"$unwind": "$ordered_products"},
        {
            "$lookup": {
                "from": "product",
                "localField": "ordered_products.product_id",
                "foreignField": "_id",
                "as": "product_ins"
            }
        },
        {
            "$unwind": "$product_ins"
        },
        {
            "$lookup": {
                "from": "category",
                "localField": "product_ins.category",
                "foreignField": "name",
                "as": "category_ins"
            }
        },
        {
            "$unwind": "$category_ins"
        },
        {
            "$group": {
                "_id": {
                    "category_id": "$category_ins._id",
                    "category_name": "$category_ins.name"
                },
                "total_order_value": {"$sum": "$ordered_products.quantity_price"}
            }
        },
        {
            "$project": {
                "_id": 0,
                "category_id": {"$toString": "$_id.category_id"},
                "category_name": "$_id.category_name",
                "total_order_value": {"$round": ["$total_order_value", 2]}
            }
        },
        {
            "$sort": {"total_order_value": -1}
        },
        {
            "$limit": limit
        }
    ]

    custom_top_categories = list(custom_order.objects.aggregate(*custom_pipeline))
    if marketplace_id == "custom":
        data['top_categories'] = custom_top_categories
    else:
        # Combine and sort categories by total_order_value
        combined_categories = {}
        for category in top_categories + custom_top_categories:
            category_id = category["category_id"]
            if category_id in combined_categories:
                combined_categories[category_id]["total_order_value"] += category["total_order_value"]
            else:
                combined_categories[category_id] = category

        sorted_categories = sorted(combined_categories.values(), key=lambda x: x["total_order_value"], reverse=True)[:limit]

        data['top_categories'] = sorted_categories
    return data

#-----------------------------------USER CREATION--------------------------------

@csrf_exempt
def createUser(request):
    data = dict()
    json_request = JSONParser().parse(request)
    email = json_request.get("email")
    old_user_obj = DatabaseModel.get_document(user.objects,{"email" : email},['id'])
    if old_user_obj == None:
        user_data = {
            "first_name": json_request.get("first_name"),
            "last_name" : json_request.get('last_name'),
            "email": email,
            "password": json_request.get("password"),  # Ensure to hash the password in production
            "role_id": ObjectId(json_request.get("role_id")),
            "profile_image" : json_request.get("profile_image", "")
            
        }
        new_user = DatabaseModel.save_documents(user, user_data)
        data["message"] = "User created successfully."
        data["user_id"] = str(new_user.id)
    else:
        data["message"] = "User Already Present."
        data["user_id"] = str(old_user_obj.id)
    return data


@csrf_exempt
def updateUser(request):
    data = dict()
    json_request = JSONParser().parse(request)
    target_user_id = json_request.get("target_user_id")
    update_obj = json_request.get('update_obj')
    old_user_obj = DatabaseModel.get_document(user.objects,{"id" : target_user_id})
    data["message"] = "User Not Updated."
    try:
        update_obj['role_id'] = ObjectId(update_obj['role_id'])
    except:
        pass
    if old_user_obj:
        DatabaseModel.update_documents(user.objects,{"id" : target_user_id},update_obj)
        data["message"] = "User updated successfully."
    return data


@csrf_exempt
def listUsers(request):
    data = dict()
    limit = int(request.GET.get("limit", 100))  # Default limit = 100
    skip = int(request.GET.get("skip", 0))  # Default skip = 0

    pipeline = [
         {
            "$lookup": {
                "from": "role",
                "localField": "role_id",
                "foreignField": "_id",
                "as": "role_ins"
            }
        },
        {
            "$unwind": {
                "path": "$role_ins",
                "preserveNullAndEmptyArrays": True  # Ensure the product is not removed if no orders exist
            }
        },
        {
        "$project": {
            "_id": 0,
            "id": {"$toString": "$_id"},
            "first_name": 1,
            "last_name": 1,
            "email" : 1,
            "role_name": "$role_ins.name",
            "creation_date" :1,
            "role_id" : {"$toString" : "$role_id"}
        }
    },
        {"$skip": skip},
        {"$limit": limit},
        {
            "$sort" : {
                "id" : -1
            }
        }
    ]
    users = list(user.objects.aggregate(*pipeline))
    data["users"] = users
    return data


def fetchUserDetails(request):
    data = dict()
    target_user_id = request.GET.get("target_user_id")
    data['user_obj'] = {}
    pipeline = [
    {
        "$match" : {
            "_id" : ObjectId(target_user_id)
        }
    },
    {
        "$lookup": {
            "from": "role",
            "localField": "role_id",
            "foreignField": "_id",
            "as": "role_ins"
        }
    },
    {
        "$unwind": "$role_ins"
    },
    {
        "$project": {
            "_id": 0,
            "id": {"$toString": "$_id"},
            "first_name": 1,
            "last_name": 1,
            "email" : 1,
            "mobile_number" : 1,
            "profile_image" : 1,
            "role_name": "$role_ins.name",
        }
    },
    ]
    user_obj = list(user.objects.aggregate(*pipeline))
    if user_obj != []:
        data['user_obj'] = user_obj[0]
   
    return data


def fetchRoles(request):
    pipeline = [
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "name": 1,
            }
        }
    ]
       
    role_list = list(role.objects.aggregate(*pipeline))
    return role_list

#-------------------------------------------INVENTRY MANAGEMENT--------------------
@csrf_exempt
def fetchInventryList(request):
    data = dict()
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    skip = int(json_request.get('skip'))
    limit = int(json_request.get('limit'))
    search_query = json_request.get('search_query')   
    # marketplace = json_request.get('marketplace')
    # category_name = json_request.get('category_name')
    # brand_id_list = json_request.get('brand_id_list')
    sort_by = json_request.get('sort_by')
    sort_by_value = json_request.get('sort_by_value')
    pipeline = []
    count_pipeline = []
    match = {}
    if marketplace_id != None and marketplace_id != "" and marketplace_id != "all":
        match['marketplace_id'] = ObjectId(marketplace_id)
    # if category_name != None and category_name != "" and category_name != []:
    #     match['category'] = {"$in":category_name}
    # if brand_id_list != None and brand_id_list != "" and brand_id_list != []:
    #     match['brand_id'] = {"$in":[ObjectId(brand_id) for brand_id in brand_id_list]}
    if search_query != None and search_query != "":
        search_query = search_query.strip() 
        match["$or"] = [
            {"product_title": {"$regex": search_query, "$options": "i"}},
            {"sku": {"$regex": search_query, "$options": "i"}},
        ]
    if match != {}:
        match_pipeline = {
            "$match" : match}
        print(match_pipeline)
        pipeline.append(match_pipeline)
        count_pipeline.append(match_pipeline)
    pipeline.extend([
        {
            "$lookup" : {
                "from" : "marketplace",
                "localField" : "marketplace_id",
                "foreignField" : "_id",
                "as" : "marketplace"
            }
        },
        {
            "$unwind" : "$marketplace"
        },
        {
            "$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "product_title" : 1,
                "sku" : 1,
                "price" : 1,
                "quantity" : 1,
                "image_url" : {"$ifNull" : ["$image_url",""]},  # If image_url is null, replace with empty string
                "marketplace_name" : "$marketplace.name",
            }
        },
        {
            "$skip" : skip
        },
        {
            "$limit" : limit+skip
        }
    ])
    if sort_by != None and sort_by != "":
        sort = {
            "$sort" : {
                sort_by : int(sort_by_value)
            }
        }
        pipeline.append(sort)
    inventry_list = list(Product.objects.aggregate(*(pipeline)))
    # Get total product count
    count_pipeline.extend([
        {
            "$count": "total_count"
        }
    ])
    total_count_result = list(Product.objects.aggregate(*(count_pipeline)))
    total_count = total_count_result[0]['total_count'] if total_count_result else 0

    data['total_count'] = total_count
    data['inventry_list'] = inventry_list
    return data



def exportOrderReport(request):
    # Fetch orders from the database
    orders = list(Order.objects.all())

    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Order Report"

    # Determine the maximum number of products in any order
    max_products = max(len(order.order_items) for order in orders) if orders else 1

    # Define base headers
    order_headers = [
        "Purchase Order Id", "Customer Order Id", "Order Date", "Marketplace Name", "Earliest Ship Date",
        "Fulfilment Channel", "Order Status", "Ship Service Level", "Customer Email Id",
        "Has Regulated Items", "Is Replacement Order", "Shipping Information"
    ]

    # Dynamically generate product headers
    product_headers = []
    for i in range(1, max_products + 1):
        product_headers.extend([
            f"Product {i} Name", f"Product {i} SKU", f"Product {i} Quantity Ordered",
            f"Product {i} Quantity Shipped", f"Product {i} Unit Price"
        ])

    # Fixed fields at the end
    fixed_headers = ["Discount", "Tax", "Total Order", "Currency"]

    # Combine all headers
    headers = order_headers + product_headers + fixed_headers

    # Apply header styles
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    bold_font = Font(bold=True)

    # Write headers
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.fill = blue_fill
        cell.font = bold_font

    # Write order data
    row = 2
    for order in orders:
        base_col = 1

        # Process Shipping Information
        if order.shipping_information:
            shipping_info = order.shipping_information
            postal_address = shipping_info.get('postalAddress', {})
            shipping_details = f"Name: {postal_address.get('name', '')}, " \
                               f"Address1: {postal_address.get('address1', '')}, " \
                               f"Address2: {postal_address.get('address2', '')}, " \
                               f"City: {postal_address.get('city', '')}, " \
                               f"State: {postal_address.get('state', '')}, " \
                               f"PostalCode: {postal_address.get('postalCode', '')}, " \
                               f"Country: {postal_address.get('country', '')}, " \
                               f"Phone: {shipping_info.get('phone', '')}, " \
                               f"MethodCode: {shipping_info.get('methodCode', '')}, " \
                               f"EstimatedShipDate: {shipping_info.get('estimatedShipDate', '')}, " \
                               f"EstimatedDeliveryDate: {shipping_info.get('estimatedDeliveryDate', '')}"
        else:
            shipping_details = None

        # Order details
        order_data = [
            order.purchase_order_id, order.customer_order_id,
            order.order_date.strftime('%Y-%m-%d') if order.order_date else None,
            order.marketplace_id.name if order.marketplace_id else None,
            order.earliest_ship_date, order.fulfillment_channel,
            order.order_status, order.ship_service_level,
            order.customer_email_id, order.has_regulated_items,
            order.is_replacement_order, shipping_details
        ]

        # Write order details
        for col_index, value in enumerate(order_data, start=base_col):
            sheet.cell(row=row, column=col_index, value=value)

        # Move to product details section
        col = base_col + len(order_data)

        # Iterate over ordered items and populate product columns dynamically
        for product_index, product in enumerate(order.order_items[:max_products]):
            product_details = [
                product.ProductDetails.Title, product.ProductDetails.SKU,
                product.ProductDetails.QuantityOrdered, product.ProductDetails.QuantityShipped,
                product.Pricing['ItemPrice']['Amount'] if product.Pricing else None
            ]

            # Write product details
            for col_index, value in enumerate(product_details, start=col + product_index * 5):
                sheet.cell(row=row, column=col_index, value=value)

        # Move to fixed fields section
        col = base_col + len(order_data) + (max_products * 5)

        # Fixed Fields: Discount, Tax, Total Order, Currency
        fixed_details = [
            order.discount if hasattr(order, 'discount') else None,
            order.tax if hasattr(order, 'tax') else None,
            order.order_total, order.currency
        ]

        # Write fixed details
        for col_index, value in enumerate(fixed_details, start=col):
            sheet.cell(row=row, column=col_index, value=value)

        # Move to the next row for the next order
        row += 1

    # Adjust column widths
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    # Save the workbook to a BytesIO stream
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Create a response with the Excel file
    response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Order_Report.xlsx"'
    return response



@csrf_exempt
def fetchSalesSummary(request):
    data = {}
    total_sales_pipeline = []
    pipeline = []
    match = {}

    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    start_date = json_request.get('start_date')  # Optional custom start date
    end_date = json_request.get('end_date')  # Optional custom end date

    # Add date range filter if provided
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        match["order_date"] = {"$gte": start_date, "$lte": end_date}

    if marketplace_id != None and marketplace_id != "all" and marketplace_id != "" and marketplace_id != "custom":
        match["marketplace_id"] = ObjectId(marketplace_id)
    if marketplace_id == "all" or marketplace_id != "custom":
        if match:
            match_stage = {"$match": match}
            total_sales_pipeline.append(match_stage)
            pipeline.append(match_stage)

        # Pipeline to calculate total units sold, total sold product count, and total sales
        total_sales_pipeline.extend([
            {
                "$group": {
                    "_id": None,  # Grouping by None to get a single summary document
                    "total_sales": {"$sum": "$order_total"},
                    "total_cogs": {"$sum": "$cogs"},  # Assuming 'cogs' field exists
                    "total_refunds": {"$sum": "$refunds"}  # Assuming 'refunds' field exists
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "total_sales": {"$round": ["$total_sales", 2]},
                    "total_cogs": {"$round": ["$total_cogs", 2]},
                    "total_refunds": {"$round": ["$total_refunds", 2]}
                }
            }
        ])

        summary = list(Order.objects.aggregate(*total_sales_pipeline))
        if summary:
            data['total_sales'] = summary[0].get('total_sales', 0)
            data['total_cogs'] = summary[0].get('total_cogs', 0)
            data['total_refunds'] = summary[0].get('total_refunds', 0)
            data['margin'] = ((data['total_sales'] - data['total_cogs']) / data['total_sales']) * 100 if data['total_sales'] else 0
        else:
            data['total_sales'] = 0
            data['total_cogs'] = 0
            data['total_refunds'] = 0
            data['margin'] = 0

        pipeline.extend([
            {
                "$unwind": "$order_items"  # Unwind the order_items list to process each item individually
            },
            {
                "$group": {
                    "_id": None,  # Grouping by None to get a single summary document
                    "ids": {"$addToSet": "$order_items"}  # Collect unique order_items
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "ids": 1
                }
            },
        ])
        summary1 = list(Order.objects.aggregate(*pipeline))
        if summary1:
            s_pipeline = [
                {"$match": {
                    "_id": {"$in": summary1[0]['ids']}
                }},
                {
                    "$group": {
                        "_id": None,
                        "total_units_sold": {"$sum": "$ProductDetails.QuantityOrdered"},
                        "unique_product_ids": {"$addToSet": "$ProductDetails.product_id"},
                    }
                },
                {
                    "$project": {
                        "_id": 0,
                        "total_units_sold": 1,
                        "total_sold_product_count": {"$size": "$unique_product_ids"},
                    }
                }
            ]
            summary2 = list(OrderItems.objects.aggregate(*s_pipeline))
            if summary2:
                data['total_units_sold'] = summary2[0].get('total_units_sold', 0)
                data['total_sold_product_count'] = summary2[0].get('total_sold_product_count', 0)
        else:
            data['total_units_sold'] = 0
            data['total_sold_product_count'] = 0

    # Add custom_order data
    custom_match = {}
    if start_date and end_date:
        custom_match["purchase_order_date"] = {"$gte": start_date, "$lte": end_date}

    custom_pipeline = [
        {"$match": custom_match},
        {"$unwind": "$ordered_products"},
        {
            "$group": {
                "_id": None,
                "total_custom_sales": {"$sum": "$ordered_products.quantity_price"},
                "total_custom_cogs": {"$sum": "$ordered_products.cogs"},  # Assuming 'cogs' field exists
                "total_custom_units_sold": {"$sum": "$ordered_products.quantity"},
                "unique_custom_product_ids": {"$addToSet": "$ordered_products.product_id"}
            }
        },
        {
            "$project": {
                "_id": 0,
                "total_custom_sales": {"$round": ["$total_custom_sales", 2]},
                "total_custom_cogs": {"$round": ["$total_custom_cogs", 2]},
                "total_custom_units_sold": 1,
                "total_custom_sold_product_count": {"$size": "$unique_custom_product_ids"}
            }
        }
    ]

    custom_summary = list(custom_order.objects.aggregate(*custom_pipeline))
    if custom_summary:
        total_custom_sales = custom_summary[0].get('total_custom_sales', 0)
        total_custom_cogs = custom_summary[0].get('total_custom_cogs', 0)
        total_custom_units_sold = custom_summary[0].get('total_custom_units_sold', 0)
        total_custom_sold_product_count = custom_summary[0].get('total_custom_sold_product_count', 0)
    else:
        total_custom_sales = 0
        total_custom_cogs = 0
        total_custom_units_sold = 0
        total_custom_sold_product_count = 0

    if marketplace_id == "custom":
        data['total_sales'] = total_custom_sales
        data['total_cogs'] = total_custom_cogs
        data['total_units_sold'] = total_custom_units_sold
        data['total_sold_product_count'] = total_custom_sold_product_count
        data['margin'] = ((total_custom_sales - total_custom_cogs) / total_custom_sales) * 100 if total_custom_sales else 0

    if marketplace_id == "": 
        # Combine totals
        data['total_sales'] += total_custom_sales
        data['total_cogs'] += total_custom_cogs
        data['total_units_sold'] += total_custom_units_sold
        data['total_sold_product_count'] += total_custom_sold_product_count
        data['margin'] = ((data['total_sales'] - data['total_cogs']) / data['total_sales']) * 100 if data['total_sales'] else 0

    return data


def getProductVariant(request):
    variant_list = list()
    product_id = request.GET.get('product_id')
    is_duplicate = request.GET.get('is_duplicate',False)
    parant_sku = DatabaseModel.get_document(Product.objects,{"id" : product_id},['parent_sku']).parent_sku
    match = {}
    match['parent_sku'] = parant_sku
    if is_duplicate == "true":
        match['_id'] = {"$ne" : ObjectId(product_id)}
    if parant_sku != None:
        pipeline = [
            {
            "$match": match
            },
            {
            "$lookup": {
                "from": "marketplace",
                "localField": "marketplace_ids",
                "foreignField": "_id",
                "as": "marketplace_ins"
            }
            },
            {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "product_title": {"$ifNull": ["$product_title", ""]},
                "product_id": {"$ifNull": ["$product_id", ""]},
                "product_id_type": {"$ifNull": ["$product_id_type", ""]},
                "sku": {"$ifNull": ["$sku", ""]},
                "price": {"$ifNull": ["$price", 0]},
                "currency": {"$ifNull": ["$currency", ""]},
                "quantity": {"$ifNull": ["$quantity", 0]},
                "marketplace_ins": {
                "$reduce": {
                    "input": "$marketplace_ins.name",
                    "initialValue": [],
                    "in": {
                    "$cond": {
                        "if": {"$in": ["$$this", "$$value"]},
                        "then": "$$value",
                        "else": {"$concatArrays": ["$$value", ["$$this"]]}
                    }
                    }
                }
                },
                "marketplace_image_url": {
                "$reduce": {
                    "input": "$marketplace_ins.image_url",
                    "initialValue": [],
                    "in": {
                    "$cond": {
                        "if": {"$in": ["$$this", "$$value"]},
                        "then": "$$value",
                        "else": {"$concatArrays": ["$$value", ["$$this"]]}
                    }
                    }
                }
                },
                "brand_name": {"$ifNull": ["$brand_name", ""]},
                "image_url": {"$ifNull": ["$image_url", ""]}
            }
            }
        ]
        variant_list = list(Product.objects.aggregate(*(pipeline)))
        
    return variant_list
